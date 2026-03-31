"""
Überstundenrechner Pro - Ein Tool zur Erfassung und Berechnung von Arbeitsstunden.
"""
import calendar
import csv
import json
import math
import os
import shutil
import sys
from datetime import datetime

from PyQt6.QtCore import QDate, Qt, QTime, QPoint
from PyQt6.QtGui import QColor, QFont, QIcon, QPalette, QPainter, QPixmap, QPolygon
from PyQt6.QtWidgets import (
    QApplication, QCheckBox, QComboBox, QDateEdit,
    QFileDialog, QFrame, QHBoxLayout, QHeaderView, QLabel,
    QLineEdit, QMainWindow, QMenu, QMessageBox, QPushButton, QSpinBox,
    QTableWidget, QTableWidgetItem, QTabWidget, QTimeEdit,
    QVBoxLayout, QWidget, QGroupBox, QProgressBar, QGridLayout
)

# Configuration must be imported before matplotlib for MPLCONFIGDIR
# pylint: disable=wrong-import-position, wrong-import-order
from config import BASE_DIR, DB_FILE, SETTINGS_FILE, ICON_PATH
import matplotlib
matplotlib.use('QtAgg')
from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg
from matplotlib.figure import Figure

from models import WorkEntry
from database import DBManager
from logic import get_holidays, calculate_timed_entries, get_login_time
from ui_components import HeatmapDelegate
from dialogs import SettingsDialog, EditDialog
# pylint: enable=wrong-import-position, wrong-import-order

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    _OPENPYXL = True
except ImportError:
    _OPENPYXL = False


# --- HAUPTANWENDUNG ---
class UeberstundenApp(QMainWindow):
    """
    Hauptklasse der Anwendung Überstunden-Rechner Pro.
    Verwaltet das Hauptfenster, die Tabs und die Geschäftslogik.
    """
    def __init__(self):
        """
        Initialisiert die Anwendung, lädt Einstellungen und baut die UI auf.
        """
        super().__init__()
        self.setWindowTitle("Überstunden-Rechner Pro")
        self.resize(1000, 750)

        if os.path.exists(ICON_PATH):
            self.setWindowIcon(QIcon(ICON_PATH))

        self.system_palette = QApplication.instance().palette()
        bg_color = self.system_palette.color(QPalette.ColorRole.Window)
        bg_lightness = bg_color.lightness()
        self.system_is_dark = bg_lightness < 128

        self.settings = self.load_settings()
        self.db = DBManager(self.settings.get("db_path", DB_FILE))

        self.entries = []
        self.current_calculated_overtime = 0
        self.current_calculated_pause = 0

        self.apply_theme()

        # Daten laden BEVOR die UI sie berechnet
        self.entries = self.db.load_all()

        self.tabs = QTabWidget()
        self.setCentralWidget(self.tabs)

        self.tab_main = QWidget()
        self.tab_goals = QWidget()
        self.tab_calendar = QWidget()
        self.tab_stats = QWidget()

        self.tabs.addTab(self.tab_main, "Eingabe && Liste")
        self.tabs.addTab(self.tab_goals, "Ziele && Dashboard")
        self.tabs.addTab(self.tab_calendar, "Kalender-Heatmap")
        self.tabs.addTab(self.tab_stats, "Diagramm && Statistik")

        self.setup_main_tab()
        self.setup_goals_tab()
        self.setup_calendar_tab()
        self.setup_stats_tab()

        self.load_data()

    def load_settings(self):
        """
        Lädt die Einstellungen aus der JSON-Datei oder gibt Standardwerte zurück.
        """
        defaults = {
            "default_start": "07:00",
            "target_work_time": "08:00",
            "state": "TH",
            "dark_mode": self.system_is_dark,
            "auto_break": True,
            "use_login_time": False,
            "workdays": [0, 1, 2, 3, 4],
            "special_days": [
                {"month": 12, "day": 24, "target": "04:00"},
                {"month": 12, "day": 31, "target": "04:00"}
            ],
            "goal_active": False,
            "goal_start_date": QDate.currentDate().addDays(30).toString("yyyy-MM-dd"),
            "goal_end_date": QDate.currentDate().addDays(35).toString("yyyy-MM-dd"),
            "goal_hours": 0,
            "db_path": DB_FILE
        }
        if os.path.exists(SETTINGS_FILE):
            try:
                with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
                    defaults.update(json.load(f))
            except (json.JSONDecodeError, OSError):
                pass
        return defaults

    def save_settings(self):
        """
        Speichert die aktuellen Einstellungen in die JSON-Datei.
        """
        try:
            with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
                json.dump(self.settings, f)
        except OSError:
            pass

    def get_light_palette(self):
        """
        Erstellt und gibt die Farbpalette für den hellen Modus zurück (Breeze Light).
        """
        palette = QPalette()
        palette.setColor(QPalette.ColorRole.Window,          QColor("#eff0f1"))
        palette.setColor(QPalette.ColorRole.WindowText,      QColor("#31363b"))
        palette.setColor(QPalette.ColorRole.Base,            QColor("#fcfcfc"))
        palette.setColor(QPalette.ColorRole.AlternateBase,   QColor("#eff0f1"))
        palette.setColor(QPalette.ColorRole.ToolTipBase,     QColor("#31363b"))
        palette.setColor(QPalette.ColorRole.ToolTipText,     QColor("#eff0f1"))
        palette.setColor(QPalette.ColorRole.Text,            QColor("#31363b"))
        palette.setColor(QPalette.ColorRole.Button,          QColor("#eff0f1"))
        palette.setColor(QPalette.ColorRole.ButtonText,      QColor("#31363b"))
        palette.setColor(QPalette.ColorRole.BrightText,      QColor("#ff0000"))
        palette.setColor(QPalette.ColorRole.Highlight,       QColor("#3daee9"))
        palette.setColor(QPalette.ColorRole.HighlightedText, QColor("#ffffff"))
        palette.setColor(QPalette.ColorRole.Link,            QColor("#3daee9"))
        return palette

    def get_dark_palette(self):
        """
        Erstellt und gibt die Farbpalette für den dunklen Modus zurück (Breeze Dark).
        """
        palette = QPalette()
        palette.setColor(QPalette.ColorRole.Window,          QColor("#31363b"))
        palette.setColor(QPalette.ColorRole.WindowText,      QColor("#eff0f1"))
        palette.setColor(QPalette.ColorRole.Base,            QColor("#232629"))
        palette.setColor(QPalette.ColorRole.AlternateBase,   QColor("#31363b"))
        palette.setColor(QPalette.ColorRole.ToolTipBase,     QColor("#31363b"))
        palette.setColor(QPalette.ColorRole.ToolTipText,     QColor("#eff0f1"))
        palette.setColor(QPalette.ColorRole.Text,            QColor("#eff0f1"))
        palette.setColor(QPalette.ColorRole.Button,          QColor("#31363b"))
        palette.setColor(QPalette.ColorRole.ButtonText,      QColor("#eff0f1"))
        palette.setColor(QPalette.ColorRole.BrightText,      QColor("#ff5555"))
        palette.setColor(QPalette.ColorRole.Highlight,       QColor("#3daee9"))
        palette.setColor(QPalette.ColorRole.HighlightedText, QColor("#ffffff"))
        palette.setColor(QPalette.ColorRole.Link,            QColor("#3daee9"))
        return palette

    def get_dark_stylesheet(self, icon_dir=''):
        """
        Gibt das CSS-Stylesheet für den Breeze Dark Modus zurück.
        """
        return """
            * { font-size: 13px; }

            QMainWindow, QDialog { background-color: #31363b; }

            QWidget { background-color: #31363b; color: #eff0f1; }

            QPushButton {
                background-color: #3b4045;
                color: #eff0f1;
                border: 1px solid #4a5056;
                border-radius: 3px;
                padding: 4px 16px;
                min-height: 24px;
            }
            QPushButton:hover  { background-color: #4d5560; border-color: #3daee9; }
            QPushButton:pressed { background-color: #3daee9; color: #fff; border-color: #3daee9; }
            QPushButton:disabled { color: #6c7176; border-color: #3a3f44; }

            QLineEdit, QComboBox {
                background-color: #232629;
                color: #eff0f1;
                border: 1px solid #4a5056;
                border-radius: 3px;
                padding: 3px 6px;
                min-height: 24px;
                selection-background-color: #3daee9;
            }
            QLineEdit:focus, QComboBox:focus { border-color: #3daee9; }

            QAbstractSpinBox {
                background-color: #232629;
                color: #eff0f1;
                border: 1px solid #4a5056;
                border-radius: 3px;
                padding: 3px 6px;
                min-height: 24px;
            }
            QAbstractSpinBox:focus { border-color: #3daee9; }
            QAbstractSpinBox::up-button {
                subcontrol-origin: border;
                subcontrol-position: top right;
                width: 20px;
                background-color: #3b4045;
                border-left: 1px solid #4a5056;
                border-bottom: 1px solid #4a5056;
                border-top-right-radius: 3px;
            }
            QAbstractSpinBox::down-button {
                subcontrol-origin: border;
                subcontrol-position: bottom right;
                width: 20px;
                background-color: #3b4045;
                border-left: 1px solid #4a5056;
                border-bottom-right-radius: 3px;
            }
            QAbstractSpinBox::up-button:hover,
            QAbstractSpinBox::down-button:hover { background-color: #3daee9; }

            QComboBox::drop-down { border: none; width: 22px; }
            QComboBox::down-arrow { width: 10px; height: 10px; }
            QComboBox QAbstractItemView {
                background-color: #232629;
                color: #eff0f1;
                border: 1px solid #4a5056;
                selection-background-color: #3daee9;
                selection-color: #fff;
                outline: none;
            }

            QTabWidget::pane { border: 1px solid #4a5056; top: -1px; }
            QTabBar::tab {
                background-color: #2e3338;
                color: #7f8c8d;
                border: 1px solid #4a5056;
                border-bottom: none;
                border-top-left-radius: 3px;
                border-top-right-radius: 3px;
                padding: 5px 14px;
                margin-right: 2px;
            }
            QTabBar::tab:selected    { background-color: #31363b; color: #eff0f1; border-bottom: 2px solid #3daee9; }
            QTabBar::tab:hover:!selected { background-color: #3b4045; color: #eff0f1; }

            QTableWidget {
                background-color: #232629;
                alternate-background-color: #2a2f34;
                color: #eff0f1;
                gridline-color: #3a3f44;
                border: 1px solid #4a5056;
            }
            QTableWidget::item:selected { background-color: #3daee9; color: #fff; }
            QHeaderView::section {
                background-color: #31363b;
                color: #eff0f1;
                border: none;
                border-right: 1px solid #4a5056;
                border-bottom: 1px solid #4a5056;
                padding: 4px 8px;
                font-weight: bold;
            }
            QHeaderView::section:first { border-left: none; }

            QGroupBox {
                border: 1px solid #4a5056;
                border-radius: 3px;
                margin-top: 12px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top left;
                left: 8px;
                color: #3daee9;
                font-weight: bold;
            }

            QCheckBox::indicator {
                width: 16px; height: 16px;
                border: 1px solid #4a5056;
                border-radius: 2px;
                background-color: #232629;
            }
            QCheckBox::indicator:checked  { background-color: #3daee9; border-color: #3daee9; }
            QCheckBox::indicator:hover    { border-color: #3daee9; }

            QScrollBar:vertical { background: #2e3338; width: 8px; border-radius: 4px; margin: 0; }
            QScrollBar::handle:vertical { background: #4a5056; border-radius: 4px; min-height: 24px; }
            QScrollBar::handle:vertical:hover { background: #3daee9; }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height: 0; }
            QScrollBar:horizontal { background: #2e3338; height: 8px; border-radius: 4px; margin: 0; }
            QScrollBar::handle:horizontal { background: #4a5056; border-radius: 4px; min-width: 24px; }
            QScrollBar::handle:horizontal:hover { background: #3daee9; }
            QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal { width: 0; }

            QProgressBar {
                border: 1px solid #4a5056;
                border-radius: 3px;
                background-color: #232629;
                text-align: center;
                color: #eff0f1;
                min-height: 16px;
            }
            QProgressBar::chunk { background-color: #3daee9; border-radius: 2px; }

            QLabel { background: transparent; }
            QFrame { background: transparent; }
            QToolTip { background-color: #31363b; color: #eff0f1; border: 1px solid #4a5056; }
            QMessageBox { background-color: #31363b; }

            /* Aufklappbarer Kalender (QDateEdit-Popup) */
            QCalendarWidget {
                background-color: #31363b;
                color: #eff0f1;
            }
            QCalendarWidget QWidget#qt_calendar_navigationbar {
                background-color: #232629;
                padding: 2px;
            }
            QCalendarWidget QToolButton {
                background-color: #3b4045;
                color: #eff0f1;
                border: 1px solid #4a5056;
                border-radius: 3px;
                padding: 3px 8px;
                margin: 2px;
            }
            QCalendarWidget QToolButton:hover  { background-color: #3daee9; border-color: #3daee9; }
            QCalendarWidget QToolButton:pressed { background-color: #2980b9; }
            QCalendarWidget QToolButton::menu-indicator { image: none; }
            QCalendarWidget QSpinBox {
                background-color: #232629;
                color: #eff0f1;
                border: 1px solid #4a5056;
                border-radius: 3px;
                padding: 2px 4px;
                selection-background-color: #3daee9;
            }
            QCalendarWidget QAbstractItemView:enabled {
                background-color: #232629;
                color: #eff0f1;
                selection-background-color: #3daee9;
                selection-color: #ffffff;
                gridline-color: #3a3f44;
            }
            QCalendarWidget QAbstractItemView:disabled { color: #4a5056; }
            QCalendarWidget QMenu {
                background-color: #232629;
                color: #eff0f1;
                border: 1px solid #4a5056;
            }
            QCalendarWidget QMenu::item:selected { background-color: #3daee9; color: #fff; }
        """ + (
            f"            QAbstractSpinBox::up-arrow   "
            f"{{ image: url(\"{icon_dir}/arrow_up.png\");   width: 10px; height: 10px; }}\n"
            f"            QAbstractSpinBox::down-arrow "
            f"{{ image: url(\"{icon_dir}/arrow_down.png\"); width: 10px; height: 10px; }}\n"
            if icon_dir else ""
        )

    def get_light_stylesheet(self, icon_dir=''):
        """Gibt das CSS-Stylesheet für den Breeze Light Modus zurück."""
        return """
            * { font-size: 13px; }

            QMainWindow, QDialog { background-color: #eff0f1; }

            QWidget { background-color: #eff0f1; color: #31363b; }

            QPushButton {
                background-color: #e3e5e7;
                color: #31363b;
                border: 1px solid #bdc3c7;
                border-radius: 3px;
                padding: 4px 16px;
                min-height: 24px;
            }
            QPushButton:hover  { background-color: #d6eaf8; border-color: #3daee9; }
            QPushButton:pressed { background-color: #3daee9; color: #fff; border-color: #3daee9; }
            QPushButton:disabled { color: #95a5a6; border-color: #bdc3c7; }

            QLineEdit, QComboBox {
                background-color: #fcfcfc;
                color: #31363b;
                border: 1px solid #bdc3c7;
                border-radius: 3px;
                padding: 3px 6px;
                min-height: 24px;
                selection-background-color: #3daee9;
                selection-color: #fff;
            }
            QLineEdit:focus, QComboBox:focus { border-color: #3daee9; }

            QAbstractSpinBox {
                background-color: #fcfcfc;
                color: #31363b;
                border: 1px solid #bdc3c7;
                border-radius: 3px;
                padding: 3px 6px;
                min-height: 24px;
            }
            QAbstractSpinBox:focus { border-color: #3daee9; }
            QAbstractSpinBox::up-button {
                subcontrol-origin: border;
                subcontrol-position: top right;
                width: 20px;
                background-color: #e3e5e7;
                border-left: 1px solid #bdc3c7;
                border-bottom: 1px solid #bdc3c7;
                border-top-right-radius: 3px;
            }
            QAbstractSpinBox::down-button {
                subcontrol-origin: border;
                subcontrol-position: bottom right;
                width: 20px;
                background-color: #e3e5e7;
                border-left: 1px solid #bdc3c7;
                border-bottom-right-radius: 3px;
            }
            QAbstractSpinBox::up-button:hover,
            QAbstractSpinBox::down-button:hover { background-color: #3daee9; }

            QComboBox::drop-down { border: none; width: 22px; }
            QComboBox::down-arrow { width: 10px; height: 10px; }
            QComboBox QAbstractItemView {
                background-color: #fcfcfc;
                color: #31363b;
                border: 1px solid #bdc3c7;
                selection-background-color: #3daee9;
                selection-color: #fff;
                outline: none;
            }

            QTabWidget::pane { border: 1px solid #bdc3c7; top: -1px; }
            QTabBar::tab {
                background-color: #dde1e3;
                color: #7f8c8d;
                border: 1px solid #bdc3c7;
                border-bottom: none;
                border-top-left-radius: 3px;
                border-top-right-radius: 3px;
                padding: 5px 14px;
                margin-right: 2px;
            }
            QTabBar::tab:selected    { background-color: #eff0f1; color: #31363b; border-bottom: 2px solid #3daee9; }
            QTabBar::tab:hover:!selected { background-color: #e8eaeb; color: #31363b; }

            QTableWidget {
                background-color: #fcfcfc;
                alternate-background-color: #f4f5f5;
                color: #31363b;
                gridline-color: #d5d8db;
                border: 1px solid #bdc3c7;
            }
            QTableWidget::item:selected { background-color: #3daee9; color: #fff; }
            QHeaderView::section {
                background-color: #e8eaeb;
                color: #31363b;
                border: none;
                border-right: 1px solid #bdc3c7;
                border-bottom: 1px solid #bdc3c7;
                padding: 4px 8px;
                font-weight: bold;
            }

            QGroupBox {
                border: 1px solid #bdc3c7;
                border-radius: 3px;
                margin-top: 12px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top left;
                left: 8px;
                color: #3daee9;
                font-weight: bold;
            }

            QCheckBox::indicator {
                width: 16px; height: 16px;
                border: 1px solid #bdc3c7;
                border-radius: 2px;
                background-color: #fcfcfc;
            }
            QCheckBox::indicator:checked { background-color: #3daee9; border-color: #3daee9; }
            QCheckBox::indicator:hover   { border-color: #3daee9; }

            QScrollBar:vertical { background: #dde1e3; width: 8px; border-radius: 4px; margin: 0; }
            QScrollBar::handle:vertical { background: #bdc3c7; border-radius: 4px; min-height: 24px; }
            QScrollBar::handle:vertical:hover { background: #3daee9; }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height: 0; }
            QScrollBar:horizontal { background: #dde1e3; height: 8px; border-radius: 4px; margin: 0; }
            QScrollBar::handle:horizontal { background: #bdc3c7; border-radius: 4px; min-width: 24px; }
            QScrollBar::handle:horizontal:hover { background: #3daee9; }
            QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal { width: 0; }

            QProgressBar {
                border: 1px solid #bdc3c7;
                border-radius: 3px;
                background-color: #dde1e3;
                text-align: center;
                color: #31363b;
                min-height: 16px;
            }
            QProgressBar::chunk { background-color: #3daee9; border-radius: 2px; }

            QLabel { background: transparent; }
            QFrame { background: transparent; }
            QToolTip { background-color: #31363b; color: #eff0f1; border: 1px solid #4a5056; }

            /* Aufklappbarer Kalender (QDateEdit-Popup) */
            QCalendarWidget {
                background-color: #eff0f1;
                color: #31363b;
            }
            QCalendarWidget QWidget#qt_calendar_navigationbar {
                background-color: #dde1e3;
                padding: 2px;
            }
            QCalendarWidget QToolButton {
                background-color: #e3e5e7;
                color: #31363b;
                border: 1px solid #bdc3c7;
                border-radius: 3px;
                padding: 3px 8px;
                margin: 2px;
            }
            QCalendarWidget QToolButton:hover  { background-color: #3daee9; color: #fff; border-color: #3daee9; }
            QCalendarWidget QToolButton:pressed { background-color: #2980b9; color: #fff; }
            QCalendarWidget QToolButton::menu-indicator { image: none; }
            QCalendarWidget QSpinBox {
                background-color: #fcfcfc;
                color: #31363b;
                border: 1px solid #bdc3c7;
                border-radius: 3px;
                padding: 2px 4px;
                selection-background-color: #3daee9;
                selection-color: #fff;
            }
            QCalendarWidget QAbstractItemView:enabled {
                background-color: #fcfcfc;
                color: #31363b;
                selection-background-color: #3daee9;
                selection-color: #ffffff;
                gridline-color: #d5d8db;
            }
            QCalendarWidget QAbstractItemView:disabled { color: #95a5a6; }
            QCalendarWidget QMenu {
                background-color: #fcfcfc;
                color: #31363b;
                border: 1px solid #bdc3c7;
            }
            QCalendarWidget QMenu::item:selected { background-color: #3daee9; color: #fff; }
        """ + (
            f"            QAbstractSpinBox::up-arrow   "
            f"{{ image: url(\"{icon_dir}/arrow_up.png\");   width: 10px; height: 10px; }}\n"
            f"            QAbstractSpinBox::down-arrow "
            f"{{ image: url(\"{icon_dir}/arrow_down.png\"); width: 10px; height: 10px; }}\n"
            if icon_dir else ""
        )

    def _create_arrow_icons(self):
        """Zeichnet kleine Dreieck-Pfeile als PNG und speichert sie für die QSS-Nutzung."""
        icon_dir = os.path.join(BASE_DIR, '.ui_icons')
        os.makedirs(icon_dir, exist_ok=True)
        is_dark = self.settings.get("dark_mode", False)
        color = QColor("#eff0f1") if is_dark else QColor("#31363b")
        arrows = {
            'up':   [QPoint(5, 1), QPoint(10, 9), QPoint(0, 9)],
            'down': [QPoint(0, 1), QPoint(10, 1), QPoint(5, 9)],
        }
        for name, pts in arrows.items():
            px = QPixmap(10, 10)
            px.fill(Qt.GlobalColor.transparent)
            p = QPainter(px)
            p.setRenderHint(QPainter.RenderHint.Antialiasing)
            p.setBrush(color)
            p.setPen(Qt.PenStyle.NoPen)
            p.drawPolygon(QPolygon(pts))
            p.end()
            px.save(os.path.join(icon_dir, f'arrow_{name}.png'))
        # Qt QSS benötigt Forward-Slashes, auch auf Windows
        return icon_dir.replace('\\', '/')

    def apply_theme(self):
        """Wendet das aktuelle Theme (Hell/Dunkel) auf die gesamte Anwendung an."""
        qt_app = QApplication.instance()
        is_dark = self.settings.get("dark_mode", False)

        # Auf Linux im Entwicklungsmodus: natives Breeze-Theme nutzen
        if not getattr(sys, 'frozen', False) and sys.platform.startswith('linux'):
            qt_app.setStyle("Breeze")
            qt_app.setPalette(self.get_dark_palette() if is_dark else self.get_light_palette())
            return

        # Alle anderen Fälle (kompiliert oder Windows/macOS): Fusion + modernes Stylesheet
        qt_app.setStyle("Fusion")
        qt_app.setPalette(self.get_dark_palette() if is_dark else self.get_light_palette())
        icon_dir = self._create_arrow_icons()
        sheet = (
            self.get_dark_stylesheet(icon_dir) if is_dark else self.get_light_stylesheet(icon_dir)
        )
        qt_app.setStyleSheet(sheet)

    def closeEvent(self, event):  # pylint: disable=invalid-name
        """
        Wird beim Schließen der Anwendung aufgerufen. Speichert Einstellungen und schließt die DB.
        """
        self.save_settings()
        self.db.close()
        super().closeEvent(event)

    def load_data(self):
        """
        Lädt alle Daten aus der Datenbank und aktualisiert die Benutzeroberfläche.
        """
        self.entries = self.db.load_all()
        self.update_ui()
        self.update_stats_chart()
        self.on_goal_changed()
        self.update_calendar_heatmap()

    def get_target_minutes(self):
        """
        Gibt die in den Einstellungen festgelegte Regelarbeitszeit in Minuten zurück.
        """
        t = QTime.fromString(self.settings.get("target_work_time", "08:00"), "HH:mm")
        return t.hour() * 60 + t.minute()

    def get_target_minutes_for_date(self, date_str):
        """
        Ermittelt das Tagessoll für ein bestimmtes Datum unter Berücksichtigung von
        individuellen Einträgen, Sonderarbeitstagen, Feiertagen und Wochenenden.
        """
        # 1. Check if any entry for this day has a custom target_minutes
        for e in self.entries:
            if e.date == date_str and e.target_minutes != -1:
                return e.target_minutes

        qdate = QDate.fromString(date_str, "yyyy-MM-dd")

        # 2. Check special days from settings (e.g. 24.12.)
        special_days = self.settings.get("special_days", [])
        for sd in special_days:
            if qdate.month() == sd["month"] and qdate.day() == sd["day"]:
                t = QTime.fromString(sd["target"], "HH:mm")
                return t.hour() * 60 + t.minute()

        year = qdate.year()
        state = self.settings.get("state", "TH")
        holidays = get_holidays(year, state)

        # Check if holiday
        if date_str in holidays:
            return 0

        # Check if workday (0=Mon, 6=Sun)
        # QDate.dayOfWeek() returns 1 (Mon) to 7 (Sun)
        day_idx = qdate.dayOfWeek() - 1
        workdays = self.settings.get("workdays", [0, 1, 2, 3, 4])
        if day_idx not in workdays:
            return 0

        return self.get_target_minutes()

    def get_max_minutes(self):
        """
        Gibt die maximal anrechenbare Arbeitszeit pro Tag in Minuten zurück.
        """
        return self.settings.get("max_work_hours", 10) * 60

    def _get_default_start_time(self):
        """Ermittelt die Startzeit für die Eingabemaske.
        Wenn Login-Zeit aktiv: Login-Zeit → default_start (Fallback).
        Sonst: last_start (heute) → default_start."""
        if self.settings.get("use_login_time", False):
            t = get_login_time()
            if t and t.isValid():
                return t
            return QTime.fromString(self.settings.get("default_start", "07:00"), "HH:mm")
        today = QDate.currentDate().toString("yyyy-MM-dd")
        if self.settings.get("last_date") == today and self.settings.get("last_start"):
            return QTime.fromString(self.settings["last_start"], "HH:mm")
        return QTime.fromString(self.settings.get("default_start", "07:00"), "HH:mm")

    # --- SETUP TABS ---
    def setup_main_tab(self):
        """
        Erstellt das Layout und die Steuerelemente für den Haupt-Tab (Eingabe & Liste).
        """
        layout = QVBoxLayout(self.tab_main)

        self.lbl_saldo = QLabel("0h 0m")
        self.lbl_saldo.setAlignment(Qt.AlignmentFlag.AlignCenter)
        font = QFont()
        font.setPointSize(28)
        font.setBold(True)
        self.lbl_saldo.setFont(font)
        layout.addWidget(QLabel("<b>Gesamt-Saldo:</b>",
                                alignment=Qt.AlignmentFlag.AlignCenter))
        layout.addWidget(self.lbl_saldo)

        frame_input = QFrame()
        frame_layout = QVBoxLayout(frame_input)

        input_row1 = QHBoxLayout()
        self.date_edit = QDateEdit()
        self.date_edit.setCalendarPopup(True)
        self.date_edit.setDate(QDate.currentDate())
        self.date_edit.dateChanged.connect(self.update_live_calc)
        input_row1.addWidget(QLabel("Datum:"))
        input_row1.addWidget(self.date_edit)

        start_to_use = self._get_default_start_time()

        self.time_start = QTimeEdit()
        self.time_start.setDisplayFormat("HH:mm")
        self.time_start.setTime(start_to_use)
        input_row1.addWidget(QLabel("Start:"))
        input_row1.addWidget(self.time_start)
        btn_now_start = QPushButton("Jetzt")
        btn_now_start.setFixedWidth(50)
        btn_now_start.setToolTip("Aktuelle Uhrzeit als Startzeit setzen")
        btn_now_start.clicked.connect(lambda: self.time_start.setTime(QTime.currentTime()))
        input_row1.addWidget(btn_now_start)

        self.time_end = QTimeEdit()
        self.time_end.setDisplayFormat("HH:mm")
        input_row1.addWidget(QLabel("Ende:"))
        input_row1.addWidget(self.time_end)
        btn_now_end = QPushButton("Jetzt")
        btn_now_end.setFixedWidth(50)
        btn_now_end.setToolTip("Aktuelle Uhrzeit als Endzeit setzen")
        btn_now_end.clicked.connect(self.set_now_as_end)
        input_row1.addWidget(btn_now_end)

        self.pause_spin = QSpinBox()
        self.pause_spin.setRange(0, 300)
        self.pause_spin.setSuffix(" Min")
        self.pause_spin.setEnabled(not self.settings.get("auto_break", True))
        input_row1.addWidget(QLabel("Pause:"))
        input_row1.addWidget(self.pause_spin)

        self.reason_edit = QLineEdit()
        self.reason_edit.setPlaceholderText("z.B. Regulär")
        input_row1.addWidget(QLabel("Anlass:"))
        input_row1.addWidget(self.reason_edit)

        btn_add = QPushButton("Eintragen")
        btn_add.clicked.connect(self.add_entry)
        input_row1.addWidget(btn_add)

        frame_layout.addLayout(input_row1)

        input_row2 = QHBoxLayout()
        self.custom_target_cb = QCheckBox("Indiv. Tagessoll:")
        self.custom_target_time = QTimeEdit()
        self.custom_target_time.setDisplayFormat("HH:mm")
        def_target = self.settings.get("target_work_time", "08:00")
        self.custom_target_time.setTime(QTime.fromString(def_target, "HH:mm"))
        self.custom_target_time.setEnabled(False)
        self.custom_target_cb.stateChanged.connect(
            lambda: self.custom_target_time.setEnabled(self.custom_target_cb.isChecked())
        )
        self.custom_target_cb.stateChanged.connect(self.update_live_calc)
        self.custom_target_time.timeChanged.connect(self.update_live_calc)

        input_row2.addWidget(self.custom_target_cb)
        input_row2.addWidget(self.custom_target_time)
        input_row2.addStretch()
        frame_layout.addLayout(input_row2)

        self.lbl_live_calc = QLabel("Berechne...")
        self.lbl_live_calc.setAlignment(Qt.AlignmentFlag.AlignCenter)
        frame_layout.addWidget(self.lbl_live_calc)
        layout.addWidget(frame_input)

        self.time_start.timeChanged.connect(self.on_start_time_changed)
        self.time_end.timeChanged.connect(self.update_live_calc)
        self.pause_spin.valueChanged.connect(self.update_live_calc)

        toolbar_layout = QHBoxLayout()
        self.month_filter = QComboBox()
        self.month_filter.addItem("Alle", "ALL")
        self.month_filter.currentIndexChanged.connect(self.on_list_filter_changed)
        toolbar_layout.addWidget(QLabel("Filter:"))
        toolbar_layout.addWidget(self.month_filter)
        toolbar_layout.addStretch()

        btn_import = QPushButton("CSV Import")
        btn_import.clicked.connect(self.import_csv)
        toolbar_layout.addWidget(btn_import)

        btn_export = QPushButton("Export")
        export_menu = QMenu(self)
        export_menu.addAction("CSV  (.csv)",  self.export_csv)
        export_menu.addAction("Excel (.xlsx)", self.export_xlsx)
        export_menu.addAction("PDF  (.pdf)",  self.export_pdf)
        btn_export.setMenu(export_menu)
        toolbar_layout.addWidget(btn_export)

        btn_settings = QPushButton("Einstellungen")
        btn_settings.clicked.connect(self.open_settings)
        toolbar_layout.addWidget(btn_settings)
        layout.addLayout(toolbar_layout)

        self.table = QTableWidget(0, 5)
        self.table.setHorizontalHeaderLabels(
            ["Datum", "Zeitraum", "Überstunden", "Anlass", "Aktion"]
        )
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        self.table.setColumnWidth(1, 160)

        self.table.cellDoubleClicked.connect(self.edit_entry)
        layout.addWidget(self.table)

        self.on_start_time_changed(start_to_use)

    def setup_goals_tab(self):
        """Erstellt den Tab für Ziele und das Fortschritts-Dashboard."""
        layout = QVBoxLayout(self.tab_goals)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(20)

        # 1. Konfigurations-Bereich
        settings_group = QGroupBox("Zeitraum und Überstunden-Ziel konfigurieren")
        settings_layout = QVBoxLayout(settings_group)

        goal_header_layout = QHBoxLayout()
        self.goal_active_cb = QCheckBox("Gleitzeit-Ziel aktivieren (Urlaubs-Sparer)")
        self.goal_active_cb.setChecked(self.settings.get("goal_active", False))
        self.goal_active_cb.stateChanged.connect(self.on_goal_changed)
        goal_header_layout.addWidget(self.goal_active_cb)
        goal_header_layout.addStretch()
        settings_layout.addLayout(goal_header_layout)

        goal_inputs_layout = QHBoxLayout()

        goal_inputs_layout.addWidget(QLabel("Urlaub / Frei von:"))
        self.goal_start_edit = QDateEdit()
        self.goal_start_edit.setCalendarPopup(True)
        self.goal_start_edit.setDate(
            QDate.fromString(self.settings.get("goal_start_date", ""), "yyyy-MM-dd")
        )
        self.goal_start_edit.dateChanged.connect(self.auto_calculate_goal_hours)
        goal_inputs_layout.addWidget(self.goal_start_edit)

        goal_inputs_layout.addWidget(QLabel("bis:"))
        self.goal_end_edit = QDateEdit()
        self.goal_end_edit.setCalendarPopup(True)
        self.goal_end_edit.setDate(
            QDate.fromString(self.settings.get("goal_end_date", ""), "yyyy-MM-dd")
        )
        self.goal_end_edit.dateChanged.connect(self.auto_calculate_goal_hours)
        goal_inputs_layout.addWidget(self.goal_end_edit)

        goal_inputs_layout.addWidget(QLabel(" | Benötigte Überstunden:"))
        self.goal_hours_spin = QSpinBox()
        self.goal_hours_spin.setRange(0, 5000)
        self.goal_hours_spin.setValue(self.settings.get("goal_hours", 0))
        self.goal_hours_spin.setSuffix(" h")
        self.goal_hours_spin.valueChanged.connect(self.on_goal_changed)
        goal_inputs_layout.addWidget(self.goal_hours_spin)

        goal_inputs_layout.addStretch()
        settings_layout.addLayout(goal_inputs_layout)

        layout.addWidget(settings_group)

        # 2. Dashboard-Bereich
        self.dashboard_group = QGroupBox("Fortschritts-Dashboard")
        dashboard_layout = QVBoxLayout(self.dashboard_group)

        self.goal_progress_bar = QProgressBar()
        self.goal_progress_bar.setRange(0, 100)
        self.goal_progress_bar.setValue(0)
        self.goal_progress_bar.setFixedHeight(30)
        self.goal_progress_bar.setTextVisible(True)
        self.goal_progress_bar.setAlignment(Qt.AlignmentFlag.AlignCenter)
        dashboard_layout.addWidget(self.goal_progress_bar)

        grid = QGridLayout()
        grid.setSpacing(15)

        # Kachel: Aktueller Stand
        self.lbl_goal_current = QLabel("0h 0m")
        self.lbl_goal_current.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_goal_current.setFont(QFont("Arial", 20, QFont.Weight.Bold))
        grid.addWidget(QLabel("Aktueller Stand", alignment=Qt.AlignmentFlag.AlignCenter), 0, 0)
        grid.addWidget(self.lbl_goal_current, 1, 0)

        # Kachel: Es fehlen noch
        self.lbl_goal_missing = QLabel("0h 0m")
        self.lbl_goal_missing.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_goal_missing.setFont(QFont("Arial", 20, QFont.Weight.Bold))
        grid.addWidget(QLabel("Es fehlen noch", alignment=Qt.AlignmentFlag.AlignCenter), 0, 1)
        grid.addWidget(self.lbl_goal_missing, 1, 1)

        # Kachel: Verbleibende Tage (bis zum Start des Zeitraums)
        self.lbl_goal_days = QLabel("0")
        self.lbl_goal_days.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_goal_days.setFont(QFont("Arial", 20, QFont.Weight.Bold))
        grid.addWidget(
            QLabel("Arbeitstage zum Ansparen", alignment=Qt.AlignmentFlag.AlignCenter), 0, 2
        )
        grid.addWidget(self.lbl_goal_days, 1, 2)

        dashboard_layout.addLayout(grid)

        # Fazit-Label unten
        self.lbl_goal_action = QLabel("-")
        self.lbl_goal_action.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_goal_action.setFont(QFont("Arial", 12))
        dashboard_layout.addSpacing(15)
        dashboard_layout.addWidget(self.lbl_goal_action)

        layout.addWidget(self.dashboard_group)
        layout.addStretch()

        if self.settings.get("goal_hours", 0) == 0:
            self.auto_calculate_goal_hours()

    def setup_calendar_tab(self):
        """Erstellt den Tab für die Kalender-Heatmap."""
        layout = QVBoxLayout(self.tab_calendar)

        cal_toolbar = QHBoxLayout()

        self.btn_cal_prev = QPushButton("< Vorheriger")
        self.btn_cal_prev.clicked.connect(self.cal_go_prev_month)
        self.btn_cal_next = QPushButton("Nächster >")
        self.btn_cal_next.clicked.connect(self.cal_go_next_month)

        self.cal_month_filter = QComboBox()
        self.cal_month_filter.currentIndexChanged.connect(self.on_cal_filter_changed)

        cal_toolbar.addWidget(QLabel("Monat:"))
        cal_toolbar.addWidget(self.btn_cal_prev)
        cal_toolbar.addWidget(self.cal_month_filter)
        cal_toolbar.addWidget(self.btn_cal_next)

        cal_toolbar.addStretch()

        self.lbl_cal_month_sum = QLabel("Monats-Saldo: 0h 0m")
        font = QFont()
        font.setPointSize(12)
        font.setBold(True)
        self.lbl_cal_month_sum.setFont(font)
        cal_toolbar.addWidget(self.lbl_cal_month_sum)

        layout.addLayout(cal_toolbar)

        self.cal_table = QTableWidget(6, 7)

        self.cal_table.setHorizontalHeaderLabels(
            ["Montag", "Dienstag", "Mittwoch", "Donnerstag", "Freitag", "Samstag", "Sonntag"]
        )
        self.cal_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.cal_table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.cal_table.verticalHeader().hide()
        self.cal_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.cal_table.setSelectionMode(QTableWidget.SelectionMode.NoSelection)
        self.heatmap_delegate = HeatmapDelegate(self.cal_table)
        self.cal_table.setItemDelegate(self.heatmap_delegate)
        layout.addWidget(self.cal_table)

    def cal_go_prev_month(self):
        """Navigiert im Kalender einen Monat zurück."""
        idx = self.cal_month_filter.currentIndex()
        if idx < self.cal_month_filter.count() - 1:
            self.cal_month_filter.setCurrentIndex(idx + 1)

    def cal_go_next_month(self):
        """Navigiert im Kalender einen Monat vor."""
        idx = self.cal_month_filter.currentIndex()
        if idx > 0:
            self.cal_month_filter.setCurrentIndex(idx - 1)

    def setup_stats_tab(self):
        """Erstellt den Tab für das Diagramm und die Statistik."""
        self.stats_layout = QVBoxLayout(self.tab_stats)
        self.figure = Figure(figsize=(8, 4))
        self.canvas = FigureCanvasQTAgg(self.figure)
        self.stats_layout.addWidget(self.canvas)

    # --- ZIELE & KALENDER LOGIK ---
    def auto_calculate_goal_hours(self):
        """Berechnet die benötigten Überstunden automatisch anhand des gewählten Zeitraums."""
        start_d = self.goal_start_edit.date()
        end_d = self.goal_end_edit.date()
        if start_d > end_d:
            return

        total_target_mins = 0
        curr = start_d

        while curr <= end_d:
            total_target_mins += self.get_target_minutes_for_date(curr.toString("yyyy-MM-dd"))
            curr = curr.addDays(1)

        total_hours_needed = total_target_mins / 60.0

        self.goal_hours_spin.blockSignals(True)
        self.goal_hours_spin.setValue(math.ceil(total_hours_needed))
        self.goal_hours_spin.blockSignals(False)

        self.on_goal_changed()

    def on_goal_changed(self):
        """Wird aufgerufen, wenn sich die Ziel-Einstellungen ändern."""
        self.settings["goal_active"] = self.goal_active_cb.isChecked()
        self.settings["goal_start_date"] = self.goal_start_edit.date().toString("yyyy-MM-dd")
        self.settings["goal_end_date"] = self.goal_end_edit.date().toString("yyyy-MM-dd")
        self.settings["goal_hours"] = self.goal_hours_spin.value()
        self.save_settings()

        self.goal_start_edit.setEnabled(self.goal_active_cb.isChecked())
        self.goal_end_edit.setEnabled(self.goal_active_cb.isChecked())
        self.goal_hours_spin.setEnabled(self.goal_active_cb.isChecked())
        self.dashboard_group.setVisible(self.goal_active_cb.isChecked())

        if self.goal_active_cb.isChecked():
            self.update_goal_status()

    def update_goal_status(self):
        """Aktualisiert die Anzeige des Fortschritts-Dashboards."""
        if not self.goal_active_cb.isChecked():
            return

        target_start_date = self.goal_start_edit.date()
        target_mins = self.goal_hours_spin.value() * 60
        current_saldo = sum(e.minutes for e in self.entries)

        progress_saldo = max(0, current_saldo)

        if target_mins == 0:
            percentage = 100
        else:
            percentage = min(100, int((progress_saldo / target_mins) * 100))

        self.goal_progress_bar.setValue(percentage)
        self.goal_progress_bar.setFormat(f"{percentage}% erreicht")
        self.lbl_goal_current.setText(self.format_time(current_saldo))

        missing_mins = target_mins - current_saldo

        if missing_mins <= 0:
            self.lbl_goal_missing.setText("0h 0m")
            self.lbl_goal_days.setText("-")
            self.lbl_goal_action.setText(
                "🎉 Herzlichen Glückwunsch! Du hast genug Überstunden für diesen Zeitraum angespart!"
            )
            self.lbl_goal_action.setStyleSheet("color: #10b981; font-weight: bold;")
            return

        self.lbl_goal_missing.setText(self.format_time(missing_mins))

        today = QDate.currentDate()
        if target_start_date <= today:
            self.lbl_goal_days.setText("0")
            self.lbl_goal_action.setText(
                "⚠️ Der gewünschte Zeitraum hat bereits begonnen oder ist heute!"
            )
            self.lbl_goal_action.setStyleSheet("color: #ef4444; font-weight: bold;")
            return

        workdays = 0
        curr = today.addDays(1)

        while curr < target_start_date:
            if self.get_target_minutes_for_date(curr.toString("yyyy-MM-dd")) > 0:
                workdays += 1
            curr = curr.addDays(1)

        self.lbl_goal_days.setText(str(workdays))

        if workdays == 0:
            self.lbl_goal_action.setText("Keine regulären Arbeitstage mehr zum Ansparen übrig!")
            self.lbl_goal_action.setStyleSheet("color: #ef4444;")
        else:
            extra_per_day = missing_mins / workdays
            self.lbl_goal_action.setText(
                f"Tipp: Wenn du ab sofort jeden Tag "
                f"<b>{int(extra_per_day)} Minuten</b> länger machst, "
                "erreichst du dein Ziel punktgenau."
            )
            self.lbl_goal_action.setStyleSheet("color: #3b82f6;")

    def update_calendar_heatmap(self):
        """Aktualisiert die Kalender-Heatmap für den ausgewählten Monat."""
        self.cal_month_filter.blockSignals(True)
        current_cal_filter = self.cal_month_filter.currentData()
        self.cal_month_filter.clear()

        months_set = set(e.date[:7] for e in self.entries if len(e.date) >= 7)
        today = QDate.currentDate()
        for i in range(-60, 61):
            months_set.add(today.addMonths(i).toString("yyyy-MM"))

        months = sorted(list(months_set), reverse=True)

        for m in months:
            self.cal_month_filter.addItem(f"{m[-2:]}/{m[:4]}", m)

        idx = self.cal_month_filter.findData(current_cal_filter)
        if idx < 0:
            idx = self.cal_month_filter.findData(today.toString("yyyy-MM"))
        if idx >= 0:
            self.cal_month_filter.setCurrentIndex(idx)

        self.btn_cal_prev.setEnabled(
            self.cal_month_filter.currentIndex() < self.cal_month_filter.count() - 1
        )
        self.btn_cal_next.setEnabled(self.cal_month_filter.currentIndex() > 0)

        self.cal_month_filter.blockSignals(False)

        sel_date_str = self.cal_month_filter.currentData()
        if not sel_date_str:
            sel_date_str = today.toString("yyyy-MM")

        year, month = map(int, sel_date_str.split('-'))
        cal = calendar.monthcalendar(year, month)

        state = self.settings.get("state", "TH")
        holidays = get_holidays(year, state)

        day_mins = {}
        monthly_sum = 0

        for e in self.entries:
            if e.date.startswith(sel_date_str):
                day_mins[e.date] = day_mins.get(e.date, 0) + e.minutes
                monthly_sum += e.minutes

        self.lbl_cal_month_sum.setText(
            f"Monats-Saldo: {self.format_time(monthly_sum, show_plus=True)}"
        )
        if monthly_sum > 0:
            self.lbl_cal_month_sum.setStyleSheet("color: #10b981;")
        elif monthly_sum < 0:
            self.lbl_cal_month_sum.setStyleSheet("color: #ef4444;")
        else:
            self.lbl_cal_month_sum.setStyleSheet("")

        self.cal_table.setRowCount(len(cal))
        is_dark = self.settings.get("dark_mode", False)
        workdays_setting = self.settings.get("workdays", [0, 1, 2, 3, 4])

        for row, week in enumerate(cal):
            for col, day in enumerate(week):
                if day == 0:
                    item = QTableWidgetItem("")
                    item.setData(Qt.ItemDataRole.UserRole + 1, False)
                    item.setBackground(QColor("#222222" if is_dark else "#f3f4f6"))
                else:
                    date_str = f"{year}-{month:02d}-{day:02d}"
                    mins = day_mins.get(date_str, 0)
                    is_holiday = date_str in holidays
                    is_workday = col in workdays_setting

                    f_mins = f"\n({self.format_time(mins, show_plus=True)})" if mins != 0 else ""
                    if is_holiday:
                        hol_name = holidays[date_str]
                        text = f"{day}\n{hol_name}{f_mins}"
                    else:
                        text = f"{day}{f_mins}"

                    item = QTableWidgetItem(text)
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

                    if is_holiday:
                        item.setToolTip(holidays[date_str])
                        if mins == 0:
                            item.setBackground(QColor("#1e3a8a" if is_dark else "#bfdbfe"))
                            item.setForeground(QColor("#60a5fa" if is_dark else "#1d4ed8"))
                        elif mins > 0:
                            alpha = min(255, 60 + (mins * 2))
                            item.setBackground(QColor(16, 185, 129, alpha))
                        else:
                            alpha = min(255, 60 + (abs(mins) * 2))
                            item.setBackground(QColor(239, 68, 68, alpha))
                    else:
                        if mins == 0:
                            if not is_workday:
                                item.setBackground(QColor("#2d3748" if is_dark else "#e5e7eb"))
                            else:
                                item.setBackground(QColor("#333333" if is_dark else "#ffffff"))
                        elif mins > 0:
                            alpha = min(255, 60 + (mins * 2))
                            item.setBackground(QColor(16, 185, 129, alpha))
                        else:
                            alpha = min(255, 60 + (abs(mins) * 2))
                            item.setBackground(QColor(239, 68, 68, alpha))

                    # Den aktuellen Tag (Heute) prüfen und markieren
                    if year == today.year() and month == today.month() and day == today.day():
                        item.setData(Qt.ItemDataRole.UserRole + 1, True)
                    else:
                        item.setData(Qt.ItemDataRole.UserRole + 1, False)

                self.cal_table.setItem(row, col, item)

    # --- ALLGEMEINE LOGIK ---
    def on_list_filter_changed(self):
        """Wird aufgerufen, wenn der Monatsfilter in der Listenansicht geändert wird."""
        filter_val = self.month_filter.currentData()
        if filter_val and filter_val != "ALL":
            idx = self.cal_month_filter.findData(filter_val)
            if idx >= 0:
                self.cal_month_filter.blockSignals(True)
                self.cal_month_filter.setCurrentIndex(idx)
                self.cal_month_filter.blockSignals(False)
        self.update_ui()

    def on_cal_filter_changed(self):
        """Wird aufgerufen, wenn der Monatsfilter im Kalender geändert wird."""
        cal_val = self.cal_month_filter.currentData()
        if cal_val:
            idx = self.month_filter.findData(cal_val)
            if idx >= 0:
                self.month_filter.blockSignals(True)
                self.month_filter.setCurrentIndex(idx)
                self.month_filter.blockSignals(False)
        self.update_calendar_heatmap()

    def open_settings(self):
        """Öffnet den Einstellungs-Dialog und übernimmt geänderte Einstellungen."""
        old_db_path = self.settings.get("db_path", DB_FILE)
        dialog = SettingsDialog(self.settings, self)
        if dialog.exec():
            new_settings = dialog.get_settings()
            new_db_path = new_settings.get("db_path", DB_FILE)

            if new_db_path != old_db_path:
                if os.path.exists(old_db_path) and not os.path.exists(new_db_path):
                    reply = QMessageBox.question(
                        self, "Datenbank verschieben?",
                        "Soll die bestehende Datenbank an den neuen Ort verschoben werden?\n"
                        "(Bei 'Nein' wird am neuen Ort eine neue, leere Datenbank erstellt.)",
                        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No |
                        QMessageBox.StandardButton.Cancel
                    )
                    if reply == QMessageBox.StandardButton.Cancel:
                        return
                    if reply == QMessageBox.StandardButton.Yes:
                        try:
                            shutil.move(old_db_path, new_db_path)
                        except OSError as e:
                            QMessageBox.critical(self, "Fehler",
                                               f"Datenbank konnte nicht verschoben werden:\n{e}")
                            return

                self.db.close()
                self.db = DBManager(new_db_path)
                self.entries = self.db.load_all()

            self.settings.update(new_settings)
            self.save_settings()
            self.apply_theme()
            self.pause_spin.setEnabled(not self.settings.get("auto_break", True))

            # Alle Tage neu berechnen, da sich das Soll geändert haben könnte
            all_dates = sorted(list(set(e.date for e in self.entries)))
            for d_str in all_dates:
                self.recalculate_day(d_str)

            self.load_data() # Lädt alles neu (inkl. UI-Update)

            new_start = self._get_default_start_time()
            self.time_start.blockSignals(True)
            self.time_start.setTime(new_start)
            self.time_start.blockSignals(False)
            self.on_start_time_changed(new_start)

    def _compute_target_end_time(self, new_start: "QTime") -> "QTime":
        """Berechnet die Endzeit so dass das Tagessoll (Regelarbeitszeit) erreicht wird.
        Berücksichtigt bereits gespeicherte Einträge (auch manuelle) für das Datum."""
        curr_date_str = self.date_edit.date().toString("yyyy-MM-dd")
        all_day = [e for e in self.entries if e.date == curr_date_str]
        timed_existing = [e for e in all_day if e.start and e.end]
        manual_sum = sum(e.minutes for e in all_day if not (e.start and e.end))

        target_mins = self.get_target_minutes_for_date(curr_date_str)
        max_mins = self.get_max_minutes()
        is_auto = self.settings.get("auto_break", True)

        # Wir suchen die kleinste Dauer (in Minuten), die das Soll erfüllt.
        for duration_mins in range(0, max_mins * 2 + 1):
            temp = WorkEntry(
                id=-1,
                date=curr_date_str,
                start=new_start.toString("HH:mm"),
                end=new_start.addSecs(duration_mins * 60).toString("HH:mm"),
                pause=self.pause_spin.value() if not is_auto else 0,
                minutes=0,
                reason=""
            )
            # calculate_timed_entries liefert uns das Netto der Zeiteinträge (timed_existing + temp)
            _, total_net_timed = calculate_timed_entries(
                timed_existing + [temp], target_mins, max_mins, is_auto
            )

            # Das gesamte Tagessaldo ist Netto-Zeit + manuelle Korrekturen
            if (total_net_timed + manual_sum) >= target_mins:
                break

        return new_start.addSecs(duration_mins * 60)

    def on_start_time_changed(self, new_start_time):
        """Wird aufgerufen, wenn die Startzeit geändert wird; aktualisiert die Endzeit-Vorschau."""
        today = QDate.currentDate().toString("yyyy-MM-dd")
        self.settings["last_date"] = today
        self.settings["last_start"] = new_start_time.toString("HH:mm")

        self.time_end.blockSignals(True)
        self.time_end.setTime(self._compute_target_end_time(new_start_time))
        self.time_end.blockSignals(False)
        self.update_live_calc()

    def update_live_calc(self):
        """Berechnet die Überstunden-Vorschau live und zeigt sie im Label an."""
        curr_date_str = self.date_edit.date().toString("yyyy-MM-dd")
        target_mins = self.get_target_minutes_for_date(curr_date_str)
        max_mins = self.get_max_minutes()
        is_auto = self.settings.get("auto_break", True)

        current_temp = WorkEntry(
            id=-1,
            date=curr_date_str,
            start=self.time_start.time().toString("HH:mm"),
            end=self.time_end.time().toString("HH:mm"),
            pause=self.pause_spin.value() if not is_auto else 0,
            minutes=0,
            reason=""
        )

        all_day = [e for e in self.entries if e.date == curr_date_str]
        timed = [e for e in all_day if e.start and e.end] + [current_temp]
        manual_sum = sum(e.minutes for e in all_day if not (e.start and e.end))

        results, total_net = calculate_timed_entries(timed, target_mins, max_mins, is_auto)
        entry_pause, entry_overtime = results[-1]

        if is_auto:
            self.pause_spin.blockSignals(True)
            self.pause_spin.setValue(entry_pause)
            self.pause_spin.blockSignals(False)

        self.current_calculated_pause = entry_pause
        self.current_calculated_overtime = entry_overtime

        final_total_overtime = (total_net - target_mins) + manual_sum

        calc_text = (
            f"Netto (Tag): {self.format_time(total_net)} ➔ "
            f"<b>{self.format_time(final_total_overtime, show_plus=True)}"
            " Überstunden (Tag-Saldo)</b>"
        )
        warnings = []
        if total_net >= max_mins:
            warnings.append(f"⚠️ Max. {max_mins // 60}h erreicht!")

        # Ruhezeit-Check (Lücke zum Vortag/letzten Eintrag davor)
        prev_entry = self.db.get_last_entry_before(curr_date_str)
        if prev_entry and prev_entry.end:
            try:
                dt_prev = datetime.strptime(f"{prev_entry.date} {prev_entry.end}", "%Y-%m-%d %H:%M")
                start_str_curr = self.time_start.time().toString('HH:mm')
                dt_curr = datetime.strptime(
                    f"{curr_date_str} {start_str_curr}", "%Y-%m-%d %H:%M"
                )
                rest_hours = (dt_curr - dt_prev).total_seconds() / 3600
                if 0 < rest_hours < 11:
                    warnings.append(f"⚠️ Ruhezeit verletzt ({rest_hours:.1f}h < 11h)")
            except ValueError:
                pass

        if warnings:
            calc_text += f" <span style='color: #ef4444;'>{' | '.join(warnings)}</span>"
            self.lbl_live_calc.setStyleSheet("color: #ef4444;")
        else:
            self.lbl_live_calc.setStyleSheet("")
        self.lbl_live_calc.setText(calc_text)

    def recalculate_day(self, date_str):
        """Verteilt Pausen und Überstunden für alle Zeiteinträge eines Tages neu.
        Manuelle Einträge (ohne Start-/Endzeit) werden nicht angefasst."""
        day_entries = [e for e in self.entries if e.date == date_str]
        if not day_entries:
            return

        timed = [e for e in day_entries if e.start and e.end]
        if not timed:
            return

        target_mins = self.get_target_minutes_for_date(date_str)
        max_mins = self.get_max_minutes()
        is_auto = self.settings.get("auto_break", True)

        results, _ = calculate_timed_entries(timed, target_mins, max_mins, is_auto)

        for e in timed:
            e.pause, e.minutes = results[e.id]
            self.db.update(e)

    def set_now_as_end(self):
        """Setzt die Endzeit auf die aktuelle Uhrzeit."""
        self.time_end.setTime(QTime.currentTime())

    def check_overlap(self, date_str, start_str, end_str, exclude_id=None):
        """Prüft, ob sich der Zeitraum mit bestehenden Einträgen am selben Tag überschneidet."""
        if not start_str or not end_str:
            return None

        s_new = QTime.fromString(start_str, "HH:mm")
        e_new = QTime.fromString(end_str, "HH:mm")

        # Falls Mitternacht überschritten wird, ist die Logik komplexer,
        # hier vereinfacht für den gleichen Tag:
        for e in self.entries:
            if e.date == date_str and e.start and e.end and e.id != exclude_id:
                s_old = QTime.fromString(e.start, "HH:mm")
                e_old = QTime.fromString(e.end, "HH:mm")

                # Standard Überlappungs-Check: (StartA < EndeB) und (EndeA > StartB)
                # Wir nutzen secsTo für den Vergleich
                if s_new.secsTo(e_old) > 0 and s_old.secsTo(e_new) > 0:
                    return f"{e.start} - {e.end} ({e.reason or 'Ohne Anlass'})"
        return None

    def add_entry(self):
        """Liest die Eingabefelder aus, prüft auf Überlappung und fügt den Eintrag in die DB ein."""
        date_str = self.date_edit.date().toString("yyyy-MM-dd")
        start_str = self.time_start.time().toString("HH:mm")
        end_str = self.time_end.time().toString("HH:mm")

        # Überlappungs-Check
        overlap = self.check_overlap(date_str, start_str, end_str)
        if overlap:
            QMessageBox.warning(
                self, "Überschneidung",
                f"Dieser Zeitraum überschneidet sich mit einem existierenden Eintrag:"
                f"\n\n{overlap}\n\nBitte korrigiere die Zeiten."
            )
            return

        entry = WorkEntry(
            id=None,
            date=date_str,
            start=start_str,
            end=end_str,
            pause=self.current_calculated_pause,
            minutes=self.current_calculated_overtime,
            reason=self.reason_edit.text().strip(),
            target_minutes=(
                self.custom_target_time.time().hour() * 60
                + self.custom_target_time.time().minute()
            ) if self.custom_target_cb.isChecked() else -1
        )
        self.db.insert(entry)
        self.reason_edit.clear()
        self.custom_target_cb.setChecked(False)
        self.date_edit.setDate(QDate.currentDate())

        # Alle Einträge neu laden und den betroffenen Tag glattziehen
        self.entries = self.db.load_all()
        self.recalculate_day(date_str)
        self.load_data()

    def delete_entry(self, entry: WorkEntry):
        """Fragt den Benutzer nach Bestätigung und löscht dann den übergebenen Eintrag."""
        date_str = entry.date
        d = QDate.fromString(date_str, "yyyy-MM-dd").toString("dd.MM.yyyy")
        reply = QMessageBox.question(self, "Löschen bestätigen",
            f"Eintrag vom {d} wirklich löschen?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            self.db.delete(entry.id)
            # Nach Löschen: Tag neu berechnen
            self.entries = [e for e in self.entries if e.id != entry.id]
            self.recalculate_day(date_str)
            self.load_data()

    def edit_entry(self, row, _column):
        """Öffnet den Bearbeitungs-Dialog für den Eintrag in der angeklickten Zeile."""
        entry_idx = self.table.item(row, 0).data(Qt.ItemDataRole.UserRole)
        entry = self.entries[entry_idx]
        old_date = entry.date

        dialog = EditDialog(
            entry, self.entries, self.get_target_minutes(),
            self.get_max_minutes(), self.settings.get("auto_break", True), self
        )
        if dialog.exec():
            # Bevor wir anwenden: Überlappungs-Check (mit den neuen Werten aus dem Dialog)
            # Wir holen uns die Werte temporär
            new_date = dialog.date_edit.date().toString("yyyy-MM-dd")
            new_start = (
                dialog.time_start.time().toString("HH:mm")
                if dialog.has_times_cb.isChecked() else ""
            )
            new_end = (
                dialog.time_end.time().toString("HH:mm")
                if dialog.has_times_cb.isChecked() else ""
            )

            overlap = self.check_overlap(new_date, new_start, new_end, exclude_id=entry.id)
            if overlap:
                QMessageBox.warning(
                    self, "Überschneidung",
                    f"Die Änderungen überschneiden sich mit einem anderen Eintrag:"
                    f"\n\n{overlap}\n\nBitte korrigiere die Zeiten."
                )
                return

            dialog.apply_to_entry()
            self.db.update(entry)

            # Neu laden und beide betroffenen Tage (alt/neu) neu berechnen
            self.entries = self.db.load_all()
            self.recalculate_day(old_date)
            if entry.date != old_date:
                self.recalculate_day(entry.date)
            self.load_data()

    def format_time(self, total_minutes, show_plus=False):
        """Formatiert Minuten in Stunden und Minuten.
        Unter 60 Min -> '45m'
        Ab 60 Min    -> '1h 5m'
        """
        sign = "+" if show_plus and total_minutes > 0 else ("-" if total_minutes < 0 else "")
        abs_m = abs(total_minutes)
        if abs_m < 60:
            return f"{sign}{abs_m}m"
        return f"{sign}{abs_m // 60}h {abs_m % 60}m"

    def update_ui(self):
        """Aktualisiert die Eintrags-Tabelle und den Gesamtsaldo entsprechend dem aktiven Filter."""
        self.month_filter.blockSignals(True)
        current_filter = self.month_filter.currentData()
        self.month_filter.clear()
        self.month_filter.addItem("Alle", "ALL")

        months = sorted(
            list(set(e.date[:7] for e in self.entries if len(e.date) >= 7)), reverse=True
        )
        for m in months:
            self.month_filter.addItem(f"{m[-2:]}/{m[:4]}", m)
        idx = self.month_filter.findData(current_filter)
        if idx >= 0:
            self.month_filter.setCurrentIndex(idx)
        self.month_filter.blockSignals(False)

        self.table.setRowCount(0)
        filter_val = self.month_filter.currentData()
        total_overall = sum(e.minutes for e in self.entries)

        row = 0
        for i, e in enumerate(self.entries):
            if filter_val != "ALL" and not e.date.startswith(filter_val):
                continue

            self.table.insertRow(row)
            item_date = QTableWidgetItem(
                QDate.fromString(e.date, "yyyy-MM-dd").toString("dd.MM.yyyy")
            )
            item_date.setData(Qt.ItemDataRole.UserRole, i)
            item_date.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

            z_str = (
                f"{e.start} - {e.end}" + (f" (-{e.pause}m)" if e.pause > 0 else "")
                if e.start else "-"
            )
            item_zeit = QTableWidgetItem(z_str)
            item_zeit.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

            item_min = QTableWidgetItem(self.format_time(e.minutes, show_plus=True))
            item_min.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            if e.minutes > 0:
                item_min.setForeground(QColor("#10b981"))
            elif e.minutes < 0:
                item_min.setForeground(QColor("#ef4444"))

            btn_del = QPushButton("Löschen")
            btn_del.clicked.connect(lambda checked, ent=e: self.delete_entry(ent))

            self.table.setItem(row, 0, item_date)
            self.table.setItem(row, 1, item_zeit)
            self.table.setItem(row, 2, item_min)
            self.table.setItem(row, 3, QTableWidgetItem(e.reason))
            self.table.setCellWidget(row, 4, btn_del)
            row += 1

        self.lbl_saldo.setText(self.format_time(total_overall))
        if total_overall > 0:
            self.lbl_saldo.setStyleSheet("color: #10b981;")
        elif total_overall < 0:
            self.lbl_saldo.setStyleSheet("color: #ef4444;")
        else:
            self.lbl_saldo.setStyleSheet("")

    def update_stats_chart(self):
        """Zeichnet das monatliche Balkendiagramm der Überstunden neu."""
        monthly_totals = {}
        for e in reversed(self.entries):
            if len(e.date) >= 7:
                m = e.date[:7]
                monthly_totals[m] = monthly_totals.get(m, 0) + e.minutes

        self.figure.clear()
        is_dark = self.settings.get("dark_mode", False)
        bg_color = '#31363b' if is_dark else '#ffffff'
        text_color = '#e0e0e0' if is_dark else '#000000'

        self.figure.patch.set_facecolor(bg_color)
        ax = self.figure.add_subplot(111)
        ax.set_facecolor(bg_color)
        ax.tick_params(colors=text_color)
        for spine in ax.spines.values():
            spine.set_edgecolor(text_color)

        if not monthly_totals:
            ax.text(0.5, 0.5, "Keine Daten vorhanden", color=text_color, ha='center', va='center')
        else:
            months = list(monthly_totals.keys())
            values = [v / 60 for v in monthly_totals.values()]

            colors = ['#10b981' if v >= 0 else '#ef4444' for v in values]
            ax.bar(months, values, color=colors)
            ax.set_ylabel("Überstunden (in Stunden)", color=text_color)
            ax.set_title("Monatlicher Überstunden-Verlauf", color=text_color)
            self.figure.autofmt_xdate()

        self.canvas.draw()

    # --- EXPORT HILFSMETHODEN ---
    def _get_export_entries(self):
        filter_val = self.month_filter.currentData()
        return [e for e in self.entries if filter_val == "ALL" or e.date.startswith(filter_val)]

    def _get_export_title(self):
        filter_val = self.month_filter.currentData()
        return "Alle Einträge" if filter_val == "ALL" else f"Monat {filter_val}"

    def _export_row_data(self, e):
        """Gibt (datum_str, zeitraum_str) für einen Eintrag zurück."""
        d = QDate.fromString(e.date, "yyyy-MM-dd").toString("dd.MM.yyyy") if e.date else ""
        if e.start and e.end:
            pause_str = f" (-{e.pause}m)" if e.pause > 0 else ""
            zeitraum = f"{e.start} – {e.end}{pause_str}"
        else:
            zeitraum = "–"
        return d, zeitraum

    # --- EXPORT ---
    def export_csv(self):
        """Exportiert die aktuell gefilterten Einträge als CSV-Datei."""
        file_name, _ = QFileDialog.getSaveFileName(
            self, "CSV Export", "ueberstunden_export.csv", "CSV Dateien (*.csv)")
        if not file_name:
            return
        try:
            export_entries = self._get_export_entries()
            with open(file_name, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f, delimiter=";")
                writer.writerow(["Datum", "Zeitraum", "Minuten", "Dauer", "Anlass"])
                for e in export_entries:
                    d, zeitraum = self._export_row_data(e)
                    writer.writerow([d, zeitraum, e.minutes, self.format_time(e.minutes), e.reason])
                total = sum(e.minutes for e in export_entries)
                writer.writerow(["Gesamt", "", "", self.format_time(total), ""])
            QMessageBox.information(self, "Erfolg", "CSV erfolgreich exportiert!")
        except Exception as ex:  # pylint: disable=broad-except
            QMessageBox.critical(self, "Fehler", f"Fehler beim CSV-Export:\n{str(ex)}")

    def export_xlsx(self):
        """Exportiert die aktuell gefilterten Einträge als Excel-Datei (xlsx)."""
        if not _OPENPYXL:
            QMessageBox.critical(self, "Fehler",
                "openpyxl ist nicht installiert.\nBitte ausführen: pip install openpyxl")
            return

        file_name, _ = QFileDialog.getSaveFileName(
            self, "Excel Export", "ueberstunden_export.xlsx", "Excel Dateien (*.xlsx)")
        if not file_name:
            return
        try:
            export_entries = self._get_export_entries()
            total_min = sum(e.minutes for e in export_entries)

            wb = Workbook()
            ws = wb.active
            ws.title = "Überstunden"

            # Titel
            ws.merge_cells("A1:E1")
            ws["A1"] = f"Überstunden-Nachweis – {self._get_export_title()}"
            ws["A1"].font = Font(bold=True, size=13)
            ws["A1"].alignment = Alignment(horizontal="center")

            # Kopfzeile
            hdr_fill = PatternFill("solid", fgColor="3b82f6")
            hdr_font = Font(bold=True, color="FFFFFF")
            hdr_align = Alignment(horizontal="center")
            for col, text in enumerate(["Datum", "Zeitraum", "Minuten", "Dauer", "Anlass"], 1):
                c = ws.cell(row=3, column=col, value=text)
                c.font = hdr_font
                c.fill = hdr_fill
                c.alignment = hdr_align

            # Datenzeilen
            alt_fill = PatternFill("solid", fgColor="f3f4f6")
            for i, e in enumerate(export_entries):
                row = i + 4
                d, zeitraum = self._export_row_data(e)
                values = [d, zeitraum, e.minutes, self.format_time(e.minutes), e.reason]
                for col, val in enumerate(values, 1):
                    c = ws.cell(row=row, column=col, value=val)
                    if i % 2 == 1:
                        c.fill = alt_fill
                ovt_cell = ws.cell(row=row, column=3)
                if e.minutes > 0:
                    ovt_cell.font = Font(color="059669")
                elif e.minutes < 0:
                    ovt_cell.font = Font(color="dc2626")

            # Summenzeile
            sum_row = len(export_entries) + 4
            sum_fill = PatternFill("solid", fgColor="dbeafe")
            ws.merge_cells(f"A{sum_row}:C{sum_row}")
            ws[f"A{sum_row}"] = "Gesamtsumme:"
            ws[f"D{sum_row}"] = self.format_time(total_min)
            for col in range(1, 6):
                c = ws.cell(row=sum_row, column=col)
                c.font = Font(bold=True)
                c.fill = sum_fill

            # Spaltenbreiten
            for col, width in zip("ABCDE", [14, 24, 18, 12, 32]):
                ws.column_dimensions[col].width = width

            wb.save(file_name)
            QMessageBox.information(self, "Erfolg", "Excel-Datei erfolgreich exportiert!")
        except Exception as ex:  # pylint: disable=broad-except
            QMessageBox.critical(self, "Fehler", f"Fehler beim Excel-Export:\n{str(ex)}")

    def export_pdf(self):
        """Exportiert die aktuell gefilterten Einträge als PDF-Datei."""
        # pylint: disable=import-outside-toplevel
        from PyQt6.QtPrintSupport import QPrinter
        from PyQt6.QtGui import QTextDocument
        from PyQt6.QtCore import QSizeF

        file_name, _ = QFileDialog.getSaveFileName(
            self, "PDF Export", "ueberstunden_export.pdf", "PDF Dateien (*.pdf)")
        if not file_name:
            return
        try:
            export_entries = self._get_export_entries()
            total_min = sum(e.minutes for e in export_entries)

            rows_html = ""
            for i, e in enumerate(export_entries):
                d, zeitraum = self._export_row_data(e)
                bg = "#f9fafb" if i % 2 == 1 else "#ffffff"
                if e.minutes > 0:
                    color = "#059669"
                elif e.minutes < 0:
                    color = "#dc2626"
                else:
                    color = "inherit"
                rows_html += (
                    f"<tr style='background:{bg}'>"
                    f"<td>{d}</td><td>{zeitraum}</td>"
                    f"<td style='color:{color};text-align:right'>"
                    f"{self.format_time(e.minutes, show_plus=True)}</td>"
                    f"<td>{e.reason}</td></tr>"
                )

            html = f"""
            <html><head><meta charset="utf-8"><style>
                body {{ font-family: Arial, sans-serif; font-size: 11px; color: #111; }}
                h2 {{ font-size: 14px; margin-bottom: 4px; }}
                p.sub {{ color: #666; font-size: 10px; margin-top: 0; }}
                table {{ width: 100%; border-collapse: collapse; margin-top: 12px; }}
                th {{ background: #3b82f6; color: #fff; padding: 5px 8px; text-align: left; }}
                td {{ border-bottom: 1px solid #e5e7eb; padding: 4px 8px; }}
                tr.sum td {{ background: #dbeafe; font-weight: bold; }}
            </style></head><body>
            <h2>Überstunden-Nachweis</h2>
            <p class="sub">{self._get_export_title()} &nbsp;·&nbsp;
               Erstellt am {QDate.currentDate().toString("dd.MM.yyyy")}</p>
            <table>
              <tr><th>Datum</th><th>Zeitraum</th><th>Überstunden</th><th>Anlass</th></tr>
              {rows_html}
              <tr class="sum">
                <td colspan="2">Gesamtsumme:</td>
                <td style='text-align:right'>
                  {self.format_time(total_min, show_plus=True)}</td><td></td>
              </tr>
            </table>
            </body></html>"""

            printer = QPrinter(QPrinter.PrinterMode.HighResolution)
            printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
            printer.setOutputFileName(file_name)

            doc = QTextDocument()
            doc.setHtml(html)
            doc.setPageSize(QSizeF(printer.pageRect(QPrinter.Unit.Point).size()))
            doc.print(printer)

            QMessageBox.information(self, "Erfolg", "PDF erfolgreich exportiert!")
        except Exception as ex:  # pylint: disable=broad-except
            QMessageBox.critical(self, "Fehler", f"Fehler beim PDF-Export:\n{str(ex)}")

    def import_csv(self):
        """Importiert Einträge aus einer CSV-Datei und fügt sie zur Datenbank hinzu."""
        file_name, _ = QFileDialog.getOpenFileName(self, "CSV Import", "", "CSV Dateien (*.csv)")
        if not file_name:
            return
        try:
            with open(file_name, newline="", encoding="utf-8", errors="replace") as csvfile:
                content = csvfile.read()
                if not content:
                    return
                delimiter = ";" if ";" in content.split("\n")[0] else ","
                csvfile.seek(0)
                reader = csv.reader(csvfile, delimiter=delimiter)

                pending = []
                affected_dates = set()
                for row in reader:
                    if len(row) >= 1:
                        date_str = row[0].strip()
                        if date_str.lower() in ["datum", "date"] or not date_str:
                            continue

                        # Minuten optional behandeln
                        try:
                            minutes = int(row[1].strip()) if len(row) > 1 and row[1].strip() else 0
                        except ValueError:
                            minutes = 0

                        reason = row[2].strip() if len(row) > 2 else ""
                        start_str = row[3].strip() if len(row) > 3 else ""
                        end_str = row[4].strip() if len(row) > 4 else ""
                        try:
                            pause = int(row[5].strip()) if len(row) > 5 else 0
                        except ValueError:
                            pause = 0

                        parsed_date = ""
                        for fmt in ("%d.%m.%y", "%d.%m.%Y", "%Y-%m-%d"):
                            try:
                                parsed_date = datetime.strptime(date_str, fmt).strftime("%Y-%m-%d")
                                break
                            except ValueError:
                                pass

                        if not parsed_date:
                            parsed_date = date_str # Fallback

                        pending.append(WorkEntry(
                            id=None, date=parsed_date, start=start_str,
                            end=end_str, pause=pause, minutes=minutes, reason=reason
                        ))
                        affected_dates.add(parsed_date)

            if not pending:
                QMessageBox.information(self, "Import", "Keine importierbaren Einträge gefunden.")
                return

            preview_lines = [f"  {e.date}  {e.start}-{e.end}  {e.reason}" for e in pending[:5]]
            if len(pending) > 5:
                preview_lines.append(f"  … und {len(pending) - 5} weitere")
            preview = "\n".join(preview_lines)

            reply = QMessageBox.question(
                self, "Import bestätigen",
                f"{len(pending)} Einträge gefunden:\n\n{preview}\n\nJetzt importieren?\n"
                "(Überstunden werden automatisch für jeden Tag berechnet/konsolidiert)",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if reply != QMessageBox.StandardButton.Yes:
                return

            current_db = self.settings.get("db_path", DB_FILE)
            if os.path.exists(current_db):
                shutil.copy2(current_db, current_db + ".backup")

            for entry in pending:
                self.db.insert(entry)

            # Alle Einträge neu laden und alle betroffenen Tage glattziehen
            self.entries = self.db.load_all()
            for d in sorted(list(affected_dates)):
                self.recalculate_day(d)

            self.load_data()
            QMessageBox.information(
                self, "Erfolg",
                f"{len(pending)} Einträge importiert!\n"
                "Tagessalden wurden automatisch berechnet.\nBackup der Datenbank angelegt."
            )

        except Exception as ex:  # pylint: disable=broad-except
            QMessageBox.critical(self, "Fehler", f"Fehler beim Import:\n{str(ex)}")


if __name__ == "__main__":
    QApplication.setHighDpiScaleFactorRoundingPolicy(
        Qt.HighDpiScaleFactorRoundingPolicy.PassThrough
    )
    app = QApplication(sys.argv)
    window = UeberstundenApp()
    window.show()
    sys.exit(app.exec())
