"""
Überstundenrechner Pro - Ein Tool zur Erfassung und Berechnung von Arbeitsstunden.
"""
import calendar
import csv
import getpass
import json
import math
import os
import re
import shutil
import sqlite3
import subprocess
import sys
from dataclasses import dataclass
from datetime import datetime

# --- KONFIGURATION & PFADE (PyInstaller-kompatibel) ---
# Muss VOR dem matplotlib-Import stehen, damit MPLCONFIGDIR gesetzt werden kann.
if getattr(sys, 'frozen', False):
    if sys.platform == 'darwin':
        BASE_DIR = os.path.abspath(os.path.join(sys.executable, '..', '..', '..', '..'))
    else:
        BASE_DIR = os.path.dirname(sys.executable)
    # pylint: disable=protected-access
    BUNDLE_DIR = sys._MEIPASS if hasattr(sys, '_MEIPASS') else BASE_DIR
    # Matplotlib Font-Cache in einen persistenten Ordner leiten,
    # sonst wird er bei jedem Start neu gebaut → starke Verzögerung.
    os.environ.setdefault('MPLCONFIGDIR', os.path.join(BASE_DIR, '.mplconfig'))
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    BUNDLE_DIR = BASE_DIR

# pylint: disable=wrong-import-position
import matplotlib
matplotlib.use('QtAgg')
from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg
from matplotlib.figure import Figure

from PyQt6.QtCore import QDate, Qt, QTime, QRect, QPoint
from PyQt6.QtGui import QColor, QFont, QIcon, QPalette, QPainter, QPen, QPixmap, QPolygon
from PyQt6.QtWidgets import (
    QApplication, QCheckBox, QComboBox, QDateEdit, QDialog,
    QFileDialog, QFrame, QHBoxLayout, QHeaderView, QLabel,
    QLineEdit, QMainWindow, QMenu, QMessageBox, QPushButton, QSpinBox,
    QTableWidget, QTableWidgetItem, QTabWidget, QTimeEdit,
    QVBoxLayout, QWidget, QGroupBox, QProgressBar, QGridLayout, QStyledItemDelegate
)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    _OPENPYXL = True
except ImportError:
    _OPENPYXL = False

DB_FILE = os.path.join(BASE_DIR, "ueberstunden_daten.db")
SETTINGS_FILE = os.path.join(BASE_DIR, "ueberstunden_settings.json")
ICON_PATH = os.path.join(BUNDLE_DIR, "icon.png")

BUNDESLAENDER = {
    "BW": "Baden-Württemberg", "BY": "Bayern", "BE": "Berlin", "BB": "Brandenburg",
    "HB": "Bremen", "HH": "Hamburg", "HE": "Hessen", "MV": "Mecklenburg-Vorpommern",
    "NI": "Niedersachsen", "NW": "Nordrhein-Westfalen", "RP": "Rheinland-Pfalz",
    "SL": "Saarland", "SN": "Sachsen", "ST": "Sachsen-Anhalt",
    "SH": "Schleswig-Holstein", "TH": "Thüringen"
}

# --- DATENKLASSEN ---
@dataclass
class WorkEntry:
    """
    Repräsentiert einen einzelnen Arbeitseintrag.
    """
    id: int
    date: str
    start: str
    end: str
    pause: int
    minutes: int
    reason: str
    target_minutes: int = -1 # -1 means use default/settings

# --- FEIERTAGS-RECHNER (DEUTSCHLAND) ---
def get_holidays(year, state):
    """
    Berechnet die gesetzlichen Feiertage für ein gegebenes Jahr und Bundesland.
    """
    a = year % 19
    b = year // 100
    c = year % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day = ((h + l - 7 * m + 114) % 31) + 1
    easter = QDate(year, month, day)

    holidays = {
        f"{year}-01-01": "Neujahr",
        easter.addDays(-2).toString("yyyy-MM-dd"): "Karfreitag",
        easter.addDays(1).toString("yyyy-MM-dd"): "Ostermontag",
        f"{year}-05-01": "Tag der Arbeit",
        easter.addDays(39).toString("yyyy-MM-dd"): "Christi Himmelfahrt",
        easter.addDays(50).toString("yyyy-MM-dd"): "Pfingstmontag",
        f"{year}-10-03": "Tag d. Dt. Einheit",
        f"{year}-12-25": "1. Weihnachtstag",
        f"{year}-12-26": "2. Weihnachtstag",
    }

    if state in ["BW", "BY", "ST"]:
        holidays[f"{year}-01-06"] = "Hl. Drei Könige"
    if state in ["BE", "MV"]:
        holidays[f"{year}-03-08"] = "Int. Frauentag"
    if state in ["BW", "HE", "NW", "RP", "SL"] or (state == "BY"):
        holidays[easter.addDays(60).toString("yyyy-MM-dd")] = "Fronleichnam"
    if state in ["BY", "SL"]:
        holidays[f"{year}-08-15"] = "Mariä Himmelfahrt"
    if state == "TH":
        holidays[f"{year}-09-20"] = "Weltkindertag"
    if state in ["BB", "HB", "HH", "MV", "NI", "SH", "SN", "ST", "TH"]:
        holidays[f"{year}-10-31"] = "Reformationstag"
    if state in ["BW", "BY", "NW", "RP", "SL"]:
        holidays[f"{year}-11-01"] = "Allerheiligen"
    if state == "SN":
        d = QDate(year, 11, 22)
        while d.dayOfWeek() != 3:
            d = d.addDays(-1)
        holidays[d.toString("yyyy-MM-dd")] = "Buß- und Bettag"

    return holidays



# --- DATENBANK MANAGER ---
class DBManager:
    """
    Verwaltet die SQLite-Datenbankverbindung und Operationen für Arbeitseinträge.
    """
    def __init__(self, db_path):
        """
        Initialisiert den DBManager und erstellt die Tabelle, falls sie nicht existiert.
        """
        self.conn = sqlite3.connect(db_path)
        self.create_table()

    def create_table(self):
        """
        Erstellt die 'entries'-Tabelle und führt notwendige Migrationen durch.
        """
        cursor = self.conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS entries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                date TEXT NOT NULL,
                start TEXT,
                end TEXT,
                pause INTEGER,
                minutes INTEGER,
                reason TEXT,
                target_minutes INTEGER DEFAULT -1
            )
        """)
        # Migration: Check if target_minutes exists
        cursor.execute("PRAGMA table_info(entries)")
        columns = [row[1] for row in cursor.fetchall()]
        if "target_minutes" not in columns:
            cursor.execute("ALTER TABLE entries ADD COLUMN target_minutes INTEGER DEFAULT -1")
        self.conn.commit()

    def load_all(self):
        """
        Lädt alle Arbeitseinträge aus der Datenbank, sortiert nach Datum und Startzeit.
        """
        cursor = self.conn.cursor()
        cursor.execute("SELECT * FROM entries ORDER BY date DESC, start DESC")
        return [WorkEntry(*row) for row in cursor.fetchall()]

    def insert(self, entry: WorkEntry):
        """
        Fügt einen neuen Arbeitseintrag in die Datenbank ein.
        """
        cursor = self.conn.cursor()
        cursor.execute("""
            INSERT INTO entries (date, start, end, pause, minutes, reason, target_minutes)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (entry.date, entry.start, entry.end, entry.pause,
              entry.minutes, entry.reason, entry.target_minutes))
        self.conn.commit()
        entry.id = cursor.lastrowid

    def update(self, entry: WorkEntry):
        """
        Aktualisiert einen bestehenden Arbeitseintrag in der Datenbank.
        """
        cursor = self.conn.cursor()
        cursor.execute("""
            UPDATE entries SET date=?, start=?, end=?, pause=?, minutes=?, reason=?,
                               target_minutes=? WHERE id=?
        """, (entry.date, entry.start, entry.end, entry.pause, entry.minutes,
              entry.reason, entry.target_minutes, entry.id))
        self.conn.commit()

    def delete(self, entry_id: int):
        """
        Löscht einen Arbeitseintrag anhand seiner ID.
        """
        cursor = self.conn.cursor()
        cursor.execute("DELETE FROM entries WHERE id=?", (entry_id,))
        self.conn.commit()

    def get_last_entry_before(self, date_str: str):
        """
        Findet den zeitlich letzten Eintrag vor dem angegebenen Datum, der eine Endzeit hat.
        """
        cursor = self.conn.cursor()
        cursor.execute("""
            SELECT * FROM entries WHERE date < ? AND end != ''
            ORDER BY date DESC, end DESC LIMIT 1
        """, (date_str,))
        row = cursor.fetchone()
        return WorkEntry(*row) if row else None

    def close(self):
        """
        Schließt die Datenbankverbindung.
        """
        self.conn.close()


# --- ZENTRALE GESCHÄFTSLOGIK ---
def calculate_timed_entries(timed_entries, target_mins, max_mins, is_auto):
    """Berechnet Pause und Überstunden für die Zeiteinträge eines Tages.

    Gibt ein Tupel (results, total_net) zurück:
      results   — dict {entry.id: (pause_min, overtime_min)}
      total_net — tatsächliche Netto-Arbeitszeit des Tages (nach Cap) in Minuten

    Nur der letzte Eintrag nach Startzeit erhält die Überstunden; alle anderen 0.
    Manuelle Einträge (ohne Start-/Endzeit) werden nicht übergeben und bleiben unberührt.
    """
    sorted_entries = sorted(timed_entries, key=lambda x: x.start or "00:00")
    results = {}
    total_accumulated_gross = 0
    total_accumulated_gap = 0
    recorded_pause_distributed = 0
    last_end_qtime = None

    for i, e in enumerate(sorted_entries):
        current_gross = 0
        if e.start and e.end:
            try:
                s = QTime.fromString(e.start, "HH:mm")
                en = QTime.fromString(e.end, "HH:mm")
                diff = s.secsTo(en) // 60
                if diff < 0:
                    diff += 24 * 60
                current_gross = diff
                if last_end_qtime:
                    gap = last_end_qtime.secsTo(s) // 60
                    if gap < 0:
                        gap += 24 * 60
                    total_accumulated_gap += max(0, gap)
                last_end_qtime = en
            except (ValueError, TypeError):
                pass

        total_accumulated_gross += current_gross

        if is_auto:
            if total_accumulated_gross > 9 * 60:
                req = 45
            elif total_accumulated_gross > 6 * 60:
                req = 30
            else:
                req = 0
            current_total_pause_needed = max(0, req - total_accumulated_gap)
            current_break = max(0, current_total_pause_needed - recorded_pause_distributed)
        else:
            current_break = e.pause

        if i == len(sorted_entries) - 1:
            total_net = total_accumulated_gross - (recorded_pause_distributed + current_break)
            ovt = int(min(max_mins, total_net) - target_mins)
        else:
            ovt = 0

        results[e.id] = (int(current_break), ovt)
        recorded_pause_distributed += current_break

    total_net = min(max_mins, total_accumulated_gross - recorded_pause_distributed)
    return results, total_net


def get_login_time():
    """Ermittelt die letzte Anmeldezeit des aktuellen Benutzers als QTime.
    Gibt None zurück wenn die Zeit nicht ermittelt werden kann.

    Linux  : journalctl (systemd-logind), Fallback: who
    macOS  : last, Fallback: who
    Windows: PowerShell (Win32_LogonSession)
    """
    _no_win = {"creationflags": subprocess.CREATE_NO_WINDOW} if sys.platform == "win32" else {}

    if sys.platform.startswith("linux"):
        # Primär: journalctl mit --output=short-iso für zuverlässiges Parsing
        try:
            user = getpass.getuser()
            r = subprocess.run(
                ["journalctl", "-u", "systemd-logind",
                 "--grep", f"New session.*{user}",
                 "-n", "1", "--output=short-iso", "--no-pager"],
                capture_output=True, text=True, timeout=5, check=False
            )
            if r.returncode == 0 and r.stdout.strip():
                m = re.search(r"T(\d{2}):(\d{2}):", r.stdout.strip().split("\n")[-1])
                if m:
                    return QTime(int(m.group(1)), int(m.group(2)))
        except (subprocess.SubprocessError, OSError):
            pass

        # Fallback: who
        try:
            user = getpass.getuser()
            r = subprocess.run(["who"], capture_output=True, text=True, timeout=3, check=False)
            if r.returncode == 0:
                for line in r.stdout.strip().splitlines():
                    if line.startswith(user + " ") or line.startswith(user + "\t"):
                        m = re.search(r"(\d{2}:\d{2})", line)
                        if m:
                            return QTime.fromString(m.group(1), "HH:mm")
        except (subprocess.SubprocessError, OSError):
            pass

    elif sys.platform == "darwin":
        # macOS: last -1 <user>
        try:
            user = getpass.getuser()
            r = subprocess.run(["last", "-1", user],
                               capture_output=True, text=True, timeout=5, check=False)
            if r.returncode == 0 and r.stdout:
                m = re.search(r"\s(\d{1,2}:\d{2})\s", r.stdout.split("\n")[0])
                if m:
                    return QTime.fromString(m.group(1).zfill(5), "HH:mm")
        except (subprocess.SubprocessError, OSError):
            pass

        # Fallback: who
        try:
            user = getpass.getuser()
            r = subprocess.run(["who"], capture_output=True, text=True, timeout=3, check=False)
            if r.returncode == 0:
                for line in r.stdout.strip().splitlines():
                    if line.startswith(user):
                        m = re.search(r"(\d{1,2}:\d{2})", line)
                        if m:
                            return QTime.fromString(m.group(1).zfill(5), "HH:mm")
        except (subprocess.SubprocessError, OSError):
            pass

    elif sys.platform == "win32":
        # Windows: PowerShell — neueste interaktive Session (LogonType 2 oder 10)
        try:
            ps_cmd = (
                "(Get-CimInstance Win32_LogonSession | "
                "Where-Object {$_.LogonType -in 2,10} | "
                "Sort-Object StartTime -Descending | "
                "Select-Object -First 1).StartTime.ToString('HH:mm')"
            )
            r = subprocess.run(
                ["powershell", "-NoProfile", "-NonInteractive", "-Command", ps_cmd],
                capture_output=True, text=True, timeout=8, check=False, **_no_win
            )
            if r.returncode == 0 and r.stdout.strip():
                return QTime.fromString(r.stdout.strip(), "HH:mm")
        except (subprocess.SubprocessError, OSError):
            pass

    return None


# --- ItemDelegate für den blauen Rahmen am heutigen Tag ---
class HeatmapDelegate(QStyledItemDelegate):
    """
    Delegate zum Zeichnen eines blauen Rahmens um den heutigen Tag in der Heatmap.
    """
    def paint(self, painter: QPainter, option, index):
        """
        Zeichnet die Zelle und fügt einen Rahmen hinzu, wenn es der heutige Tag ist.
        """
        super().paint(painter, option, index)

        is_today = index.data(Qt.ItemDataRole.UserRole + 1)

        if is_today:
            pen = QPen(QColor("#60a5fa"), 2)  # Blauer Rahmen für den heutigen Tag
            painter.setPen(pen)
            r = option.rect
            painter.drawRect(r.x() + 1, r.y() + 1, r.width() - 2, r.height() - 2)

# --- DIALOGE ---
class SettingsDialog(QDialog):
    """
    Dialog zum Verwalten der Benutzereinstellungen wie Standardarbeitszeiten,
    Bundesland und Sonderarbeitstage.
    """
    def __init__(self, current_settings, parent=None):
        """
        Initialisiert den Einstellungsdialog mit den aktuellen Werten.
        """
        super().__init__(parent)
        self.setWindowTitle("Einstellungen")
        self.resize(380, 450)
        layout = QVBoxLayout(self)

        layout.addWidget(QLabel("<b>Tages-Standardwerte:</b>"))

        self.login_time_cb = QCheckBox("Login-Zeit als Startzeit verwenden")
        self.login_time_cb.setToolTip(
            "Liest beim Programmstart die letzte Anmeldezeit des Benutzers aus.\n"
            "Die Standard-Startzeit dient als Fallback, falls die Anmeldezeit\n"
            "nicht ermittelt werden kann."
        )
        self.login_time_cb.setChecked(current_settings.get("use_login_time", False))
        layout.addWidget(self.login_time_cb)

        layout.addSpacing(10)
        layout.addWidget(QLabel("<b>Arbeitstage (Soll-Tage):</b>"))
        self.workday_checkboxes = []
        workdays_layout = QHBoxLayout()
        days = ["Mo", "Di", "Mi", "Do", "Fr", "Sa", "So"]
        selected_workdays = current_settings.get("workdays", [0, 1, 2, 3, 4])
        for i, day_name in enumerate(days):
            cb = QCheckBox(day_name)
            cb.setChecked(i in selected_workdays)
            workdays_layout.addWidget(cb)
            self.workday_checkboxes.append(cb)
        layout.addLayout(workdays_layout)

        layout.addSpacing(10)
        time_layout = QHBoxLayout()
        self.time_start = QTimeEdit()
        self.time_start.setDisplayFormat("HH:mm")
        default_start = current_settings.get("default_start", "07:00")
        self.time_start.setTime(QTime.fromString(default_start, "HH:mm"))
        self.time_start.setEnabled(not current_settings.get("use_login_time", False))
        self.login_time_cb.stateChanged.connect(
            self._on_login_time_state_changed
        )
        time_layout.addWidget(QLabel("Fallback Startzeit:"))
        time_layout.addWidget(self.time_start)
        layout.addLayout(time_layout)

        target_layout = QHBoxLayout()
        self.time_target = QTimeEdit()
        self.time_target.setDisplayFormat("HH:mm")
        target_work_time = current_settings.get("target_work_time", "08:00")
        self.time_target.setTime(QTime.fromString(target_work_time, "HH:mm"))
        target_layout.addWidget(QLabel("Regelarbeitszeit (Soll):"))
        target_layout.addWidget(self.time_target)
        layout.addLayout(target_layout)

        max_layout = QHBoxLayout()
        self.max_hours_spin = QSpinBox()
        self.max_hours_spin.setRange(6, 24)
        self.max_hours_spin.setSuffix(" h")
        self.max_hours_spin.setValue(current_settings.get("max_work_hours", 10))
        max_layout.addWidget(QLabel("Max. anrechenbare Arbeitszeit:"))
        max_layout.addWidget(self.max_hours_spin)
        layout.addLayout(max_layout)

        layout.addSpacing(10)
        layout.addWidget(QLabel("<b>Pausen-Regelung:</b>"))
        self.auto_break_cb = QCheckBox("Automatische Pausen-Berechnung")
        self.auto_break_cb.setToolTip("6-9h: 30 Min, >9h: 45 Min")
        self.auto_break_cb.setChecked(current_settings.get("auto_break", True))
        layout.addWidget(self.auto_break_cb)

        layout.addSpacing(10)
        layout.addWidget(QLabel("<b>Regionales:</b>"))

        state_layout = QHBoxLayout()
        self.state_combo = QComboBox()
        for code, name in BUNDESLAENDER.items():
            self.state_combo.addItem(name, code)

        curr_state = current_settings.get("state", "TH")
        idx = self.state_combo.findData(curr_state)
        if idx >= 0:
            self.state_combo.setCurrentIndex(idx)

        state_layout.addWidget(QLabel("Bundesland (für Feiertage):"))
        state_layout.addWidget(self.state_combo)
        layout.addLayout(state_layout)

        layout.addSpacing(10)
        layout.addWidget(QLabel("<b>Sonder-Arbeitstage (z.B. 24.12.):</b>"))
        self.special_days_table = QTableWidget(0, 3)
        self.special_days_table.setHorizontalHeaderLabels(["Tag", "Monat", "Soll-Zeit"])
        self.special_days_table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Stretch
        )
        self.special_days_table.setFixedHeight(120)
        layout.addWidget(self.special_days_table)

        special_btn_layout = QHBoxLayout()
        btn_add_special = QPushButton("Hinzufügen")
        btn_add_special.clicked.connect(self.add_special_day_row)
        btn_remove_special = QPushButton("Löschen")
        btn_remove_special.clicked.connect(self.remove_special_day_row)
        special_btn_layout.addWidget(btn_add_special)
        special_btn_layout.addWidget(btn_remove_special)
        layout.addLayout(special_btn_layout)

        # Load special days
        special_days = current_settings.get("special_days", [])
        for sd in special_days:
            self.add_special_day_row(sd)

        layout.addSpacing(10)
        layout.addWidget(QLabel("<b>Darstellung:</b>"))
        self.dark_mode_cb = QCheckBox("Dark Mode aktivieren")
        self.dark_mode_cb.setChecked(current_settings.get("dark_mode", False))
        layout.addWidget(self.dark_mode_cb)

        self.btn_save = QPushButton("Speichern")
        self.btn_save.clicked.connect(self.accept)
        layout.addSpacing(15)
        layout.addWidget(self.btn_save)

    def _on_login_time_state_changed(self):
        """
        Aktiviert oder deaktiviert die Startzeit-Eingabe basierend auf der Login-Zeit-Option.
        """
        self.time_start.setEnabled(not self.login_time_cb.isChecked())

    def add_special_day_row(self, data=None):
        """
        Fügt eine Zeile für einen Sonderarbeitstag hinzu.
        """
        if not isinstance(data, dict):
            data = None  # Handle QPushButton click
        row = self.special_days_table.rowCount()
        self.special_days_table.insertRow(row)

        day_spin = QSpinBox()
        day_spin.setRange(1, 31)
        day_spin.setValue(data["day"] if data else 1)
        self.special_days_table.setCellWidget(row, 0, day_spin)

        month_spin = QSpinBox()
        month_spin.setRange(1, 12)
        month_spin.setValue(data["month"] if data else 1)
        self.special_days_table.setCellWidget(row, 1, month_spin)

        time_edit = QTimeEdit()
        time_edit.setDisplayFormat("HH:mm")
        time_edit.setTime(QTime.fromString(data["target"] if data else "04:00", "HH:mm"))
        self.special_days_table.setCellWidget(row, 2, time_edit)

    def remove_special_day_row(self):
        """
        Entfernt die aktuell ausgewählte Zeile der Sonderarbeitstage.
        """
        curr = self.special_days_table.currentRow()
        if curr >= 0:
            self.special_days_table.removeRow(curr)

    def get_settings(self):
        """
        Gibt die im Dialog eingestellten Werte als Dictionary zurück.
        """
        workdays = [i for i, cb in enumerate(self.workday_checkboxes) if cb.isChecked()]
        special_days = []
        for r in range(self.special_days_table.rowCount()):
            day = self.special_days_table.cellWidget(r, 0).value()
            month = self.special_days_table.cellWidget(r, 1).value()
            target = self.special_days_table.cellWidget(r, 2).time().toString("HH:mm")
            special_days.append({"day": day, "month": month, "target": target})

        return {
            "default_start": self.time_start.time().toString("HH:mm"),
            "target_work_time": self.time_target.time().toString("HH:mm"),
            "state": self.state_combo.currentData(),
            "dark_mode": self.dark_mode_cb.isChecked(),
            "max_work_hours": self.max_hours_spin.value(),
            "auto_break": self.auto_break_cb.isChecked(),
            "use_login_time": self.login_time_cb.isChecked(),
            "workdays": workdays,
            "special_days": special_days,
        }


class EditDialog(QDialog):
    """
    Dialog zum Bearbeiten oder Hinzufügen eines Arbeitseintrags.
    """
    def __init__(self, entry: WorkEntry, all_entries: list, target_minutes: int,
                 max_minutes: int = 600, auto_break: bool = True, parent=None):
        """
        Initialisiert den Bearbeitungsdialog.
        """
        super().__init__(parent)
        self.setWindowTitle("Eintrag bearbeiten")
        self.resize(450, 300)
        self.entry = entry
        self.all_entries = all_entries  # All entries to filter by date
        self.target_minutes = target_minutes
        self.max_minutes = max_minutes
        self.auto_break = auto_break
        layout = QVBoxLayout(self)

        self.date_edit = QDateEdit()
        self.date_edit.setCalendarPopup(True)
        if entry.date:
            self.date_edit.setDate(QDate.fromString(entry.date, "yyyy-MM-dd"))
        layout.addWidget(QLabel("Datum:"))
        layout.addWidget(self.date_edit)

        self.has_times_cb = QCheckBox("Start- und Endzeit verwenden")
        has_times = bool(entry.start and entry.end)
        self.has_times_cb.setChecked(has_times)
        self.has_times_cb.stateChanged.connect(self.toggle_times)
        layout.addWidget(self.has_times_cb)

        time_layout = QHBoxLayout()
        self.time_start = QTimeEdit()
        self.time_start.setDisplayFormat("HH:mm")
        time_layout.addWidget(QLabel("Start:"))
        time_layout.addWidget(self.time_start)

        self.time_end = QTimeEdit()
        self.time_end.setDisplayFormat("HH:mm")
        time_layout.addWidget(QLabel("Ende:"))
        time_layout.addWidget(self.time_end)

        self.pause_spin = QSpinBox()
        self.pause_spin.setRange(0, 300)
        self.pause_spin.setSuffix(" Min")
        self.pause_spin.setValue(entry.pause)
        self.pause_spin.setEnabled(not self.auto_break)
        time_layout.addWidget(QLabel("Pause:"))
        time_layout.addWidget(self.pause_spin)
        layout.addLayout(time_layout)

        if has_times:
            self.time_start.setTime(QTime.fromString(entry.start, "HH:mm"))
            self.time_end.setTime(QTime.fromString(entry.end, "HH:mm"))
        else:
            self.time_start.setTime(QTime(7, 0))
            self.time_end.setTime(QTime(15, 30))
            self.time_start.setEnabled(False)
            self.time_end.setEnabled(False)
            self.pause_spin.setEnabled(False)

        self.time_start.timeChanged.connect(self.recalc_minutes)
        self.time_end.timeChanged.connect(self.recalc_minutes)
        self.pause_spin.valueChanged.connect(self.recalc_minutes)
        self.date_edit.dateChanged.connect(self.recalc_minutes)

        self.min_spinbox = QSpinBox()
        self.min_spinbox.setRange(-2000, 2000)
        self.min_spinbox.setValue(entry.minutes)
        layout.addWidget(QLabel("Minuten (Überstunden):"))
        layout.addWidget(self.min_spinbox)

        self.reason_edit = QLineEdit()
        self.reason_edit.setText(entry.reason)
        layout.addWidget(QLabel("Anlass:"))
        layout.addWidget(self.reason_edit)

        # Custom Target Time for this entry
        self.custom_target_cb = QCheckBox("Individuelles Tagessoll für diesen Tag")
        self.custom_target_cb.setChecked(entry.target_minutes != -1)
        self.custom_target_time = QTimeEdit()
        self.custom_target_time.setDisplayFormat("HH:mm")
        if entry.target_minutes != -1:
            self.custom_target_time.setTime(QTime(entry.target_minutes // 60,
                                                  entry.target_minutes % 60))
        else:
            def_target = self.parent().settings.get("target_work_time", "08:00")
            self.custom_target_time.setTime(QTime.fromString(def_target, "HH:mm"))

        self.custom_target_time.setEnabled(self.custom_target_cb.isChecked())
        self.custom_target_cb.stateChanged.connect(
            self._on_custom_target_changed
        )
        self.custom_target_cb.stateChanged.connect(self.recalc_minutes)
        self.custom_target_time.timeChanged.connect(self.recalc_minutes)

        target_layout = QHBoxLayout()
        target_layout.addWidget(self.custom_target_cb)
        target_layout.addWidget(self.custom_target_time)
        layout.addLayout(target_layout)

        btn_save = QPushButton("Speichern")
        btn_save.clicked.connect(self.accept)
        layout.addWidget(btn_save)

        if has_times:
            self.recalc_minutes()

    def _on_custom_target_changed(self):
        """
        Aktiviert oder deaktiviert das Zeit-Eingabefeld für das individuelle Tagessoll.
        """
        self.custom_target_time.setEnabled(self.custom_target_cb.isChecked())

    def toggle_times(self, state):
        """
        Aktiviert oder deaktiviert die Zeit-Eingabefelder basierend auf der Checkbox.
        """
        is_checked = self.has_times_cb.isChecked()
        self.time_start.setEnabled(is_checked)
        self.time_end.setEnabled(is_checked)
        self.pause_spin.setEnabled(is_checked and not self.auto_break)
        if is_checked:
            self.recalc_minutes()

    def recalc_minutes(self):
        """
        Berechnet die Überstunden automatisch basierend auf Start-, Endzeit und Pause.
        """
        if not self.has_times_cb.isChecked():
            return

        curr_date_str = self.date_edit.date().toString("yyyy-MM-dd")
        target_mins = self.parent().get_target_minutes_for_date(curr_date_str)

        current_temp = WorkEntry(
            id=self.entry.id,
            date=curr_date_str,
            start=self.time_start.time().toString("HH:mm"),
            end=self.time_end.time().toString("HH:mm"),
            pause=self.pause_spin.value() if not self.auto_break else 0,
            minutes=0,
            reason=""
        )

        other_timed = [e for e in self.all_entries
                       if e.date == curr_date_str and e.id != self.entry.id and e.start and e.end]
        results, _ = calculate_timed_entries(other_timed + [current_temp],
                                             target_mins, self.max_minutes, self.auto_break)
        entry_pause, entry_overtime = results[self.entry.id]

        if self.auto_break:
            self.pause_spin.blockSignals(True)
            self.pause_spin.setValue(entry_pause)
            self.pause_spin.blockSignals(False)

        self.min_spinbox.setValue(entry_overtime)

    def apply_to_entry(self):
        """
        Übernimmt die Werte aus dem Dialog in das übergebene WorkEntry-Objekt.
        """
        self.entry.date = self.date_edit.date().toString("yyyy-MM-dd")
        self.entry.minutes = self.min_spinbox.value()
        self.entry.reason = self.reason_edit.text().strip()

        if self.custom_target_cb.isChecked():
            t = self.custom_target_time.time()
            self.entry.target_minutes = t.hour() * 60 + t.minute()
        else:
            self.entry.target_minutes = -1

        if self.has_times_cb.isChecked():
            self.entry.start = self.time_start.time().toString("HH:mm")
            self.entry.end = self.time_end.time().toString("HH:mm")
            self.entry.pause = self.pause_spin.value()
        else:
            self.entry.start = ""
            self.entry.end = ""
            self.entry.pause = 0


# --- HAUPTANWENDUNG ---
class UeberstundenApp(QMainWindow):
    """
    Hauptklasse der Anwendung Überstunden-Rechner Pro.
    Verwaltet das Hauptfenster, die Tabs und die Geschäftslogik.
    """
    def __init__(self):
        """
        Initialisiert die Anwendung, lädt Einstellungen und baut die UI auf.
        """
        super().__init__()
        self.setWindowTitle("Überstunden-Rechner Pro")
        self.resize(1000, 750)

        if os.path.exists(ICON_PATH):
            self.setWindowIcon(QIcon(ICON_PATH))

        self.db = DBManager(DB_FILE)

        self.system_palette = QApplication.instance().palette()
        bg_color = self.system_palette.color(QPalette.ColorRole.Window)
        bg_lightness = bg_color.lightness()
        self.system_is_dark = bg_lightness < 128

        self.settings = self.load_settings()

        self.entries = []
        self.current_calculated_overtime = 0
        self.current_calculated_pause = 0

        self.apply_theme()

        # Daten laden BEVOR die UI sie berechnet
        self.entries = self.db.load_all()

        self.tabs = QTabWidget()
        self.setCentralWidget(self.tabs)

        self.tab_main = QWidget()
        self.tab_goals = QWidget()
        self.tab_calendar = QWidget()
        self.tab_stats = QWidget()

        self.tabs.addTab(self.tab_main, "Eingabe && Liste")
        self.tabs.addTab(self.tab_goals, "Ziele && Dashboard")
        self.tabs.addTab(self.tab_calendar, "Kalender-Heatmap")
        self.tabs.addTab(self.tab_stats, "Diagramm && Statistik")

        self.setup_main_tab()
        self.setup_goals_tab()
        self.setup_calendar_tab()
        self.setup_stats_tab()

        self.load_data()

    def load_settings(self):
        """
        Lädt die Einstellungen aus der JSON-Datei oder gibt Standardwerte zurück.
        """
        defaults = {
            "default_start": "07:00",
            "target_work_time": "08:00",
            "state": "TH",
            "dark_mode": self.system_is_dark,
            "auto_break": True,
            "use_login_time": False,
            "workdays": [0, 1, 2, 3, 4],
            "special_days": [
                {"month": 12, "day": 24, "target": "04:00"},
                {"month": 12, "day": 31, "target": "04:00"}
            ],
            "goal_active": False,
            "goal_start_date": QDate.currentDate().addDays(30).toString("yyyy-MM-dd"),
            "goal_end_date": QDate.currentDate().addDays(35).toString("yyyy-MM-dd"),
            "goal_hours": 0
        }
        if os.path.exists(SETTINGS_FILE):
            try:
                with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
                    defaults.update(json.load(f))
            except (json.JSONDecodeError, OSError):
                pass
        return defaults

    def save_settings(self):
        """
        Speichert die aktuellen Einstellungen in die JSON-Datei.
        """
        try:
            with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
                json.dump(self.settings, f)
        except OSError:
            pass

    def get_light_palette(self):
        """
        Erstellt und gibt die Farbpalette für den hellen Modus zurück (Breeze Light).
        """
        palette = QPalette()
        palette.setColor(QPalette.ColorRole.Window,          QColor("#eff0f1"))
        palette.setColor(QPalette.ColorRole.WindowText,      QColor("#31363b"))
        palette.setColor(QPalette.ColorRole.Base,            QColor("#fcfcfc"))
        palette.setColor(QPalette.ColorRole.AlternateBase,   QColor("#eff0f1"))
        palette.setColor(QPalette.ColorRole.ToolTipBase,     QColor("#31363b"))
        palette.setColor(QPalette.ColorRole.ToolTipText,     QColor("#eff0f1"))
        palette.setColor(QPalette.ColorRole.Text,            QColor("#31363b"))
        palette.setColor(QPalette.ColorRole.Button,          QColor("#eff0f1"))
        palette.setColor(QPalette.ColorRole.ButtonText,      QColor("#31363b"))
        palette.setColor(QPalette.ColorRole.BrightText,      QColor("#ff0000"))
        palette.setColor(QPalette.ColorRole.Highlight,       QColor("#3daee9"))
        palette.setColor(QPalette.ColorRole.HighlightedText, QColor("#ffffff"))
        palette.setColor(QPalette.ColorRole.Link,            QColor("#3daee9"))
        return palette

    def get_dark_palette(self):
        """
        Erstellt und gibt die Farbpalette für den dunklen Modus zurück (Breeze Dark).
        """
        palette = QPalette()
        palette.setColor(QPalette.ColorRole.Window,          QColor("#31363b"))
        palette.setColor(QPalette.ColorRole.WindowText,      QColor("#eff0f1"))
        palette.setColor(QPalette.ColorRole.Base,            QColor("#232629"))
        palette.setColor(QPalette.ColorRole.AlternateBase,   QColor("#31363b"))
        palette.setColor(QPalette.ColorRole.ToolTipBase,     QColor("#31363b"))
        palette.setColor(QPalette.ColorRole.ToolTipText,     QColor("#eff0f1"))
        palette.setColor(QPalette.ColorRole.Text,            QColor("#eff0f1"))
        palette.setColor(QPalette.ColorRole.Button,          QColor("#31363b"))
        palette.setColor(QPalette.ColorRole.ButtonText,      QColor("#eff0f1"))
        palette.setColor(QPalette.ColorRole.BrightText,      QColor("#ff5555"))
        palette.setColor(QPalette.ColorRole.Highlight,       QColor("#3daee9"))
        palette.setColor(QPalette.ColorRole.HighlightedText, QColor("#ffffff"))
        palette.setColor(QPalette.ColorRole.Link,            QColor("#3daee9"))
        return palette

    def get_dark_stylesheet(self):
        """
        Gibt das CSS-Stylesheet für den Breeze Dark Modus zurück.
        """
        return """
            * { font-size: 13px; }

            QMainWindow, QDialog { background-color: #31363b; }

            QWidget { background-color: #31363b; color: #eff0f1; }

            QPushButton {
                background-color: #3b4045;
                color: #eff0f1;
                border: 1px solid #4a5056;
                border-radius: 3px;
                padding: 4px 16px;
                min-height: 24px;
            }
            QPushButton:hover  { background-color: #4d5560; border-color: #3daee9; }
            QPushButton:pressed { background-color: #3daee9; color: #fff; border-color: #3daee9; }
            QPushButton:disabled { color: #6c7176; border-color: #3a3f44; }

            QLineEdit, QComboBox {
                background-color: #232629;
                color: #eff0f1;
                border: 1px solid #4a5056;
                border-radius: 3px;
                padding: 3px 6px;
                min-height: 24px;
                selection-background-color: #3daee9;
            }
            QLineEdit:focus, QComboBox:focus { border-color: #3daee9; }

            QAbstractSpinBox {
                background-color: #232629;
                color: #eff0f1;
                border: 1px solid #4a5056;
                border-radius: 3px;
                padding: 3px 6px;
                min-height: 24px;
            }
            QAbstractSpinBox:focus { border-color: #3daee9; }
            QAbstractSpinBox::up-button {
                subcontrol-origin: border;
                subcontrol-position: top right;
                width: 20px;
                background-color: #3b4045;
                border-left: 1px solid #4a5056;
                border-bottom: 1px solid #4a5056;
                border-top-right-radius: 3px;
            }
            QAbstractSpinBox::down-button {
                subcontrol-origin: border;
                subcontrol-position: bottom right;
                width: 20px;
                background-color: #3b4045;
                border-left: 1px solid #4a5056;
                border-bottom-right-radius: 3px;
            }
            QAbstractSpinBox::up-button:hover,
            QAbstractSpinBox::down-button:hover { background-color: #3daee9; }

            QComboBox::drop-down { border: none; width: 22px; }
            QComboBox::down-arrow { width: 10px; height: 10px; }
            QComboBox QAbstractItemView {
                background-color: #232629;
                color: #eff0f1;
                border: 1px solid #4a5056;
                selection-background-color: #3daee9;
                selection-color: #fff;
                outline: none;
            }

            QTabWidget::pane { border: 1px solid #4a5056; top: -1px; }
            QTabBar::tab {
                background-color: #2e3338;
                color: #7f8c8d;
                border: 1px solid #4a5056;
                border-bottom: none;
                border-top-left-radius: 3px;
                border-top-right-radius: 3px;
                padding: 5px 14px;
                margin-right: 2px;
            }
            QTabBar::tab:selected    { background-color: #31363b; color: #eff0f1; border-bottom: 2px solid #3daee9; }
            QTabBar::tab:hover:!selected { background-color: #3b4045; color: #eff0f1; }

            QTableWidget {
                background-color: #232629;
                alternate-background-color: #2a2f34;
                color: #eff0f1;
                gridline-color: #3a3f44;
                border: 1px solid #4a5056;
            }
            QTableWidget::item:selected { background-color: #3daee9; color: #fff; }
            QHeaderView::section {
                background-color: #31363b;
                color: #eff0f1;
                border: none;
                border-right: 1px solid #4a5056;
                border-bottom: 1px solid #4a5056;
                padding: 4px 8px;
                font-weight: bold;
            }
            QHeaderView::section:first { border-left: none; }

            QGroupBox {
                border: 1px solid #4a5056;
                border-radius: 3px;
                margin-top: 12px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top left;
                left: 8px;
                color: #3daee9;
                font-weight: bold;
            }

            QCheckBox::indicator {
                width: 16px; height: 16px;
                border: 1px solid #4a5056;
                border-radius: 2px;
                background-color: #232629;
            }
            QCheckBox::indicator:checked  { background-color: #3daee9; border-color: #3daee9; }
            QCheckBox::indicator:hover    { border-color: #3daee9; }

            QScrollBar:vertical { background: #2e3338; width: 8px; border-radius: 4px; margin: 0; }
            QScrollBar::handle:vertical { background: #4a5056; border-radius: 4px; min-height: 24px; }
            QScrollBar::handle:vertical:hover { background: #3daee9; }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height: 0; }
            QScrollBar:horizontal { background: #2e3338; height: 8px; border-radius: 4px; margin: 0; }
            QScrollBar::handle:horizontal { background: #4a5056; border-radius: 4px; min-width: 24px; }
            QScrollBar::handle:horizontal:hover { background: #3daee9; }
            QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal { width: 0; }

            QProgressBar {
                border: 1px solid #4a5056;
                border-radius: 3px;
                background-color: #232629;
                text-align: center;
                color: #eff0f1;
                min-height: 16px;
            }
            QProgressBar::chunk { background-color: #3daee9; border-radius: 2px; }

            QLabel { background: transparent; }
            QFrame { background: transparent; }
            QToolTip { background-color: #31363b; color: #eff0f1; border: 1px solid #4a5056; }
            QMessageBox { background-color: #31363b; }

            /* Aufklappbarer Kalender (QDateEdit-Popup) */
            QCalendarWidget {
                background-color: #31363b;
                color: #eff0f1;
            }
            QCalendarWidget QWidget#qt_calendar_navigationbar {
                background-color: #232629;
                padding: 2px;
            }
            QCalendarWidget QToolButton {
                background-color: #3b4045;
                color: #eff0f1;
                border: 1px solid #4a5056;
                border-radius: 3px;
                padding: 3px 8px;
                margin: 2px;
            }
            QCalendarWidget QToolButton:hover  { background-color: #3daee9; border-color: #3daee9; }
            QCalendarWidget QToolButton:pressed { background-color: #2980b9; }
            QCalendarWidget QToolButton::menu-indicator { image: none; }
            QCalendarWidget QSpinBox {
                background-color: #232629;
                color: #eff0f1;
                border: 1px solid #4a5056;
                border-radius: 3px;
                padding: 2px 4px;
                selection-background-color: #3daee9;
            }
            QCalendarWidget QAbstractItemView:enabled {
                background-color: #232629;
                color: #eff0f1;
                selection-background-color: #3daee9;
                selection-color: #ffffff;
                gridline-color: #3a3f44;
            }
            QCalendarWidget QAbstractItemView:disabled { color: #4a5056; }
            QCalendarWidget QMenu {
                background-color: #232629;
                color: #eff0f1;
                border: 1px solid #4a5056;
            }
            QCalendarWidget QMenu::item:selected { background-color: #3daee9; color: #fff; }
        """ + (f"""
            QAbstractSpinBox::up-arrow   {{ image: url("{icon_dir}/arrow_up.png");   width: 10px; height: 10px; }}
            QAbstractSpinBox::down-arrow {{ image: url("{icon_dir}/arrow_down.png"); width: 10px; height: 10px; }}
        """ if icon_dir else "")

    def get_light_stylesheet(self, icon_dir=''):
        # Breeze Light
        return """
            * { font-size: 13px; }

            QMainWindow, QDialog { background-color: #eff0f1; }

            QWidget { background-color: #eff0f1; color: #31363b; }

            QPushButton {
                background-color: #e3e5e7;
                color: #31363b;
                border: 1px solid #bdc3c7;
                border-radius: 3px;
                padding: 4px 16px;
                min-height: 24px;
            }
            QPushButton:hover  { background-color: #d6eaf8; border-color: #3daee9; }
            QPushButton:pressed { background-color: #3daee9; color: #fff; border-color: #3daee9; }
            QPushButton:disabled { color: #95a5a6; border-color: #bdc3c7; }

            QLineEdit, QComboBox {
                background-color: #fcfcfc;
                color: #31363b;
                border: 1px solid #bdc3c7;
                border-radius: 3px;
                padding: 3px 6px;
                min-height: 24px;
                selection-background-color: #3daee9;
                selection-color: #fff;
            }
            QLineEdit:focus, QComboBox:focus { border-color: #3daee9; }

            QAbstractSpinBox {
                background-color: #fcfcfc;
                color: #31363b;
                border: 1px solid #bdc3c7;
                border-radius: 3px;
                padding: 3px 6px;
                min-height: 24px;
            }
            QAbstractSpinBox:focus { border-color: #3daee9; }
            QAbstractSpinBox::up-button {
                subcontrol-origin: border;
                subcontrol-position: top right;
                width: 20px;
                background-color: #e3e5e7;
                border-left: 1px solid #bdc3c7;
                border-bottom: 1px solid #bdc3c7;
                border-top-right-radius: 3px;
            }
            QAbstractSpinBox::down-button {
                subcontrol-origin: border;
                subcontrol-position: bottom right;
                width: 20px;
                background-color: #e3e5e7;
                border-left: 1px solid #bdc3c7;
                border-bottom-right-radius: 3px;
            }
            QAbstractSpinBox::up-button:hover,
            QAbstractSpinBox::down-button:hover { background-color: #3daee9; }

            QComboBox::drop-down { border: none; width: 22px; }
            QComboBox::down-arrow { width: 10px; height: 10px; }
            QComboBox QAbstractItemView {
                background-color: #fcfcfc;
                color: #31363b;
                border: 1px solid #bdc3c7;
                selection-background-color: #3daee9;
                selection-color: #fff;
                outline: none;
            }

            QTabWidget::pane { border: 1px solid #bdc3c7; top: -1px; }
            QTabBar::tab {
                background-color: #dde1e3;
                color: #7f8c8d;
                border: 1px solid #bdc3c7;
                border-bottom: none;
                border-top-left-radius: 3px;
                border-top-right-radius: 3px;
                padding: 5px 14px;
                margin-right: 2px;
            }
            QTabBar::tab:selected    { background-color: #eff0f1; color: #31363b; border-bottom: 2px solid #3daee9; }
            QTabBar::tab:hover:!selected { background-color: #e8eaeb; color: #31363b; }

            QTableWidget {
                background-color: #fcfcfc;
                alternate-background-color: #f4f5f5;
                color: #31363b;
                gridline-color: #d5d8db;
                border: 1px solid #bdc3c7;
            }
            QTableWidget::item:selected { background-color: #3daee9; color: #fff; }
            QHeaderView::section {
                background-color: #e8eaeb;
                color: #31363b;
                border: none;
                border-right: 1px solid #bdc3c7;
                border-bottom: 1px solid #bdc3c7;
                padding: 4px 8px;
                font-weight: bold;
            }

            QGroupBox {
                border: 1px solid #bdc3c7;
                border-radius: 3px;
                margin-top: 12px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top left;
                left: 8px;
                color: #3daee9;
                font-weight: bold;
            }

            QCheckBox::indicator {
                width: 16px; height: 16px;
                border: 1px solid #bdc3c7;
                border-radius: 2px;
                background-color: #fcfcfc;
            }
            QCheckBox::indicator:checked { background-color: #3daee9; border-color: #3daee9; }
            QCheckBox::indicator:hover   { border-color: #3daee9; }

            QScrollBar:vertical { background: #dde1e3; width: 8px; border-radius: 4px; margin: 0; }
            QScrollBar::handle:vertical { background: #bdc3c7; border-radius: 4px; min-height: 24px; }
            QScrollBar::handle:vertical:hover { background: #3daee9; }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height: 0; }
            QScrollBar:horizontal { background: #dde1e3; height: 8px; border-radius: 4px; margin: 0; }
            QScrollBar::handle:horizontal { background: #bdc3c7; border-radius: 4px; min-width: 24px; }
            QScrollBar::handle:horizontal:hover { background: #3daee9; }
            QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal { width: 0; }

            QProgressBar {
                border: 1px solid #bdc3c7;
                border-radius: 3px;
                background-color: #dde1e3;
                text-align: center;
                color: #31363b;
                min-height: 16px;
            }
            QProgressBar::chunk { background-color: #3daee9; border-radius: 2px; }

            QLabel { background: transparent; }
            QFrame { background: transparent; }
            QToolTip { background-color: #31363b; color: #eff0f1; border: 1px solid #4a5056; }

            /* Aufklappbarer Kalender (QDateEdit-Popup) */
            QCalendarWidget {
                background-color: #eff0f1;
                color: #31363b;
            }
            QCalendarWidget QWidget#qt_calendar_navigationbar {
                background-color: #dde1e3;
                padding: 2px;
            }
            QCalendarWidget QToolButton {
                background-color: #e3e5e7;
                color: #31363b;
                border: 1px solid #bdc3c7;
                border-radius: 3px;
                padding: 3px 8px;
                margin: 2px;
            }
            QCalendarWidget QToolButton:hover  { background-color: #3daee9; color: #fff; border-color: #3daee9; }
            QCalendarWidget QToolButton:pressed { background-color: #2980b9; color: #fff; }
            QCalendarWidget QToolButton::menu-indicator { image: none; }
            QCalendarWidget QSpinBox {
                background-color: #fcfcfc;
                color: #31363b;
                border: 1px solid #bdc3c7;
                border-radius: 3px;
                padding: 2px 4px;
                selection-background-color: #3daee9;
                selection-color: #fff;
            }
            QCalendarWidget QAbstractItemView:enabled {
                background-color: #fcfcfc;
                color: #31363b;
                selection-background-color: #3daee9;
                selection-color: #ffffff;
                gridline-color: #d5d8db;
            }
            QCalendarWidget QAbstractItemView:disabled { color: #95a5a6; }
            QCalendarWidget QMenu {
                background-color: #fcfcfc;
                color: #31363b;
                border: 1px solid #bdc3c7;
            }
            QCalendarWidget QMenu::item:selected { background-color: #3daee9; color: #fff; }
        """ + (f"""
            QAbstractSpinBox::up-arrow   {{ image: url("{icon_dir}/arrow_up.png");   width: 10px; height: 10px; }}
            QAbstractSpinBox::down-arrow {{ image: url("{icon_dir}/arrow_down.png"); width: 10px; height: 10px; }}
        """ if icon_dir else "")

    def _create_arrow_icons(self):
        """Zeichnet kleine Dreieck-Pfeile als PNG und speichert sie für die QSS-Nutzung."""
        icon_dir = os.path.join(BASE_DIR, '.ui_icons')
        os.makedirs(icon_dir, exist_ok=True)
        is_dark = self.settings.get("dark_mode", False)
        color = QColor("#eff0f1") if is_dark else QColor("#31363b")
        arrows = {
            'up':   [QPoint(5, 1), QPoint(10, 9), QPoint(0, 9)],
            'down': [QPoint(0, 1), QPoint(10, 1), QPoint(5, 9)],
        }
        for name, pts in arrows.items():
            px = QPixmap(10, 10)
            px.fill(Qt.GlobalColor.transparent)
            p = QPainter(px)
            p.setRenderHint(QPainter.RenderHint.Antialiasing)
            p.setBrush(color)
            p.setPen(Qt.PenStyle.NoPen)
            p.drawPolygon(QPolygon(pts))
            p.end()
            px.save(os.path.join(icon_dir, f'arrow_{name}.png'))
        # Qt QSS benötigt Forward-Slashes, auch auf Windows
        return icon_dir.replace('\\', '/')

    def apply_theme(self):
        app = QApplication.instance()
        is_dark = self.settings.get("dark_mode", False)

        # Auf Linux im Entwicklungsmodus: natives Breeze-Theme nutzen
        if not getattr(sys, 'frozen', False) and sys.platform.startswith('linux'):
            app.setStyle("Breeze")
            app.setPalette(self.get_dark_palette() if is_dark else self.get_light_palette())
            return

        # Alle anderen Fälle (kompiliert oder Windows/macOS): Fusion + modernes Stylesheet
        app.setStyle("Fusion")
        app.setPalette(self.get_dark_palette() if is_dark else self.get_light_palette())
        icon_dir = self._create_arrow_icons()
        sheet = self.get_dark_stylesheet(icon_dir) if is_dark else self.get_light_stylesheet(icon_dir)
        app.setStyleSheet(sheet)

    def closeEvent(self, event):
        """
        Wird beim Schließen der Anwendung aufgerufen. Speichert Einstellungen und schließt die DB.
        """
        self.save_settings()
        self.db.close()
        super().closeEvent(event)

    def load_data(self):
        """
        Lädt alle Daten aus der Datenbank und aktualisiert die Benutzeroberfläche.
        """
        self.entries = self.db.load_all()
        self.update_ui()
        self.update_stats_chart()
        self.on_goal_changed()
        self.update_calendar_heatmap()

    def get_target_minutes(self):
        """
        Gibt die in den Einstellungen festgelegte Regelarbeitszeit in Minuten zurück.
        """
        t = QTime.fromString(self.settings.get("target_work_time", "08:00"), "HH:mm")
        return t.hour() * 60 + t.minute()

    def get_target_minutes_for_date(self, date_str):
        """
        Ermittelt das Tagessoll für ein bestimmtes Datum unter Berücksichtigung von
        individuellen Einträgen, Sonderarbeitstagen, Feiertagen und Wochenenden.
        """
        # 1. Check if any entry for this day has a custom target_minutes
        for e in self.entries:
            if e.date == date_str and e.target_minutes != -1:
                return e.target_minutes

        qdate = QDate.fromString(date_str, "yyyy-MM-dd")

        # 2. Check special days from settings (e.g. 24.12.)
        special_days = self.settings.get("special_days", [])
        for sd in special_days:
            if qdate.month() == sd["month"] and qdate.day() == sd["day"]:
                t = QTime.fromString(sd["target"], "HH:mm")
                return t.hour() * 60 + t.minute()

        year = qdate.year()
        state = self.settings.get("state", "TH")
        holidays = get_holidays(year, state)

        # Check if holiday
        if date_str in holidays:
            return 0

        # Check if workday (0=Mon, 6=Sun)
        # QDate.dayOfWeek() returns 1 (Mon) to 7 (Sun)
        day_idx = qdate.dayOfWeek() - 1
        workdays = self.settings.get("workdays", [0, 1, 2, 3, 4])
        if day_idx not in workdays:
            return 0

        return self.get_target_minutes()

    def get_max_minutes(self):
        """
        Gibt die maximal anrechenbare Arbeitszeit pro Tag in Minuten zurück.
        """
        return self.settings.get("max_work_hours", 10) * 60

    def _get_default_start_time(self):
        """Ermittelt die Startzeit für die Eingabemaske.
        Wenn Login-Zeit aktiv: Login-Zeit → default_start (Fallback).
        Sonst: last_start (heute) → default_start."""
        if self.settings.get("use_login_time", False):
            t = get_login_time()
            if t and t.isValid():
                return t
            return QTime.fromString(self.settings.get("default_start", "07:00"), "HH:mm")
        today = QDate.currentDate().toString("yyyy-MM-dd")
        if self.settings.get("last_date") == today and self.settings.get("last_start"):
            return QTime.fromString(self.settings["last_start"], "HH:mm")
        return QTime.fromString(self.settings.get("default_start", "07:00"), "HH:mm")

    # --- SETUP TABS ---
    def setup_main_tab(self):
        """
        Erstellt das Layout und die Steuerelemente für den Haupt-Tab (Eingabe & Liste).
        """
        layout = QVBoxLayout(self.tab_main)

        self.lbl_saldo = QLabel("0h 0m")
        self.lbl_saldo.setAlignment(Qt.AlignmentFlag.AlignCenter)
        font = QFont()
        font.setPointSize(28)
        font.setBold(True)
        self.lbl_saldo.setFont(font)
        layout.addWidget(QLabel("<b>Gesamt-Saldo:</b>",
                                alignment=Qt.AlignmentFlag.AlignCenter))
        layout.addWidget(self.lbl_saldo)

        frame_input = QFrame()
        frame_layout = QVBoxLayout(frame_input)

        input_row1 = QHBoxLayout()
        self.date_edit = QDateEdit()
        self.date_edit.setCalendarPopup(True)
        self.date_edit.setDate(QDate.currentDate())
        self.date_edit.dateChanged.connect(self.update_live_calc)
        input_row1.addWidget(QLabel("Datum:"))
        input_row1.addWidget(self.date_edit)

        start_to_use = self._get_default_start_time()

        self.time_start = QTimeEdit()
        self.time_start.setDisplayFormat("HH:mm")
        self.time_start.setTime(start_to_use)
        input_row1.addWidget(QLabel("Start:"))
        input_row1.addWidget(self.time_start)
        btn_now_start = QPushButton("Jetzt")
        btn_now_start.setFixedWidth(50)
        btn_now_start.setToolTip("Aktuelle Uhrzeit als Startzeit setzen")
        btn_now_start.clicked.connect(lambda: self.time_start.setTime(QTime.currentTime()))
        input_row1.addWidget(btn_now_start)

        self.time_end = QTimeEdit()
        self.time_end.setDisplayFormat("HH:mm")
        input_row1.addWidget(QLabel("Ende:"))
        input_row1.addWidget(self.time_end)
        btn_now_end = QPushButton("Jetzt")
        btn_now_end.setFixedWidth(50)
        btn_now_end.setToolTip("Aktuelle Uhrzeit als Endzeit setzen")
        btn_now_end.clicked.connect(self.set_now_as_end)
        input_row1.addWidget(btn_now_end)

        self.pause_spin = QSpinBox()
        self.pause_spin.setRange(0, 300)
        self.pause_spin.setSuffix(" Min")
        self.pause_spin.setEnabled(not self.settings.get("auto_break", True))
        input_row1.addWidget(QLabel("Pause:"))
        input_row1.addWidget(self.pause_spin)

        self.reason_edit = QLineEdit()
        self.reason_edit.setPlaceholderText("z.B. Regulär")
        input_row1.addWidget(QLabel("Anlass:"))
        input_row1.addWidget(self.reason_edit)

        btn_add = QPushButton("Eintragen")
        btn_add.clicked.connect(self.add_entry)
        input_row1.addWidget(btn_add)

        frame_layout.addLayout(input_row1)

        input_row2 = QHBoxLayout()
        self.custom_target_cb = QCheckBox("Indiv. Tagessoll:")
        self.custom_target_time = QTimeEdit()
        self.custom_target_time.setDisplayFormat("HH:mm")
        def_target = self.settings.get("target_work_time", "08:00")
        self.custom_target_time.setTime(QTime.fromString(def_target, "HH:mm"))
        self.custom_target_time.setEnabled(False)
        self.custom_target_cb.stateChanged.connect(
            lambda: self.custom_target_time.setEnabled(self.custom_target_cb.isChecked())
        )
        self.custom_target_cb.stateChanged.connect(self.update_live_calc)
        self.custom_target_time.timeChanged.connect(self.update_live_calc)

        input_row2.addWidget(self.custom_target_cb)
        input_row2.addWidget(self.custom_target_time)
        input_row2.addStretch()
        frame_layout.addLayout(input_row2)

        self.lbl_live_calc = QLabel("Berechne...")
        self.lbl_live_calc.setAlignment(Qt.AlignmentFlag.AlignCenter)
        frame_layout.addWidget(self.lbl_live_calc)
        layout.addWidget(frame_input)

        self.time_start.timeChanged.connect(self.on_start_time_changed)
        self.time_end.timeChanged.connect(self.update_live_calc)
        self.pause_spin.valueChanged.connect(self.update_live_calc)

        toolbar_layout = QHBoxLayout()
        self.month_filter = QComboBox()
        self.month_filter.addItem("Alle", "ALL")
        self.month_filter.currentIndexChanged.connect(self.on_list_filter_changed)
        toolbar_layout.addWidget(QLabel("Filter:"))
        toolbar_layout.addWidget(self.month_filter)
        toolbar_layout.addStretch()

        btn_import = QPushButton("CSV Import")
        btn_import.clicked.connect(self.import_csv)
        toolbar_layout.addWidget(btn_import)

        btn_export = QPushButton("Export")
        export_menu = QMenu(self)
        export_menu.addAction("CSV  (.csv)",  self.export_csv)
        export_menu.addAction("Excel (.xlsx)", self.export_xlsx)
        export_menu.addAction("PDF  (.pdf)",  self.export_pdf)
        btn_export.setMenu(export_menu)
        toolbar_layout.addWidget(btn_export)

        btn_settings = QPushButton("Einstellungen")
        btn_settings.clicked.connect(self.open_settings)
        toolbar_layout.addWidget(btn_settings)
        layout.addLayout(toolbar_layout)

        self.table = QTableWidget(0, 5)
        self.table.setHorizontalHeaderLabels(["Datum", "Zeitraum", "Überstunden", "Anlass", "Aktion"])
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        self.table.setColumnWidth(1, 160)

        self.table.cellDoubleClicked.connect(self.edit_entry)
        layout.addWidget(self.table)

        self.on_start_time_changed(start_to_use)

    def setup_goals_tab(self):
        layout = QVBoxLayout(self.tab_goals)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(20)

        # 1. Konfigurations-Bereich
        settings_group = QGroupBox("Zeitraum und Überstunden-Ziel konfigurieren")
        settings_layout = QVBoxLayout(settings_group)

        goal_header_layout = QHBoxLayout()
        self.goal_active_cb = QCheckBox("Gleitzeit-Ziel aktivieren (Urlaubs-Sparer)")
        self.goal_active_cb.setChecked(self.settings.get("goal_active", False))
        self.goal_active_cb.stateChanged.connect(self.on_goal_changed)
        goal_header_layout.addWidget(self.goal_active_cb)
        goal_header_layout.addStretch()
        settings_layout.addLayout(goal_header_layout)

        goal_inputs_layout = QHBoxLayout()

        goal_inputs_layout.addWidget(QLabel("Urlaub / Frei von:"))
        self.goal_start_edit = QDateEdit()
        self.goal_start_edit.setCalendarPopup(True)
        self.goal_start_edit.setDate(QDate.fromString(self.settings.get("goal_start_date", ""), "yyyy-MM-dd"))
        self.goal_start_edit.dateChanged.connect(self.auto_calculate_goal_hours)
        goal_inputs_layout.addWidget(self.goal_start_edit)

        goal_inputs_layout.addWidget(QLabel("bis:"))
        self.goal_end_edit = QDateEdit()
        self.goal_end_edit.setCalendarPopup(True)
        self.goal_end_edit.setDate(QDate.fromString(self.settings.get("goal_end_date", ""), "yyyy-MM-dd"))
        self.goal_end_edit.dateChanged.connect(self.auto_calculate_goal_hours)
        goal_inputs_layout.addWidget(self.goal_end_edit)

        goal_inputs_layout.addWidget(QLabel(" | Benötigte Überstunden:"))
        self.goal_hours_spin = QSpinBox()
        self.goal_hours_spin.setRange(0, 5000)
        self.goal_hours_spin.setValue(self.settings.get("goal_hours", 0))
        self.goal_hours_spin.setSuffix(" h")
        self.goal_hours_spin.valueChanged.connect(self.on_goal_changed)
        goal_inputs_layout.addWidget(self.goal_hours_spin)

        goal_inputs_layout.addStretch()
        settings_layout.addLayout(goal_inputs_layout)

        layout.addWidget(settings_group)

        # 2. Dashboard-Bereich
        self.dashboard_group = QGroupBox("Fortschritts-Dashboard")
        dashboard_layout = QVBoxLayout(self.dashboard_group)

        self.goal_progress_bar = QProgressBar()
        self.goal_progress_bar.setRange(0, 100)
        self.goal_progress_bar.setValue(0)
        self.goal_progress_bar.setFixedHeight(30)
        self.goal_progress_bar.setTextVisible(True)
        self.goal_progress_bar.setAlignment(Qt.AlignmentFlag.AlignCenter)
        dashboard_layout.addWidget(self.goal_progress_bar)

        grid = QGridLayout()
        grid.setSpacing(15)

        # Kachel: Aktueller Stand
        self.lbl_goal_current = QLabel("0h 0m")
        self.lbl_goal_current.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_goal_current.setFont(QFont("Arial", 20, QFont.Weight.Bold))
        grid.addWidget(QLabel("Aktueller Stand", alignment=Qt.AlignmentFlag.AlignCenter), 0, 0)
        grid.addWidget(self.lbl_goal_current, 1, 0)

        # Kachel: Es fehlen noch
        self.lbl_goal_missing = QLabel("0h 0m")
        self.lbl_goal_missing.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_goal_missing.setFont(QFont("Arial", 20, QFont.Weight.Bold))
        grid.addWidget(QLabel("Es fehlen noch", alignment=Qt.AlignmentFlag.AlignCenter), 0, 1)
        grid.addWidget(self.lbl_goal_missing, 1, 1)

        # Kachel: Verbleibende Tage (bis zum Start des Zeitraums)
        self.lbl_goal_days = QLabel("0")
        self.lbl_goal_days.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_goal_days.setFont(QFont("Arial", 20, QFont.Weight.Bold))
        grid.addWidget(QLabel("Arbeitstage zum Ansparen", alignment=Qt.AlignmentFlag.AlignCenter), 0, 2)
        grid.addWidget(self.lbl_goal_days, 1, 2)

        dashboard_layout.addLayout(grid)

        # Fazit-Label unten
        self.lbl_goal_action = QLabel("-")
        self.lbl_goal_action.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_goal_action.setFont(QFont("Arial", 12))
        dashboard_layout.addSpacing(15)
        dashboard_layout.addWidget(self.lbl_goal_action)

        layout.addWidget(self.dashboard_group)
        layout.addStretch()

        if self.settings.get("goal_hours", 0) == 0:
            self.auto_calculate_goal_hours()

    def setup_calendar_tab(self):
        layout = QVBoxLayout(self.tab_calendar)

        cal_toolbar = QHBoxLayout()

        self.btn_cal_prev = QPushButton("< Vorheriger")
        self.btn_cal_prev.clicked.connect(self.cal_go_prev_month)
        self.btn_cal_next = QPushButton("Nächster >")
        self.btn_cal_next.clicked.connect(self.cal_go_next_month)

        self.cal_month_filter = QComboBox()
        self.cal_month_filter.currentIndexChanged.connect(self.on_cal_filter_changed)

        cal_toolbar.addWidget(QLabel("Monat:"))
        cal_toolbar.addWidget(self.btn_cal_prev)
        cal_toolbar.addWidget(self.cal_month_filter)
        cal_toolbar.addWidget(self.btn_cal_next)

        cal_toolbar.addStretch()

        self.lbl_cal_month_sum = QLabel("Monats-Saldo: 0h 0m")
        font = QFont()
        font.setPointSize(12)
        font.setBold(True)
        self.lbl_cal_month_sum.setFont(font)
        cal_toolbar.addWidget(self.lbl_cal_month_sum)

        layout.addLayout(cal_toolbar)

        self.cal_table = QTableWidget(6, 7)

        self.cal_table.setHorizontalHeaderLabels(["Montag", "Dienstag", "Mittwoch", "Donnerstag", "Freitag", "Samstag", "Sonntag"])
        self.cal_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.cal_table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.cal_table.verticalHeader().hide()
        self.cal_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.cal_table.setSelectionMode(QTableWidget.SelectionMode.NoSelection)
        self.heatmap_delegate = HeatmapDelegate(self.cal_table)
        self.cal_table.setItemDelegate(self.heatmap_delegate)
        layout.addWidget(self.cal_table)

    def cal_go_prev_month(self):
        idx = self.cal_month_filter.currentIndex()
        if idx < self.cal_month_filter.count() - 1:
            self.cal_month_filter.setCurrentIndex(idx + 1)

    def cal_go_next_month(self):
        idx = self.cal_month_filter.currentIndex()
        if idx > 0:
            self.cal_month_filter.setCurrentIndex(idx - 1)

    def setup_stats_tab(self):
        self.stats_layout = QVBoxLayout(self.tab_stats)
        self.figure = Figure(figsize=(8, 4))
        self.canvas = FigureCanvasQTAgg(self.figure)
        self.stats_layout.addWidget(self.canvas)

    # --- ZIELE & KALENDER LOGIK ---
    def auto_calculate_goal_hours(self):
        start_d = self.goal_start_edit.date()
        end_d = self.goal_end_edit.date()
        if start_d > end_d: return

        total_target_mins = 0
        curr = start_d

        while curr <= end_d:
            total_target_mins += self.get_target_minutes_for_date(curr.toString("yyyy-MM-dd"))
            curr = curr.addDays(1)

        total_hours_needed = total_target_mins / 60.0

        self.goal_hours_spin.blockSignals(True)
        self.goal_hours_spin.setValue(math.ceil(total_hours_needed))
        self.goal_hours_spin.blockSignals(False)

        self.on_goal_changed()

    def on_goal_changed(self):
        self.settings["goal_active"] = self.goal_active_cb.isChecked()
        self.settings["goal_start_date"] = self.goal_start_edit.date().toString("yyyy-MM-dd")
        self.settings["goal_end_date"] = self.goal_end_edit.date().toString("yyyy-MM-dd")
        self.settings["goal_hours"] = self.goal_hours_spin.value()
        self.save_settings()

        self.goal_start_edit.setEnabled(self.goal_active_cb.isChecked())
        self.goal_end_edit.setEnabled(self.goal_active_cb.isChecked())
        self.goal_hours_spin.setEnabled(self.goal_active_cb.isChecked())
        self.dashboard_group.setVisible(self.goal_active_cb.isChecked())

        if self.goal_active_cb.isChecked():
            self.update_goal_status()

    def update_goal_status(self):
        if not self.goal_active_cb.isChecked(): return

        target_start_date = self.goal_start_edit.date()
        target_mins = self.goal_hours_spin.value() * 60
        current_saldo = sum(e.minutes for e in self.entries)

        progress_saldo = max(0, current_saldo)

        if target_mins == 0: percentage = 100
        else: percentage = min(100, int((progress_saldo / target_mins) * 100))

        self.goal_progress_bar.setValue(percentage)
        self.goal_progress_bar.setFormat(f"{percentage}% erreicht")
        self.lbl_goal_current.setText(self.format_time(current_saldo))

        missing_mins = target_mins - current_saldo

        if missing_mins <= 0:
            self.lbl_goal_missing.setText("0h 0m")
            self.lbl_goal_days.setText("-")
            self.lbl_goal_action.setText("🎉 Herzlichen Glückwunsch! Du hast genug Überstunden für diesen Zeitraum angespart!")
            self.lbl_goal_action.setStyleSheet("color: #10b981; font-weight: bold;")
            return

        self.lbl_goal_missing.setText(self.format_time(missing_mins))

        today = QDate.currentDate()
        if target_start_date <= today:
            self.lbl_goal_days.setText("0")
            self.lbl_goal_action.setText("⚠️ Der gewünschte Zeitraum hat bereits begonnen oder ist heute!")
            self.lbl_goal_action.setStyleSheet("color: #ef4444; font-weight: bold;")
            return

        state = self.settings.get("state", "TH")
        workdays = 0
        curr = today.addDays(1)

        while curr < target_start_date:
            if self.get_target_minutes_for_date(curr.toString("yyyy-MM-dd")) > 0:
                workdays += 1
            curr = curr.addDays(1)

        self.lbl_goal_days.setText(str(workdays))

        if workdays == 0:
            self.lbl_goal_action.setText("Keine regulären Arbeitstage mehr zum Ansparen übrig!")
            self.lbl_goal_action.setStyleSheet("color: #ef4444;")
        else:
            extra_per_day = missing_mins / workdays
            self.lbl_goal_action.setText(
                f"Tipp: Wenn du ab sofort jeden Tag <b>{int(extra_per_day)} Minuten</b> länger machst, erreichst du dein Ziel punktgenau."
            )
            self.lbl_goal_action.setStyleSheet("color: #3b82f6;")

    def update_calendar_heatmap(self):
        self.cal_month_filter.blockSignals(True)
        current_cal_filter = self.cal_month_filter.currentData()
        self.cal_month_filter.clear()

        months_set = set(e.date[:7] for e in self.entries if len(e.date) >= 7)
        today = QDate.currentDate()
        for i in range(-60, 61):
            months_set.add(today.addMonths(i).toString("yyyy-MM"))

        months = sorted(list(months_set), reverse=True)

        for m in months: self.cal_month_filter.addItem(f"{m[-2:]}/{m[:4]}", m)

        idx = self.cal_month_filter.findData(current_cal_filter)
        if idx < 0: idx = self.cal_month_filter.findData(today.toString("yyyy-MM"))
        if idx >= 0: self.cal_month_filter.setCurrentIndex(idx)

        self.btn_cal_prev.setEnabled(self.cal_month_filter.currentIndex() < self.cal_month_filter.count() - 1)
        self.btn_cal_next.setEnabled(self.cal_month_filter.currentIndex() > 0)

        self.cal_month_filter.blockSignals(False)

        sel_date_str = self.cal_month_filter.currentData()
        if not sel_date_str: sel_date_str = today.toString("yyyy-MM")

        year, month = map(int, sel_date_str.split('-'))
        cal = calendar.monthcalendar(year, month)

        state = self.settings.get("state", "TH")
        holidays = get_holidays(year, state)

        day_mins = {}
        monthly_sum = 0

        for e in self.entries:
            if e.date.startswith(sel_date_str):
                day_mins[e.date] = day_mins.get(e.date, 0) + e.minutes
                monthly_sum += e.minutes

        sign = "+" if monthly_sum > 0 else ""
        self.lbl_cal_month_sum.setText(f"Monats-Saldo: {self.format_time(monthly_sum, show_plus=True)}")
        if monthly_sum > 0:
            self.lbl_cal_month_sum.setStyleSheet("color: #10b981;")
        elif monthly_sum < 0:
            self.lbl_cal_month_sum.setStyleSheet("color: #ef4444;")
        else:
            self.lbl_cal_month_sum.setStyleSheet("")

        self.cal_table.setRowCount(len(cal))
        is_dark = self.settings.get("dark_mode", False)
        workdays_setting = self.settings.get("workdays", [0, 1, 2, 3, 4])

        for row, week in enumerate(cal):
            for col, day in enumerate(week):
                if day == 0:
                    item = QTableWidgetItem("")
                    item.setData(Qt.ItemDataRole.UserRole + 1, False)
                    item.setBackground(QColor("#222222" if is_dark else "#f3f4f6"))
                else:
                    date_str = f"{year}-{month:02d}-{day:02d}"
                    mins = day_mins.get(date_str, 0)
                    is_holiday = date_str in holidays
                    is_workday = col in workdays_setting

                    f_mins = f"\n({self.format_time(mins, show_plus=True)})" if mins != 0 else ""
                    if is_holiday:
                        hol_name = holidays[date_str]
                        text = f"{day}\n{hol_name}{f_mins}"
                    else:
                        text = f"{day}{f_mins}"

                    item = QTableWidgetItem(text)
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

                    if is_holiday:
                        item.setToolTip(holidays[date_str])
                        if mins == 0:
                            item.setBackground(QColor("#1e3a8a" if is_dark else "#bfdbfe"))
                            item.setForeground(QColor("#60a5fa" if is_dark else "#1d4ed8"))
                        elif mins > 0:
                            alpha = min(255, 60 + (mins * 2))
                            item.setBackground(QColor(16, 185, 129, alpha))
                        else:
                            alpha = min(255, 60 + (abs(mins) * 2))
                            item.setBackground(QColor(239, 68, 68, alpha))
                    else:
                        if mins == 0:
                            if not is_workday:
                                item.setBackground(QColor("#2d3748" if is_dark else "#e5e7eb"))
                            else:
                                item.setBackground(QColor("#333333" if is_dark else "#ffffff"))
                        elif mins > 0:
                            alpha = min(255, 60 + (mins * 2))
                            item.setBackground(QColor(16, 185, 129, alpha))
                        else:
                            alpha = min(255, 60 + (abs(mins) * 2))
                            item.setBackground(QColor(239, 68, 68, alpha))

                    # Den aktuellen Tag (Heute) prüfen und markieren
                    if year == today.year() and month == today.month() and day == today.day():
                        item.setData(Qt.ItemDataRole.UserRole + 1, True)
                        hol_suffix = f"\n{holidays[date_str]}" if date_str in holidays else ""
                        if mins != 0:
                            item.setText(f"{day}\nHeute{hol_suffix}\n({'+' if mins > 0 else ''}{mins}m)")
                        else:
                            item.setText(f"{day}\nHeute{hol_suffix}")
                        item.setBackground(QColor("#1e3a8a"))
                        item.setForeground(QColor("#ffffff"))
                    else:
                        item.setData(Qt.ItemDataRole.UserRole + 1, False)

                self.cal_table.setItem(row, col, item)

    # --- ALLGEMEINE LOGIK ---
    def on_list_filter_changed(self):
        filter_val = self.month_filter.currentData()
        if filter_val and filter_val != "ALL":
            idx = self.cal_month_filter.findData(filter_val)
            if idx >= 0:
                self.cal_month_filter.blockSignals(True)
                self.cal_month_filter.setCurrentIndex(idx)
                self.cal_month_filter.blockSignals(False)
        self.update_ui()

    def on_cal_filter_changed(self):
        cal_val = self.cal_month_filter.currentData()
        if cal_val:
            idx = self.month_filter.findData(cal_val)
            if idx >= 0:
                self.month_filter.blockSignals(True)
                self.month_filter.setCurrentIndex(idx)
                self.month_filter.blockSignals(False)
        self.update_calendar_heatmap()

    def open_settings(self):
        dialog = SettingsDialog(self.settings, self)
        if dialog.exec():
            self.settings.update(dialog.get_settings())
            self.save_settings()
            self.apply_theme()
            self.pause_spin.setEnabled(not self.settings.get("auto_break", True))

            # Alle Tage neu berechnen, da sich das Soll geändert haben könnte
            all_dates = sorted(list(set(e.date for e in self.entries)))
            for d_str in all_dates:
                self.recalculate_day(d_str)

            self.load_data() # Lädt alles neu (inkl. UI-Update)

            new_start = self._get_default_start_time()
            self.time_start.blockSignals(True)
            self.time_start.setTime(new_start)
            self.time_start.blockSignals(False)
            self.on_start_time_changed(new_start)

    def _compute_target_end_time(self, new_start: "QTime") -> "QTime":
        """Berechnet die Endzeit so dass das Tagessoll (Regelarbeitszeit) erreicht wird.
        Berücksichtigt bereits gespeicherte Einträge (auch manuelle) für das Datum."""
        curr_date_str = self.date_edit.date().toString("yyyy-MM-dd")
        all_day = [e for e in self.entries if e.date == curr_date_str]
        timed_existing = [e for e in all_day if e.start and e.end]
        manual_sum = sum(e.minutes for e in all_day if not (e.start and e.end))

        target_mins = self.get_target_minutes_for_date(curr_date_str)
        max_mins = self.get_max_minutes()
        is_auto = self.settings.get("auto_break", True)

        # Wenn das Soll (inkl. manueller Einträge) bereits erreicht ist,
        # geben wir einfach die Startzeit zurück (0 Min. Dauer für den neuen Eintrag).
        # Die Live-Vorschau wird dann korrekt 0m oder den aktuellen Stand anzeigen.

        # Wir suchen die kleinste Dauer G (in Minuten), die das Soll erfüllt.
        for G in range(0, max_mins * 2 + 1):
            temp = WorkEntry(
                id=-1,
                date=curr_date_str,
                start=new_start.toString("HH:mm"),
                end=new_start.addSecs(G * 60).toString("HH:mm"),
                pause=self.pause_spin.value() if not is_auto else 0,
                minutes=0,
                reason=""
            )
            # calculate_timed_entries liefert uns das Netto der Zeiteinträge (timed_existing + temp)
            _, total_net_timed = calculate_timed_entries(timed_existing + [temp], target_mins, max_mins, is_auto)

            # Das gesamte Tagessaldo ist Netto-Zeit + manuelle Korrekturen
            if (total_net_timed + manual_sum) >= target_mins:
                break

        return new_start.addSecs(G * 60)

    def on_start_time_changed(self, new_start_time):
        today = QDate.currentDate().toString("yyyy-MM-dd")
        self.settings["last_date"] = today
        self.settings["last_start"] = new_start_time.toString("HH:mm")

        self.time_end.blockSignals(True)
        self.time_end.setTime(self._compute_target_end_time(new_start_time))
        self.time_end.blockSignals(False)
        self.update_live_calc()

    def update_live_calc(self):
        curr_date_str = self.date_edit.date().toString("yyyy-MM-dd")
        target_mins = self.get_target_minutes_for_date(curr_date_str)
        max_mins = self.get_max_minutes()
        is_auto = self.settings.get("auto_break", True)

        current_temp = WorkEntry(
            id=-1,
            date=curr_date_str,
            start=self.time_start.time().toString("HH:mm"),
            end=self.time_end.time().toString("HH:mm"),
            pause=self.pause_spin.value() if not is_auto else 0,
            minutes=0,
            reason=""
        )

        all_day = [e for e in self.entries if e.date == curr_date_str]
        timed = [e for e in all_day if e.start and e.end] + [current_temp]
        manual_sum = sum(e.minutes for e in all_day if not (e.start and e.end))

        results, total_net = calculate_timed_entries(timed, target_mins, max_mins, is_auto)
        entry_pause, entry_overtime = results[-1]

        if is_auto:
            self.pause_spin.blockSignals(True)
            self.pause_spin.setValue(entry_pause)
            self.pause_spin.blockSignals(False)

        self.current_calculated_pause = entry_pause
        self.current_calculated_overtime = entry_overtime

        final_total_overtime = (total_net - target_mins) + manual_sum

        calc_text = f"Netto (Tag): {self.format_time(total_net)} ➔ <b>{self.format_time(final_total_overtime, show_plus=True)} Überstunden (Tag-Saldo)</b>"
        warnings = []
        if total_net >= max_mins: warnings.append(f"⚠️ Max. {max_mins // 60}h erreicht!")

        # Ruhezeit-Check (Lücke zum Vortag/letzten Eintrag davor)
        prev_entry = self.db.get_last_entry_before(curr_date_str)
        if prev_entry and prev_entry.end:
            try:
                dt_prev = datetime.strptime(f"{prev_entry.date} {prev_entry.end}", "%Y-%m-%d %H:%M")
                dt_curr = datetime.strptime(f"{curr_date_str} {self.time_start.time().toString('HH:mm')}", "%Y-%m-%d %H:%M")
                rest_hours = (dt_curr - dt_prev).total_seconds() / 3600
                if 0 < rest_hours < 11:
                    warnings.append(f"⚠️ Ruhezeit verletzt ({rest_hours:.1f}h < 11h)")
            except ValueError: pass

        if warnings:
            calc_text += f" <span style='color: #ef4444;'>{' | '.join(warnings)}</span>"
            self.lbl_live_calc.setStyleSheet("color: #ef4444;")
        else:
            self.lbl_live_calc.setStyleSheet("")
        self.lbl_live_calc.setText(calc_text)

    def recalculate_day(self, date_str):
        """Verteilt Pausen und Überstunden für alle Zeiteinträge eines Tages neu.
        Manuelle Einträge (ohne Start-/Endzeit) werden nicht angefasst."""
        day_entries = [e for e in self.entries if e.date == date_str]
        if not day_entries: return

        timed = [e for e in day_entries if e.start and e.end]
        if not timed: return

        target_mins = self.get_target_minutes_for_date(date_str)
        max_mins = self.get_max_minutes()
        is_auto = self.settings.get("auto_break", True)

        results, _ = calculate_timed_entries(timed, target_mins, max_mins, is_auto)

        for e in timed:
            e.pause, e.minutes = results[e.id]
            self.db.update(e)

    def set_now_as_end(self):
        """Setzt die Endzeit auf die aktuelle Uhrzeit."""
        self.time_end.setTime(QTime.currentTime())

    def check_overlap(self, date_str, start_str, end_str, exclude_id=None):
        """Prüft, ob sich der Zeitraum mit bestehenden Einträgen am selben Tag überschneidet."""
        if not start_str or not end_str:
            return None

        s_new = QTime.fromString(start_str, "HH:mm")
        e_new = QTime.fromString(end_str, "HH:mm")

        # Falls Mitternacht überschritten wird, ist die Logik komplexer,
        # hier vereinfacht für den gleichen Tag:
        for e in self.entries:
            if e.date == date_str and e.start and e.end and e.id != exclude_id:
                s_old = QTime.fromString(e.start, "HH:mm")
                e_old = QTime.fromString(e.end, "HH:mm")

                # Standard Überlappungs-Check: (StartA < EndeB) und (EndeA > StartB)
                # Wir nutzen secsTo für den Vergleich
                if s_new.secsTo(e_old) > 0 and s_old.secsTo(e_new) > 0:
                    return f"{e.start} - {e.end} ({e.reason or 'Ohne Anlass'})"
        return None

    def add_entry(self):
        date_str = self.date_edit.date().toString("yyyy-MM-dd")
        start_str = self.time_start.time().toString("HH:mm")
        end_str = self.time_end.time().toString("HH:mm")

        # Überlappungs-Check
        overlap = self.check_overlap(date_str, start_str, end_str)
        if overlap:
            QMessageBox.warning(self, "Überschneidung",
                f"Dieser Zeitraum überschneidet sich mit einem existierenden Eintrag:\n\n{overlap}\n\nBitte korrigiere die Zeiten.")
            return

        entry = WorkEntry(
            id=None,
            date=date_str,
            start=start_str,
            end=end_str,
            pause=self.current_calculated_pause,
            minutes=self.current_calculated_overtime,
            reason=self.reason_edit.text().strip(),
            target_minutes=(self.custom_target_time.time().hour() * 60 + self.custom_target_time.time().minute()) if self.custom_target_cb.isChecked() else -1
        )
        self.db.insert(entry)
        self.reason_edit.clear()
        self.custom_target_cb.setChecked(False)
        self.date_edit.setDate(QDate.currentDate())

        # Alle Einträge neu laden und den betroffenen Tag glattziehen
        self.entries = self.db.load_all()
        self.recalculate_day(date_str)
        self.load_data()

    def delete_entry(self, entry: WorkEntry):
        date_str = entry.date
        d = QDate.fromString(date_str, "yyyy-MM-dd").toString("dd.MM.yyyy")
        reply = QMessageBox.question(self, "Löschen bestätigen",
            f"Eintrag vom {d} wirklich löschen?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            self.db.delete(entry.id)
            # Nach Löschen: Tag neu berechnen
            self.entries = [e for e in self.entries if e.id != entry.id]
            self.recalculate_day(date_str)
            self.load_data()

    def edit_entry(self, row, column):
        entry_idx = self.table.item(row, 0).data(Qt.ItemDataRole.UserRole)
        entry = self.entries[entry_idx]
        old_date = entry.date

        dialog = EditDialog(entry, self.entries, self.get_target_minutes(), self.get_max_minutes(), self.settings.get("auto_break", True), self)
        if dialog.exec():
            # Bevor wir anwenden: Überlappungs-Check (mit den neuen Werten aus dem Dialog)
            # Wir holen uns die Werte temporär
            new_date = dialog.date_edit.date().toString("yyyy-MM-dd")
            new_start = dialog.time_start.time().toString("HH:mm") if dialog.has_times_cb.isChecked() else ""
            new_end = dialog.time_end.time().toString("HH:mm") if dialog.has_times_cb.isChecked() else ""

            overlap = self.check_overlap(new_date, new_start, new_end, exclude_id=entry.id)
            if overlap:
                QMessageBox.warning(self, "Überschneidung",
                    f"Die Änderungen überschneiden sich mit einem anderen Eintrag:\n\n{overlap}\n\nBitte korrigiere die Zeiten.")
                return

            dialog.apply_to_entry()
            self.db.update(entry)

            # Neu laden und beide betroffenen Tage (alt/neu) neu berechnen
            self.entries = self.db.load_all()
            self.recalculate_day(old_date)
            if entry.date != old_date:
                self.recalculate_day(entry.date)
            self.load_data()

    def format_time(self, total_minutes, show_plus=False):
        """Formatiert Minuten in Stunden und Minuten.
        Unter 60 Min -> '45m'
        Ab 60 Min    -> '1h 5m'
        """
        sign = "+" if show_plus and total_minutes > 0 else ("-" if total_minutes < 0 else "")
        abs_m = abs(total_minutes)
        if abs_m < 60:
            return f"{sign}{abs_m}m"
        return f"{sign}{abs_m // 60}h {abs_m % 60}m"

    def update_ui(self):
        self.month_filter.blockSignals(True)
        current_filter = self.month_filter.currentData()
        self.month_filter.clear()
        self.month_filter.addItem("Alle", "ALL")

        months = sorted(list(set(e.date[:7] for e in self.entries if len(e.date) >= 7)), reverse=True)
        for m in months: self.month_filter.addItem(f"{m[-2:]}/{m[:4]}", m)
        idx = self.month_filter.findData(current_filter)
        if idx >= 0: self.month_filter.setCurrentIndex(idx)
        self.month_filter.blockSignals(False)

        self.table.setRowCount(0)
        filter_val = self.month_filter.currentData()
        total_overall = sum(e.minutes for e in self.entries)

        row = 0
        for i, e in enumerate(self.entries):
            if filter_val != "ALL" and not e.date.startswith(filter_val): continue

            self.table.insertRow(row)
            item_date = QTableWidgetItem(QDate.fromString(e.date, "yyyy-MM-dd").toString("dd.MM.yyyy"))
            item_date.setData(Qt.ItemDataRole.UserRole, i)
            item_date.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

            z_str = f"{e.start} - {e.end}" + (f" (-{e.pause}m)" if e.pause > 0 else "") if e.start else "-"
            item_zeit = QTableWidgetItem(z_str)
            item_zeit.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

            item_min = QTableWidgetItem(self.format_time(e.minutes, show_plus=True))
            item_min.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            if e.minutes > 0: item_min.setForeground(QColor("#10b981"))
            elif e.minutes < 0: item_min.setForeground(QColor("#ef4444"))

            btn_del = QPushButton("Löschen")
            btn_del.clicked.connect(lambda checked, ent=e: self.delete_entry(ent))

            self.table.setItem(row, 0, item_date)
            self.table.setItem(row, 1, item_zeit)
            self.table.setItem(row, 2, item_min)
            self.table.setItem(row, 3, QTableWidgetItem(e.reason))
            self.table.setCellWidget(row, 4, btn_del)
            row += 1

        self.lbl_saldo.setText(self.format_time(total_overall))
        self.lbl_saldo.setStyleSheet("color: #10b981;" if total_overall > 0 else ("color: #ef4444;" if total_overall < 0 else ""))

    def update_stats_chart(self):
        monthly_totals = {}
        for e in reversed(self.entries):
            if len(e.date) >= 7:
                m = e.date[:7]
                monthly_totals[m] = monthly_totals.get(m, 0) + e.minutes

        self.figure.clear()
        is_dark = self.settings.get("dark_mode", False)
        bg_color = '#31363b' if is_dark else '#ffffff'
        text_color = '#e0e0e0' if is_dark else '#000000'

        self.figure.patch.set_facecolor(bg_color)
        ax = self.figure.add_subplot(111)
        ax.set_facecolor(bg_color)
        ax.tick_params(colors=text_color)
        for spine in ax.spines.values(): spine.set_edgecolor(text_color)

        if not monthly_totals:
            ax.text(0.5, 0.5, "Keine Daten vorhanden", color=text_color, ha='center', va='center')
        else:
            months = list(monthly_totals.keys())
            values = [v / 60 for v in monthly_totals.values()]

            colors = ['#10b981' if v >= 0 else '#ef4444' for v in values]
            ax.bar(months, values, color=colors)
            ax.set_ylabel("Überstunden (in Stunden)", color=text_color)
            ax.set_title("Monatlicher Überstunden-Verlauf", color=text_color)
            self.figure.autofmt_xdate()

        self.canvas.draw()

    # --- EXPORT HILFSMETHODEN ---
    def _get_export_entries(self):
        filter_val = self.month_filter.currentData()
        return [e for e in self.entries if filter_val == "ALL" or e.date.startswith(filter_val)]

    def _get_export_title(self):
        filter_val = self.month_filter.currentData()
        return "Alle Einträge" if filter_val == "ALL" else f"Monat {filter_val}"

    def _export_row_data(self, e):
        """Gibt (datum_str, zeitraum_str) für einen Eintrag zurück."""
        d = QDate.fromString(e.date, "yyyy-MM-dd").toString("dd.MM.yyyy") if e.date else ""
        if e.start and e.end:
            pause_str = f" (-{e.pause}m)" if e.pause > 0 else ""
            zeitraum = f"{e.start} – {e.end}{pause_str}"
        else:
            zeitraum = "–"
        return d, zeitraum

    # --- EXPORT ---
    def export_csv(self):
        file_name, _ = QFileDialog.getSaveFileName(
            self, "CSV Export", "ueberstunden_export.csv", "CSV Dateien (*.csv)")
        if not file_name:
            return
        try:
            export_entries = self._get_export_entries()
            with open(file_name, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f, delimiter=";")
                writer.writerow(["Datum", "Zeitraum", "Minuten", "Dauer", "Anlass"])
                for e in export_entries:
                    d, zeitraum = self._export_row_data(e)
                    writer.writerow([d, zeitraum, e.minutes, self.format_time(e.minutes), e.reason])
                total = sum(e.minutes for e in export_entries)
                writer.writerow(["Gesamt", "", "", self.format_time(total), ""])
            QMessageBox.information(self, "Erfolg", "CSV erfolgreich exportiert!")
        except Exception as ex:
            QMessageBox.critical(self, "Fehler", f"Fehler beim CSV-Export:\n{str(ex)}")

    def export_xlsx(self):
        if not _OPENPYXL:
            QMessageBox.critical(self, "Fehler",
                "openpyxl ist nicht installiert.\nBitte ausführen: pip install openpyxl")
            return

        file_name, _ = QFileDialog.getSaveFileName(
            self, "Excel Export", "ueberstunden_export.xlsx", "Excel Dateien (*.xlsx)")
        if not file_name:
            return
        try:
            export_entries = self._get_export_entries()
            total_min = sum(e.minutes for e in export_entries)

            wb = Workbook()
            ws = wb.active
            ws.title = "Überstunden"

            # Titel
            ws.merge_cells("A1:E1")
            ws["A1"] = f"Überstunden-Nachweis – {self._get_export_title()}"
            ws["A1"].font = Font(bold=True, size=13)
            ws["A1"].alignment = Alignment(horizontal="center")

            # Kopfzeile
            hdr_fill = PatternFill("solid", fgColor="3b82f6")
            hdr_font = Font(bold=True, color="FFFFFF")
            hdr_align = Alignment(horizontal="center")
            for col, text in enumerate(["Datum", "Zeitraum", "Minuten", "Dauer", "Anlass"], 1):
                c = ws.cell(row=3, column=col, value=text)
                c.font = hdr_font
                c.fill = hdr_fill
                c.alignment = hdr_align

            # Datenzeilen
            alt_fill = PatternFill("solid", fgColor="f3f4f6")
            for i, e in enumerate(export_entries):
                row = i + 4
                d, zeitraum = self._export_row_data(e)
                values = [d, zeitraum, e.minutes, self.format_time(e.minutes), e.reason]
                for col, val in enumerate(values, 1):
                    c = ws.cell(row=row, column=col, value=val)
                    if i % 2 == 1:
                        c.fill = alt_fill
                ovt_cell = ws.cell(row=row, column=3)
                if e.minutes > 0:
                    ovt_cell.font = Font(color="059669")
                elif e.minutes < 0:
                    ovt_cell.font = Font(color="dc2626")

            # Summenzeile
            sum_row = len(export_entries) + 4
            sum_fill = PatternFill("solid", fgColor="dbeafe")
            ws.merge_cells(f"A{sum_row}:C{sum_row}")
            ws[f"A{sum_row}"] = "Gesamtsumme:"
            ws[f"D{sum_row}"] = self.format_time(total_min)
            for col in range(1, 6):
                c = ws.cell(row=sum_row, column=col)
                c.font = Font(bold=True)
                c.fill = sum_fill

            # Spaltenbreiten
            for col, width in zip("ABCDE", [14, 24, 18, 12, 32]):
                ws.column_dimensions[col].width = width

            wb.save(file_name)
            QMessageBox.information(self, "Erfolg", "Excel-Datei erfolgreich exportiert!")
        except Exception as ex:
            QMessageBox.critical(self, "Fehler", f"Fehler beim Excel-Export:\n{str(ex)}")

    def export_pdf(self):
        from PyQt6.QtPrintSupport import QPrinter
        from PyQt6.QtGui import QTextDocument, QPageSize
        from PyQt6.QtCore import QSizeF

        file_name, _ = QFileDialog.getSaveFileName(
            self, "PDF Export", "ueberstunden_export.pdf", "PDF Dateien (*.pdf)")
        if not file_name:
            return
        try:
            export_entries = self._get_export_entries()
            total_min = sum(e.minutes for e in export_entries)

            rows_html = ""
            for i, e in enumerate(export_entries):
                d, zeitraum = self._export_row_data(e)
                bg = "#f9fafb" if i % 2 == 1 else "#ffffff"
                if e.minutes > 0:
                    color = "#059669"
                elif e.minutes < 0:
                    color = "#dc2626"
                else:
                    color = "inherit"
                rows_html += (
                    f"<tr style='background:{bg}'>"
                    f"<td>{d}</td><td>{zeitraum}</td>"
                    f"<td style='color:{color};text-align:right'>{self.format_time(e.minutes, show_plus=True)}</td>"
                    f"<td>{e.reason}</td></tr>"
                )

            html = f"""
            <html><head><meta charset="utf-8"><style>
                body {{ font-family: Arial, sans-serif; font-size: 11px; color: #111; }}
                h2 {{ font-size: 14px; margin-bottom: 4px; }}
                p.sub {{ color: #666; font-size: 10px; margin-top: 0; }}
                table {{ width: 100%; border-collapse: collapse; margin-top: 12px; }}
                th {{ background: #3b82f6; color: #fff; padding: 5px 8px; text-align: left; }}
                td {{ border-bottom: 1px solid #e5e7eb; padding: 4px 8px; }}
                tr.sum td {{ background: #dbeafe; font-weight: bold; }}
            </style></head><body>
            <h2>Überstunden-Nachweis</h2>
            <p class="sub">{self._get_export_title()} &nbsp;·&nbsp;
               Erstellt am {QDate.currentDate().toString("dd.MM.yyyy")}</p>
            <table>
              <tr><th>Datum</th><th>Zeitraum</th><th>Überstunden</th><th>Anlass</th></tr>
              {rows_html}
              <tr class="sum">
                <td colspan="2">Gesamtsumme:</td>
                <td style='text-align:right'>{self.format_time(total_min, show_plus=True)}</td><td></td>
              </tr>
            </table>
            </body></html>"""

            printer = QPrinter(QPrinter.PrinterMode.HighResolution)
            printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
            printer.setOutputFileName(file_name)

            doc = QTextDocument()
            doc.setHtml(html)
            doc.setPageSize(QSizeF(printer.pageRect(QPrinter.Unit.Point).size()))
            doc.print(printer)

            QMessageBox.information(self, "Erfolg", "PDF erfolgreich exportiert!")
        except Exception as ex:
            QMessageBox.critical(self, "Fehler", f"Fehler beim PDF-Export:\n{str(ex)}")

    def import_csv(self):
        file_name, _ = QFileDialog.getOpenFileName(self, "CSV Import", "", "CSV Dateien (*.csv)")
        if not file_name:
            return
        try:
            with open(file_name, newline="", encoding="utf-8", errors="replace") as csvfile:
                content = csvfile.read()
                if not content:
                    return
                delimiter = ";" if ";" in content.split("\n")[0] else ","
                csvfile.seek(0)
                reader = csv.reader(csvfile, delimiter=delimiter)

                pending = []
                affected_dates = set()
                for row in reader:
                    if len(row) >= 1:
                        date_str = row[0].strip()
                        if date_str.lower() in ["datum", "date"] or not date_str:
                            continue

                        # Minuten optional behandeln
                        try:
                            minutes = int(row[1].strip()) if len(row) > 1 and row[1].strip() else 0
                        except ValueError:
                            minutes = 0

                        reason = row[2].strip() if len(row) > 2 else ""
                        start_str = row[3].strip() if len(row) > 3 else ""
                        end_str = row[4].strip() if len(row) > 4 else ""
                        try:
                            pause = int(row[5].strip()) if len(row) > 5 else 0
                        except ValueError:
                            pause = 0

                        parsed_date = ""
                        for fmt in ("%d.%m.%y", "%d.%m.%Y", "%Y-%m-%d"):
                            try:
                                parsed_date = datetime.strptime(date_str, fmt).strftime("%Y-%m-%d")
                                break
                            except ValueError:
                                pass

                        if not parsed_date:
                            parsed_date = date_str # Fallback

                        pending.append(WorkEntry(id=None, date=parsed_date, start=start_str, end=end_str, pause=pause, minutes=minutes, reason=reason))
                        affected_dates.add(parsed_date)

            if not pending:
                QMessageBox.information(self, "Import", "Keine importierbaren Einträge gefunden.")
                return

            preview_lines = [f"  {e.date}  {e.start}-{e.end}  {e.reason}" for e in pending[:5]]
            if len(pending) > 5:
                preview_lines.append(f"  … und {len(pending) - 5} weitere")
            preview = "\n".join(preview_lines)

            reply = QMessageBox.question(self, "Import bestätigen",
                f"{len(pending)} Einträge gefunden:\n\n{preview}\n\nJetzt importieren?\n(Überstunden werden automatisch für jeden Tag berechnet/konsolidiert)",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            if reply != QMessageBox.StandardButton.Yes:
                return

            if os.path.exists(DB_FILE):
                shutil.copy2(DB_FILE, DB_FILE + ".backup")

            for entry in pending:
                self.db.insert(entry)

            # Alle Einträge neu laden und alle betroffenen Tage glattziehen
            self.entries = self.db.load_all()
            for d in sorted(list(affected_dates)):
                self.recalculate_day(d)

            self.load_data()
            QMessageBox.information(self, "Erfolg", f"{len(pending)} Einträge importiert!\nTagessalden wurden automatisch berechnet.\nBackup der Datenbank angelegt.")

        except Exception as ex:
            QMessageBox.critical(self, "Fehler", f"Fehler beim Import:\n{str(ex)}")


if __name__ == "__main__":
    QApplication.setHighDpiScaleFactorRoundingPolicy(
        Qt.HighDpiScaleFactorRoundingPolicy.PassThrough
    )
    app = QApplication(sys.argv)
    window = UeberstundenApp()
    window.show()
    sys.exit(app.exec())
