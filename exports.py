
"""
Hilfsfunktionen für den Export von Arbeitseinträgen (CSV, Excel, PDF).
"""
import csv
import html
import logging
from PyQt6.QtCore import QDate, QLocale, QSizeF
from PyQt6.QtGui import QTextDocument
from PyQt6.QtPrintSupport import QPrinter
from PyQt6.QtWidgets import QFileDialog, QMessageBox

from i18n import get_locale, tr
from logic import format_time, fmt_date, fmt_time_hhmm, get_target_minutes_for_date, \
    TYPE_LABELS, TYPE_WORK, ABSENCE_TYPES

logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    _OPENPYXL = True
except ImportError:
    _OPENPYXL = False


def _get_export_row_data(e):
    """Gibt (datum_str, zeitraum_str) für einen Eintrag zurück."""
    d = fmt_date(e.date) if e.date else ""
    if e.start and e.end:
        pause_str = f" (-{e.pause}m)" if e.pause > 0 else ""
        zeitraum = f"{fmt_time_hhmm(e.start)} – {fmt_time_hhmm(e.end)}{pause_str}"
    else:
        zeitraum = "–"
    return d, zeitraum


def export_csv(parent, entries):
    """Exportiert Einträge als CSV-Datei."""
    file_name, _ = QFileDialog.getSaveFileName(
        parent, tr("CSV Export"), "ueberstunden_export.csv", tr("CSV Dateien (*.csv)"))
    if not file_name:
        return
    try:
        with open(file_name, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f, delimiter=";")
            writer.writerow([
                tr("Datum"), tr("Zeitraum"), "Minuten", tr("Dauer"), tr("Anlass")
            ])
            for e in entries:
                d, zeitraum = _get_export_row_data(e)
                writer.writerow(
                    [d, zeitraum, e.minutes, format_time(e.minutes), e.reason]
                )
            total = sum(e.minutes for e in entries)
            writer.writerow([tr("Gesamt"), "", "", format_time(total), ""])
        QMessageBox.information(parent, tr("Erfolg"), tr("CSV erfolgreich exportiert!"))
    except Exception as ex:  # pylint: disable=broad-except
        QMessageBox.critical(parent, tr("Fehler"),
                             tr("Fehler beim CSV-Export:\n{ex}").format(ex=str(ex)))


def _setup_xlsx_header(ws, title):
    """Hilfsfunktion zum Erstellen des Excel-Headers."""
    ws.merge_cells("A1:E1")
    ws["A1"] = tr("Überstunden-Nachweis – {title}").format(title=title)
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center")

    hdr_fill = PatternFill("solid", fgColor="3b82f6")
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_align = Alignment(horizontal="center")
    headers = [tr("Datum"), tr("Zeitraum"), "Minuten", tr("Dauer"), tr("Anlass")]
    for col, text in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=text)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = hdr_align


def _write_xlsx_entries(ws, entries):
    """Hilfsfunktion zum Schreiben der Einträge in Excel."""
    alt_fill = PatternFill("solid", fgColor="f3f4f6")
    for i, e in enumerate(entries):
        r = i + 4
        d, zeitraum = _get_export_row_data(e)
        values = [d, zeitraum, e.minutes, format_time(e.minutes), e.reason]
        for col, val in enumerate(values, 1):
            c = ws.cell(row=r, column=col, value=val)
            if i % 2 == 1:
                c.fill = alt_fill
        ovt_cell = ws.cell(row=r, column=3)
        if e.minutes > 0:
            ovt_cell.font = Font(color="059669")
        elif e.minutes < 0:
            ovt_cell.font = Font(color="dc2626")


def export_xlsx(parent, entries, title):
    """Exportiert Einträge als Excel-Datei (xlsx)."""
    if not _OPENPYXL:
        msg = tr("openpyxl ist nicht installiert.\nBitte ausführen: pip install openpyxl")
        QMessageBox.critical(parent, tr("Fehler"), msg)
        return

    file_name, _ = QFileDialog.getSaveFileName(
        parent, tr("Excel Export"), "ueberstunden_export.xlsx", tr("Excel Dateien (*.xlsx)"))
    if not file_name:
        return
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Überstunden"
        _setup_xlsx_header(ws, title)
        _write_xlsx_entries(ws, entries)

        sum_row = len(entries) + 4
        sum_fill = PatternFill("solid", fgColor="dbeafe")
        ws.merge_cells(f"A{sum_row}:C{sum_row}")
        ws[f"A{sum_row}"] = tr("Gesamtsumme:")
        ws[f"D{sum_row}"] = format_time(sum(e.minutes for e in entries))
        for col in range(1, 6):
            c = ws.cell(row=sum_row, column=col)
            c.font = Font(bold=True)
            c.fill = sum_fill

        for col, width in zip("ABCDE", [14, 24, 18, 12, 32]):
            ws.column_dimensions[col].width = width

        wb.save(file_name)
        QMessageBox.information(parent, tr("Erfolg"), tr("Excel-Datei erfolgreich exportiert!"))
    except Exception as ex:  # pylint: disable=broad-except
        QMessageBox.critical(parent, tr("Fehler"),
                             tr("Fehler beim Excel-Export:\n{ex}").format(ex=str(ex)))


def _generate_pdf_html(entries, title):
    """Hilfsfunktion zum Erstellen des HTML-Inhalts für das PDF."""
    rows_html = ""
    for i, e in enumerate(entries):
        d, zeitraum = _get_export_row_data(e)
        bg = "#f9fafb" if i % 2 == 1 else "#ffffff"
        col_ov = "#059669" if e.minutes > 0 else ("#dc2626" if e.minutes < 0 else "inherit")
        rows_html += (
            f"<tr style='background:{bg}'>"
            f"<td>{html.escape(d)}</td><td>{html.escape(zeitraum)}</td>"
            f"<td style='color:{col_ov};text-align:right'>"
            f"{format_time(e.minutes, show_plus=True)}</td>"
            f"<td>{html.escape(e.reason or '')}</td></tr>"
        )

    return f"""
    <html><head><meta charset="utf-8"><style>
        body {{ font-family: Arial, sans-serif; font-size: 11px; color: #111; }}
        h2 {{ font-size: 14px; margin-bottom: 4px; }}
        p.sub {{ color: #666; font-size: 10px; margin-top: 0; }}
        table {{ width: 100%; border-collapse: collapse; margin-top: 12px; }}
        th {{ background: #3b82f6; color: #fff; padding: 5px 8px; text-align: left; }}
        td {{ border-bottom: 1px solid #e5e7eb; padding: 4px 8px; }}
        tr.sum td {{ background: #dbeafe; font-weight: bold; }}
    </style></head><body>
    <h2>{tr('Überstunden-Nachweis')}</h2>
    <p class="sub">{html.escape(title)} &nbsp;·&nbsp;
       {tr('Erstellt am {d}').format(
           d=get_locale().toString(QDate.currentDate(), QLocale.FormatType.ShortFormat)
       )}</p>
    <table>
      <tr><th>{tr('Datum')}</th><th>{tr('Zeitraum')}</th>
          <th>{tr('Überstunden')}</th><th>{tr('Anlass')}</th></tr>
      {rows_html}
      <tr class="sum">
        <td colspan="2">{tr('Gesamtsumme:')}</td>
        <td style='text-align:right'>
          {format_time(sum(e.minutes for e in entries), show_plus=True)}</td><td></td>
      </tr>
    </table>
    </body></html>"""


def export_pdf(parent, entries, title):
    """Exportiert Einträge als PDF-Datei."""
    file_name, _ = QFileDialog.getSaveFileName(
        parent, tr("PDF Export"), "ueberstunden_export.pdf", tr("PDF Dateien (*.pdf)"))
    if not file_name:
        return
    try:
        html = _generate_pdf_html(entries, title)
        printer = QPrinter(QPrinter.PrinterMode.HighResolution)
        printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
        printer.setOutputFileName(file_name)

        doc = QTextDocument()
        doc.setHtml(html)
        doc.setPageSize(QSizeF(printer.pageRect(QPrinter.Unit.Point).size()))
        doc.print(printer)

        QMessageBox.information(parent, tr("Erfolg"), tr("PDF erfolgreich exportiert!"))
    except Exception as ex:  # pylint: disable=broad-except
        QMessageBox.critical(parent, tr("Fehler"),
                             tr("Fehler beim PDF-Export:\n{ex}").format(ex=str(ex)))


def export_monthly_pdf(parent, entries, settings, month_str=None):
    """Monats-Stundennachweis als PDF (Tagestabelle mit Soll/lfd. Saldo)."""
    if month_str is None:
        today = QDate.currentDate()
        month_str = today.toString("yyyy-MM")
    year = int(month_str[:4])
    month = int(month_str[5:7])
    month_name = get_locale().monthName(month, QLocale.FormatType.LongFormat)
    title = tr("Stundennachweis {m} {y}").format(m=month_name, y=year)

    file_name, _ = QFileDialog.getSaveFileName(
        parent, tr("Monats-PDF exportieren"),
        f"stundennachweis_{month_str}.pdf", tr("PDF Dateien (*.pdf)"))
    if not file_name:
        return
    try:
        month_entries = [e for e in entries if e.date.startswith(month_str)]
        html = _generate_monthly_pdf_html(month_entries, settings, month_str, title)
        printer = QPrinter(QPrinter.PrinterMode.HighResolution)
        printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
        printer.setOutputFileName(file_name)
        doc = QTextDocument()
        doc.setHtml(html)
        doc.setPageSize(QSizeF(printer.pageRect(QPrinter.Unit.Point).size()))
        doc.print(printer)
        QMessageBox.information(parent, tr("Erfolg"),
                                tr("Monats-PDF erfolgreich exportiert!"))
    except Exception as ex:
        QMessageBox.critical(parent, tr("Fehler"),
                             tr("Fehler beim PDF-Export:\n{ex}").format(ex=str(ex)))


def _generate_monthly_pdf_html(entries, settings, month_str, title):
    """HTML für eine Monats-Tagestabelle."""
    year, month = int(month_str[:4]), int(month_str[5:7])
    locale = get_locale()
    date_entries = {}
    for e in entries:
        date_entries.setdefault(e.date, []).append(e)
    first = QDate(year, month, 1)
    last = QDate(year, month, first.daysInMonth())
    curr = first
    running = 0
    rows_html = ""
    while curr <= last:
        ds = curr.toString("yyyy-MM-dd")
        day_ents = date_entries.get(ds, [])
        dow = locale.dayName(curr.dayOfWeek(), QLocale.FormatType.ShortFormat)
        target = get_target_minutes_for_date(ds, day_ents, settings)
        if not day_ents:
            rows_html += (
                f"<tr><td>{html.escape(fmt_date(ds))}</td>"
                f"<td>{dow}</td><td>–</td>"
                f"<td style='text-align:right'>{format_time(target)}</td>"
                f"<td style='text-align:right'>0</td>"
                f"<td style='text-align:right'>{format_time(running, show_plus=True)}</td></tr>"
            )
        else:
            for e in day_ents:
                if e.entry_type in ABSENCE_TYPES:
                    z_str = TYPE_LABELS.get(e.entry_type, e.entry_type)
                elif e.start:
                    z_str = f"{fmt_time_hhmm(e.start)} - {fmt_time_hhmm(e.end)}"
                else:
                    z_str = "–"
                col_ov = "#059669" if e.minutes > 0 else ("#dc2626" if e.minutes < 0 else "inherit")
                running += e.minutes
                rows_html += (
                    f"<tr><td>{html.escape(fmt_date(ds))}</td>"
                    f"<td>{dow}</td><td>{html.escape(z_str)}</td>"
                    f"<td style='text-align:right'>{format_time(target)}</td>"
                    f"<td style='text-align:right;color:{col_ov}'>"
                    f"{format_time(e.minutes, show_plus=True)}</td>"
                    f"<td style='text-align:right'>{format_time(running, show_plus=True)}</td></tr>"
                )
        curr = curr.addDays(1)
    total_mins = sum(e.minutes for e in entries)
    return f"""<html><head><meta charset="utf-8"><style>
        body {{ font-family: Arial, sans-serif; font-size: 10px; color: #111; }}
        h2 {{ font-size: 14px; margin-bottom: 2px; }}
        p.sub {{ color: #666; font-size: 9px; margin-top: 0; margin-bottom: 8px; }}
        table {{ width: 100%; border-collapse: collapse; margin-top: 8px; }}
        th {{ background: #3b82f6; color: #fff; padding: 4px 6px; text-align: left; font-size: 9px; }}
        td {{ border-bottom: 1px solid #e5e7eb; padding: 3px 6px; }}
        tr.sum td {{ background: #dbeafe; font-weight: bold; }}
        tr.sign td {{ border: none; padding-top: 24px; font-size: 10px; }}
    </style></head><body>
    <h2>{html.escape(title)}</h2>
    <p class="sub">Monat {month:02d}/{year}
       &nbsp;·&nbsp; {tr('Saldo:')} {format_time(total_mins, show_plus=True)}</p>
    <table>
      <tr><th>{tr('Datum')}</th><th>{tr('Tag')}</th><th>{tr('Zeiten / Typ')}</th>
          <th style='text-align:right'>{tr('Soll')}</th>
          <th style='text-align:right'>{tr('Saldo')}</th>
          <th style='text-align:right'>{tr('Lfd.')}</th></tr>
      {rows_html}
      <tr class="sum">
        <td colspan="4">{tr('Monatssumme')}</td>
        <td style='text-align:right'>{format_time(total_mins, show_plus=True)}</td>
        <td></td></tr>
      <tr class="sign">
        <td colspan="2">{tr('Datum:')} ___________________</td>
        <td colspan="2">{tr('Arbeitnehmer:')} ___________________</td>
        <td colspan="2">{tr('Arbeitgeber:')} ___________________</td></tr>
    </table>
    </body></html>"""
