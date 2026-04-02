"""
Eigenständiges Widget für den Haupt-Tab (Eingabe & Liste).
"""
import csv
import logging
import os
import shutil
from datetime import datetime
from io import StringIO

from PyQt6.QtCore import QDate, QLocale, Qt, QTime, pyqtSignal
from PyQt6.QtGui import QColor, QFont
from PyQt6.QtWidgets import (
    QCheckBox, QComboBox, QDateEdit,
    QFileDialog, QFrame, QHBoxLayout, QHeaderView, QLabel,
    QLineEdit, QMenu, QMessageBox, QPushButton, QSpinBox,
    QTableWidget, QTableWidgetItem, QTimeEdit,
    QVBoxLayout, QWidget
)

from config import DB_FILE
from dialogs import EditDialog
from i18n import get_locale, tr
from logic import (
    calculate_timed_entries, get_login_time,
    format_time, fmt_date, fmt_time_hhmm,
    get_target_minutes, get_max_minutes, get_target_minutes_for_date,
    COLOR_POSITIVE, COLOR_NEGATIVE,
)
from models import WorkEntry
from ui_components import set_overtime_color

logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    _OPENPYXL = True
except ImportError:
    _OPENPYXL = False


# pylint: disable=too-many-instance-attributes
class MainTab(QWidget):  # pylint: disable=too-many-public-methods
    """Haupt-Tab für Zeiterfassung, Eintrags-Liste und Export/Import."""

    data_changed = pyqtSignal()
    filter_changed = pyqtSignal(str)

    # pylint: disable=too-many-arguments, too-many-positional-arguments
    def __init__(self, db, settings, save_settings_cb, open_settings_cb, parent=None):
        """
        Initialisiert das Haupt-Tab-Widget.

        Args:
            db:               DBManager-Instanz für Datenbankzugriffe.
            settings:         Einstellungs-Dictionary (gemeinsame Referenz).
            save_settings_cb: Callable zum Speichern der Einstellungen.
            open_settings_cb: Callable zum Öffnen des Einstellungs-Dialogs.
            parent:           Eltern-Widget.
        """
        super().__init__(parent)
        self.db = db
        self.settings = settings
        self._save_settings = save_settings_cb
        self._open_settings_cb = open_settings_cb
        self.entries = []
        self.current_calculated_overtime = 0
        self.current_calculated_pause = 0
        self._last_added_params = None

        self._build_ui()

    # pylint: disable=too-many-locals, too-many-statements
    def _build_ui(self):
        """Erstellt das Layout des Haupt-Tabs."""
        layout = QVBoxLayout(self)

        self.lbl_saldo = QLabel("0h 0m")
        self.lbl_saldo.setAlignment(Qt.AlignmentFlag.AlignCenter)
        font = QFont()
        font.setPointSize(28)
        font.setBold(True)
        self.lbl_saldo.setFont(font)
        layout.addWidget(QLabel(f"<b>{tr('Gesamt-Saldo:')}</b>",
                                alignment=Qt.AlignmentFlag.AlignCenter))
        layout.addWidget(self.lbl_saldo)

        frame_input = QFrame()
        frame_layout = QVBoxLayout(frame_input)

        input_row1 = QHBoxLayout()
        self.date_edit = QDateEdit()
        self.date_edit.setCalendarPopup(True)
        self.date_edit.setDisplayFormat(
            get_locale().dateFormat(QLocale.FormatType.ShortFormat)
        )
        self.date_edit.setDate(QDate.currentDate())
        self.date_edit.dateChanged.connect(self.update_live_calc)
        input_row1.addWidget(QLabel(tr("Datum:")))
        input_row1.addWidget(self.date_edit)

        start_to_use = self._get_default_start_time()

        _time_fmt = get_locale().timeFormat(QLocale.FormatType.ShortFormat)
        self.time_start = QTimeEdit()
        self.time_start.setDisplayFormat(_time_fmt)
        self.time_start.setTime(start_to_use)
        input_row1.addWidget(QLabel(tr("Start:")))
        input_row1.addWidget(self.time_start)
        btn_now_start = QPushButton(tr("Jetzt"))
        btn_now_start.setFixedWidth(50)
        btn_now_start.setToolTip(tr("Aktuelle Uhrzeit als Startzeit setzen"))
        btn_now_start.clicked.connect(lambda: self.time_start.setTime(QTime.currentTime()))
        input_row1.addWidget(btn_now_start)

        self.time_end = QTimeEdit()
        self.time_end.setDisplayFormat(_time_fmt)
        input_row1.addWidget(QLabel(tr("Ende:")))
        input_row1.addWidget(self.time_end)
        btn_now_end = QPushButton(tr("Jetzt"))
        btn_now_end.setFixedWidth(50)
        btn_now_end.setToolTip(tr("Aktuelle Uhrzeit als Endzeit setzen"))
        btn_now_end.clicked.connect(self._set_now_as_end)
        input_row1.addWidget(btn_now_end)

        self.pause_spin = QSpinBox()
        self.pause_spin.setRange(0, 300)
        self.pause_spin.setSuffix(" Min")
        self.pause_spin.setEnabled(not self.settings.get("auto_break", True))
        input_row1.addWidget(QLabel(tr("Pause:")))
        input_row1.addWidget(self.pause_spin)

        self.reason_edit = QLineEdit()
        self.reason_edit.setPlaceholderText(tr("z.B. Regulär"))
        input_row1.addWidget(QLabel(tr("Anlass:")))
        input_row1.addWidget(self.reason_edit)

        btn_add = QPushButton(tr("Eintragen"))
        btn_add.clicked.connect(self.add_entry)
        input_row1.addWidget(btn_add)

        frame_layout.addLayout(input_row1)

        input_row2 = QHBoxLayout()
        self.custom_target_cb = QCheckBox(tr("Indiv. Tagessoll:"))
        self.custom_target_time = QTimeEdit()
        self.custom_target_time.setDisplayFormat("HH:mm")
        def_target = self.settings.get("target_work_time", "08:00")
        self.custom_target_time.setTime(QTime.fromString(def_target, "HH:mm"))
        self.custom_target_time.setEnabled(False)
        self.custom_target_cb.stateChanged.connect(
            lambda: self.custom_target_time.setEnabled(self.custom_target_cb.isChecked())
        )
        self.custom_target_cb.stateChanged.connect(self.update_live_calc)
        self.custom_target_time.timeChanged.connect(self.update_live_calc)

        input_row2.addWidget(self.custom_target_cb)
        input_row2.addWidget(self.custom_target_time)
        input_row2.addStretch()
        frame_layout.addLayout(input_row2)

        self.lbl_live_calc = QLabel(tr("Berechne..."))
        self.lbl_live_calc.setAlignment(Qt.AlignmentFlag.AlignCenter)
        frame_layout.addWidget(self.lbl_live_calc)
        layout.addWidget(frame_input)

        self.time_start.timeChanged.connect(self.on_start_time_changed)
        self.time_end.timeChanged.connect(self.update_live_calc)
        self.pause_spin.valueChanged.connect(self.update_live_calc)

        toolbar_layout = QHBoxLayout()
        self.month_filter = QComboBox()
        self.month_filter.addItem(tr("Alle"), "ALL")
        self.month_filter.currentIndexChanged.connect(self._on_list_filter_changed)
        toolbar_layout.addWidget(QLabel(tr("Filter:")))
        toolbar_layout.addWidget(self.month_filter)
        toolbar_layout.addStretch()

        btn_import = QPushButton(tr("CSV Import"))
        btn_import.clicked.connect(self.import_csv)
        toolbar_layout.addWidget(btn_import)

        btn_export = QPushButton(tr("Export"))
        export_menu = QMenu(self)
        export_menu.addAction(tr("CSV  (.csv)"),  self.export_csv)
        export_menu.addAction(tr("Excel (.xlsx)"), self.export_xlsx)
        export_menu.addAction(tr("PDF  (.pdf)"),  self.export_pdf)
        btn_export.setMenu(export_menu)
        toolbar_layout.addWidget(btn_export)

        btn_settings = QPushButton(tr("Einstellungen"))
        btn_settings.clicked.connect(self._open_settings_cb)
        toolbar_layout.addWidget(btn_settings)
        layout.addLayout(toolbar_layout)

        self.table = QTableWidget(0, 5)
        self.table.setHorizontalHeaderLabels([
            tr("Datum"), tr("Zeitraum"), tr("Überstunden"), tr("Anlass"), tr("Aktion")
        ])
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        self.table.setColumnWidth(1, 160)
        self.table.cellDoubleClicked.connect(self.edit_entry)
        layout.addWidget(self.table)

        self.on_start_time_changed(start_to_use)

    # --- Public interface ---

    def refresh(self, entries):
        """Aktualisiert die Einträge und die Tabellen-Anzeige.

        Args:
            entries: Aktuelle Liste aller WorkEntry-Objekte.
        """
        self.entries = entries
        self.update_ui()
        self.update_live_calc()

    def set_filter(self, month_str):
        """Setzt den Monatsfilter ohne das filter_changed-Signal auszulösen.

        Args:
            month_str: Monatsstring im Format 'yyyy-MM' oder 'ALL'.
        """
        idx = self.month_filter.findData(month_str)
        if idx >= 0:
            self.month_filter.blockSignals(True)
            self.month_filter.setCurrentIndex(idx)
            self.month_filter.blockSignals(False)
            self.update_ui()

    def set_db(self, db):
        """Ersetzt die Datenbankverbindung (z.B. nach DB-Wechsel in Einstellungen).

        Args:
            db: Neue DBManager-Instanz.
        """
        self.db = db

    def on_settings_changed(self):
        """Aktualisiert UI-Elemente nach einer Einstellungsänderung."""
        self.pause_spin.setEnabled(not self.settings.get("auto_break", True))
        new_start = self._get_default_start_time()
        self.time_start.blockSignals(True)
        self.time_start.setTime(new_start)
        self.time_start.blockSignals(False)
        self.on_start_time_changed(new_start)

    def recalculate_day(self, date_str):
        """Verteilt Pausen und Überstunden für alle Zeiteinträge eines Tages neu.

        Manuelle Einträge (ohne Start-/Endzeit) werden nicht angefasst.

        Args:
            date_str: Datum im Format 'yyyy-MM-dd'.
        """
        day_entries = [e for e in self.entries if e.date == date_str]
        if not day_entries:
            return
        timed = [e for e in day_entries if e.start and e.end]
        if not timed:
            return

        target_mins = get_target_minutes_for_date(date_str, self.entries, self.settings)
        max_mins = get_max_minutes(self.settings)
        is_auto = self.settings.get("auto_break", True)

        break_rules = self.settings.get("break_rules")
        results, _ = calculate_timed_entries(timed, target_mins, max_mins, is_auto, break_rules)
        for e in timed:
            e.pause, e.minutes = results[e.id]
            self.db.update(e)

    def recalculate_all_days(self):
        """Berechnet alle Tage in der Eintrags-Liste neu (z.B. nach Einstellungsänderung)."""
        all_dates = sorted(set(e.date for e in self.entries))
        for date_str in all_dates:
            self.recalculate_day(date_str)

    # --- UI update ---

    # pylint: disable=too-many-locals, too-many-statements, too-many-branches
    def update_ui(self):
        """Aktualisiert die Eintrags-Tabelle und den Gesamtsaldo entsprechend dem aktiven Filter."""
        self.month_filter.blockSignals(True)
        current_filter = self.month_filter.currentData()
        self.month_filter.clear()
        self.month_filter.addItem(tr("Alle"), "ALL")

        months = sorted(
            list(set(e.date[:7] for e in self.entries if len(e.date) >= 7)), reverse=True
        )
        for m in months:
            self.month_filter.addItem(f"{m[-2:]}/{m[:4]}", m)
        idx = self.month_filter.findData(current_filter)
        if idx >= 0:
            self.month_filter.setCurrentIndex(idx)
        self.month_filter.blockSignals(False)

        self.table.setRowCount(0)
        filter_val = self.month_filter.currentData()
        total_overall = sum(e.minutes for e in self.entries)

        row = 0
        for i, e in enumerate(self.entries):
            if filter_val != "ALL" and not e.date.startswith(filter_val):
                continue

            self.table.insertRow(row)
            item_date = QTableWidgetItem(fmt_date(e.date))
            item_date.setData(Qt.ItemDataRole.UserRole, i)
            item_date.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

            z_str = (
                f"{fmt_time_hhmm(e.start)} - {fmt_time_hhmm(e.end)}"
                + (f" (-{e.pause}m)" if e.pause > 0 else "")
                if e.start else "-"
            )
            item_zeit = QTableWidgetItem(z_str)
            item_zeit.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

            item_min = QTableWidgetItem(format_time(e.minutes, show_plus=True))
            item_min.setTextAlignment(
                Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter
            )
            if e.minutes > 0:
                item_min.setForeground(QColor(COLOR_POSITIVE))
            elif e.minutes < 0:
                item_min.setForeground(QColor(COLOR_NEGATIVE))

            btn_del = QPushButton(tr("Löschen"))
            btn_del.clicked.connect(lambda checked, ent=e: self.delete_entry(ent))

            self.table.setItem(row, 0, item_date)
            self.table.setItem(row, 1, item_zeit)
            self.table.setItem(row, 2, item_min)
            self.table.setItem(row, 3, QTableWidgetItem(e.reason))
            self.table.setCellWidget(row, 4, btn_del)
            row += 1

        self.lbl_saldo.setText(format_time(total_overall))
        set_overtime_color(self.lbl_saldo, total_overall)

    # --- Input helpers ---

    def _get_default_start_time(self):
        """Ermittelt die Startzeit für die Eingabemaske."""
        if self.settings.get("use_login_time", False):
            t = get_login_time()
            if t and t.isValid():
                return t
            return QTime.fromString(self.settings.get("default_start", "07:00"), "HH:mm")
        today = QDate.currentDate().toString("yyyy-MM-dd")
        if self.settings.get("last_date") == today and self.settings.get("last_start"):
            return QTime.fromString(self.settings["last_start"], "HH:mm")
        return QTime.fromString(self.settings.get("default_start", "07:00"), "HH:mm")

    def _compute_target_end_time(self, new_start):
        """Berechnet die Endzeit so, dass das Tagessoll erreicht wird."""
        curr_date_str = self.date_edit.date().toString("yyyy-MM-dd")
        all_day = [e for e in self.entries if e.date == curr_date_str]
        timed_existing = [e for e in all_day if e.start and e.end]
        manual_sum = sum(e.minutes for e in all_day if not (e.start and e.end))

        if self.custom_target_cb.isChecked():
            t_target = self.custom_target_time.time()
            target_mins = t_target.hour() * 60 + t_target.minute()
        else:
            target_mins = get_target_minutes_for_date(curr_date_str, self.entries, self.settings)

        max_mins = get_max_minutes(self.settings)
        is_auto = self.settings.get("auto_break", True)

        duration_mins = 0
        for duration_mins in range(0, max_mins * 2 + 1):
            temp = WorkEntry(
                id=-1,
                date=curr_date_str,
                start=new_start.toString("HH:mm"),
                end=new_start.addSecs(duration_mins * 60).toString("HH:mm"),
                pause=self.pause_spin.value() if not is_auto else 0,
                minutes=0,
                reason=""
            )
            _, total_net_timed = calculate_timed_entries(
                timed_existing + [temp], target_mins, max_mins, is_auto,
                self.settings.get("break_rules")
            )
            if (total_net_timed + manual_sum) >= target_mins:
                break

        return new_start.addSecs(duration_mins * 60)

    def _set_now_as_end(self):
        """Setzt die Endzeit auf die aktuelle Uhrzeit."""
        self.time_end.setTime(QTime.currentTime())

    def on_start_time_changed(self, new_start_time):
        """Wird aufgerufen, wenn die Startzeit geändert wird."""
        today = QDate.currentDate().toString("yyyy-MM-dd")
        self.settings["last_date"] = today
        self.settings["last_start"] = new_start_time.toString("HH:mm")

        self.time_end.blockSignals(True)
        self.time_end.setTime(self._compute_target_end_time(new_start_time))
        self.time_end.blockSignals(False)
        self.update_live_calc()

    # pylint: disable=too-many-locals, too-many-statements, too-many-branches
    def update_live_calc(self):
        """Berechnet die Überstunden-Vorschau live und zeigt sie im Label an."""
        curr_date_str = self.date_edit.date().toString("yyyy-MM-dd")

        if self.custom_target_cb.isChecked():
            t_target = self.custom_target_time.time()
            target_mins = t_target.hour() * 60 + t_target.minute()
        else:
            target_mins = get_target_minutes_for_date(curr_date_str, self.entries, self.settings)

        max_mins = get_max_minutes(self.settings)
        is_auto = self.settings.get("auto_break", True)

        current_temp = WorkEntry(
            id=-1,
            date=curr_date_str,
            start=self.time_start.time().toString("HH:mm"),
            end=self.time_end.time().toString("HH:mm"),
            pause=self.pause_spin.value() if not is_auto else 0,
            minutes=0,
            reason=""
        )

        all_day = [e for e in self.entries if e.date == curr_date_str]

        # Intelligenz: Wenn die aktuellen Formular-Zeiten exakt dem entsprechen,
        # was wir gerade gespeichert haben, zählen wir sie in der Vorschau nicht doppelt.
        is_duplicate = (
            hasattr(self, "_last_added_params") and
            self._last_added_params == (
                curr_date_str,
                current_temp.start,
                current_temp.end
            )
        )

        timed = [e for e in all_day if e.start and e.end]
        if not is_duplicate:
            timed.append(current_temp)

        manual_sum = sum(e.minutes for e in all_day if not (e.start and e.end))

        results, total_net = calculate_timed_entries(
            timed, target_mins, max_mins, is_auto, self.settings.get("break_rules")
        )

        # Wenn wir den temp-Eintrag übersprungen haben, nehmen wir die Werte
        # vom letzten echten Eintrag für die Anzeige/Speicherung
        if not is_duplicate:
            entry_pause, entry_overtime = results[-1]
        else:
            # Dummy-Werte oder wir lassen sie wie sie sind
            entry_pause = self.current_calculated_pause
            entry_overtime = self.current_calculated_overtime

        if is_auto and not is_duplicate:
            self.pause_spin.blockSignals(True)
            self.pause_spin.setValue(entry_pause)
            self.pause_spin.blockSignals(False)

        if not is_duplicate:
            self.current_calculated_pause = entry_pause
            self.current_calculated_overtime = entry_overtime

        final_total_overtime = (total_net - target_mins) + manual_sum
        calc_text = tr(
            "Netto (Tag): {net} ➔ <b>{ot} Überstunden (Tag-Saldo)</b>"
        ).format(
            net=format_time(total_net),
            ot=format_time(final_total_overtime, show_plus=True),
        )
        warnings = []
        if total_net >= max_mins:
            warnings.append(tr("⚠️ Max. {h}h erreicht!").format(h=max_mins // 60))

        prev_entry = self.db.get_last_entry_before(curr_date_str)
        if prev_entry and prev_entry.end:
            try:
                dt_prev = datetime.strptime(
                    f"{prev_entry.date} {prev_entry.end}", "%Y-%m-%d %H:%M"
                )
                start_str_curr = self.time_start.time().toString('HH:mm')
                dt_curr = datetime.strptime(
                    f"{curr_date_str} {start_str_curr}", "%Y-%m-%d %H:%M"
                )
                rest_hours = (dt_curr - dt_prev).total_seconds() / 3600
                if 0 < rest_hours < 11:
                    warnings.append(
                        tr("⚠️ Ruhezeit verletzt ({r}h < 11h)").format(r=f"{rest_hours:.1f}")
                    )
            except ValueError as exc:
                logger.debug("Ruhezeit konnte nicht berechnet werden (ungültige Zeitdaten): %s",
                             exc)

        if warnings:
            calc_text += f" <span style='color: {COLOR_NEGATIVE};'>{' | '.join(warnings)}</span>"
            self.lbl_live_calc.setStyleSheet(f"color: {COLOR_NEGATIVE};")
        else:
            self.lbl_live_calc.setStyleSheet("")
        self.lbl_live_calc.setText(calc_text)

    # --- Data modification ---

    def add_entry(self):
        """Liest die Eingabefelder aus, prüft auf Überlappung und fügt den Eintrag in die DB ein."""
        date_str = self.date_edit.date().toString("yyyy-MM-dd")
        start_str = self.time_start.time().toString("HH:mm")
        end_str = self.time_end.time().toString("HH:mm")

        overlap = self.check_overlap(date_str, start_str, end_str)
        if overlap:
            QMessageBox.warning(
                self, tr("Überschneidung"),
                tr(
                    "Dieser Zeitraum überschneidet sich mit einem existierenden Eintrag:"
                    "\n\n{overlap}\n\nBitte korrigiere die Zeiten."
                ).format(overlap=overlap)
            )
            return

        entry = WorkEntry(
            id=None,
            date=date_str,
            start=start_str,
            end=end_str,
            pause=self.current_calculated_pause,
            minutes=self.current_calculated_overtime,
            reason=self.reason_edit.text().strip(),
            target_minutes=(
                self.custom_target_time.time().hour() * 60
                + self.custom_target_time.time().minute()
            ) if self.custom_target_cb.isChecked() else -1
        )
        self.db.insert(entry)
        self.reason_edit.clear()
        self.custom_target_cb.setChecked(False)
        self.date_edit.setDate(QDate.currentDate())

        self.entries = self.db.load_all()

        # Wir speichern die Parameter des gerade hinzugefügten Eintrags.
        # update_live_calc nutzt dies, um eine Doppelzählung in der Vorschau zu vermeiden,
        # auch wenn Start- und Endzeit im Formular unverändert bleiben.
        self._last_added_params = (date_str, start_str, end_str)

        self.recalculate_day(date_str)
        self.data_changed.emit()

    # pylint: disable=too-many-locals, too-many-statements, too-many-branches
    def edit_entry(self, row, _column):
        """Öffnet den Bearbeitungs-Dialog für den Eintrag in der angeklickten Zeile."""
        entry_idx = self.table.item(row, 0).data(Qt.ItemDataRole.UserRole)
        entry = self.entries[entry_idx]
        old_date = entry.date

        dialog = EditDialog(
            entry, self.entries, get_target_minutes(self.settings),
            get_max_minutes(self.settings), self.settings.get("auto_break", True),
            self.settings.get("break_rules"), self
        )
        if dialog.exec():
            new_date = dialog.date_edit.date().toString("yyyy-MM-dd")
            new_start = (
                dialog.time_start.time().toString("HH:mm")
                if dialog.has_times_cb.isChecked() else ""
            )
            new_end = (
                dialog.time_end.time().toString("HH:mm")
                if dialog.has_times_cb.isChecked() else ""
            )

            overlap = self.check_overlap(new_date, new_start, new_end, exclude_id=entry.id)
            if overlap:
                QMessageBox.warning(
                    self, tr("Überschneidung"),
                    tr(
                        "Die Änderungen überschneiden sich mit einem anderen Eintrag:"
                        "\n\n{overlap}\n\nBitte korrigiere die Zeiten."
                    ).format(overlap=overlap)
                )
                return

            dialog.apply_to_entry()
            self.db.update(entry)

            self.entries = self.db.load_all()
            self.recalculate_day(old_date)
            if entry.date != old_date:
                self.recalculate_day(entry.date)
            self.data_changed.emit()

    def delete_entry(self, entry):
        """Fragt den Benutzer nach Bestätigung und löscht dann den übergebenen Eintrag."""
        date_str = entry.date
        d = fmt_date(date_str)
        reply = QMessageBox.question(
            self, tr("Löschen bestätigen"),
            tr("Eintrag vom {d} wirklich löschen?").format(d=d),
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply == QMessageBox.StandardButton.Yes:
            self.db.delete(entry.id)
            self.entries = [e for e in self.entries if e.id != entry.id]
            self.recalculate_day(date_str)
            self.data_changed.emit()

    def check_overlap(self, date_str, start_str, end_str, exclude_id=None):
        """Prüft, ob sich der Zeitraum mit bestehenden Einträgen am selben Tag überschneidet."""
        if not start_str or not end_str:
            return None

        s_new = QTime.fromString(start_str, "HH:mm")
        e_new = QTime.fromString(end_str, "HH:mm")

        if s_new >= e_new:
            return tr("Startzeit muss vor Endzeit liegen")

        for e in self.entries:
            if e.date == date_str and e.start and e.end and e.id != exclude_id:
                s_old = QTime.fromString(e.start, "HH:mm")
                e_old = QTime.fromString(e.end, "HH:mm")
                if s_new.secsTo(e_old) > 0 and s_old.secsTo(e_new) > 0:
                    return f"{e.start} - {e.end} ({e.reason or tr('Ohne Anlass')})"
        return None

    # --- Filter sync ---

    def _on_list_filter_changed(self):
        """Wird aufgerufen, wenn der Monatsfilter in der Listenansicht geändert wird."""
        filter_val = self.month_filter.currentData()
        if filter_val and filter_val != "ALL":
            self.filter_changed.emit(filter_val)
        self.update_ui()

    # --- Export / Import ---

    def _get_export_entries(self):
        """Gibt die Einträge gemäß aktivem Filter zurück."""
        filter_val = self.month_filter.currentData()
        return [e for e in self.entries if filter_val == "ALL" or e.date.startswith(filter_val)]

    def _get_export_title(self):
        """Gibt den Titel für Export-Dokumente zurück."""
        filter_val = self.month_filter.currentData()
        return tr("Alle Einträge") if filter_val == "ALL" else tr("Monat {m}").format(m=filter_val)

    def _export_row_data(self, e):
        """Gibt (datum_str, zeitraum_str) für einen Eintrag zurück."""
        d = fmt_date(e.date) if e.date else ""
        if e.start and e.end:
            pause_str = f" (-{e.pause}m)" if e.pause > 0 else ""
            zeitraum = f"{fmt_time_hhmm(e.start)} – {fmt_time_hhmm(e.end)}{pause_str}"
        else:
            zeitraum = "–"
        return d, zeitraum

    def export_csv(self):
        """Exportiert die aktuell gefilterten Einträge als CSV-Datei."""
        file_name, _ = QFileDialog.getSaveFileName(
            self, tr("CSV Export"), "ueberstunden_export.csv", tr("CSV Dateien (*.csv)"))
        if not file_name:
            return
        try:
            export_entries = self._get_export_entries()
            with open(file_name, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f, delimiter=";")
                writer.writerow([
                    tr("Datum"), tr("Zeitraum"), "Minuten", tr("Dauer"), tr("Anlass")
                ])
                for e in export_entries:
                    d, zeitraum = self._export_row_data(e)
                    writer.writerow(
                        [d, zeitraum, e.minutes, format_time(e.minutes), e.reason]
                    )
                total = sum(e.minutes for e in export_entries)
                writer.writerow([tr("Gesamt"), "", "", format_time(total), ""])
            QMessageBox.information(self, tr("Erfolg"), tr("CSV erfolgreich exportiert!"))
        except Exception as ex:  # pylint: disable=broad-except
            QMessageBox.critical(self, tr("Fehler"),
                                 tr("Fehler beim CSV-Export:\n{ex}").format(ex=str(ex)))

    # pylint: disable=too-many-locals, too-many-statements, too-many-branches
    def export_xlsx(self):
        """Exportiert die aktuell gefilterten Einträge als Excel-Datei (xlsx)."""
        if not _OPENPYXL:
            QMessageBox.critical(self, tr("Fehler"),
                tr("openpyxl ist nicht installiert.\nBitte ausführen: pip install openpyxl"))
            return

        file_name, _ = QFileDialog.getSaveFileName(
            self, tr("Excel Export"), "ueberstunden_export.xlsx", tr("Excel Dateien (*.xlsx)"))
        if not file_name:
            return
        try:
            export_entries = self._get_export_entries()
            total_min = sum(e.minutes for e in export_entries)

            wb = Workbook()
            ws = wb.active
            ws.title = "Überstunden"

            ws.merge_cells("A1:E1")
            ws["A1"] = tr("Überstunden-Nachweis – {title}").format(
                title=self._get_export_title()
            )
            ws["A1"].font = Font(bold=True, size=13)
            ws["A1"].alignment = Alignment(horizontal="center")

            hdr_fill = PatternFill("solid", fgColor="3b82f6")
            hdr_font = Font(bold=True, color="FFFFFF")
            hdr_align = Alignment(horizontal="center")
            for col, text in enumerate(
                [tr("Datum"), tr("Zeitraum"), "Minuten", tr("Dauer"), tr("Anlass")], 1
            ):
                c = ws.cell(row=3, column=col, value=text)
                c.font = hdr_font
                c.fill = hdr_fill
                c.alignment = hdr_align

            alt_fill = PatternFill("solid", fgColor="f3f4f6")
            for i, e in enumerate(export_entries):
                r = i + 4
                d, zeitraum = self._export_row_data(e)
                values = [d, zeitraum, e.minutes, format_time(e.minutes), e.reason]
                for col, val in enumerate(values, 1):
                    c = ws.cell(row=r, column=col, value=val)
                    if i % 2 == 1:
                        c.fill = alt_fill
                ovt_cell = ws.cell(row=r, column=3)
                if e.minutes > 0:
                    ovt_cell.font = Font(color="059669")
                elif e.minutes < 0:
                    ovt_cell.font = Font(color="dc2626")

            sum_row = len(export_entries) + 4
            sum_fill = PatternFill("solid", fgColor="dbeafe")
            ws.merge_cells(f"A{sum_row}:C{sum_row}")
            ws[f"A{sum_row}"] = tr("Gesamtsumme:")
            ws[f"D{sum_row}"] = format_time(total_min)
            for col in range(1, 6):
                c = ws.cell(row=sum_row, column=col)
                c.font = Font(bold=True)
                c.fill = sum_fill

            for col, width in zip("ABCDE", [14, 24, 18, 12, 32]):
                ws.column_dimensions[col].width = width

            wb.save(file_name)
            QMessageBox.information(self, tr("Erfolg"), tr("Excel-Datei erfolgreich exportiert!"))
        except Exception as ex:  # pylint: disable=broad-except
            QMessageBox.critical(self, tr("Fehler"),
                                 tr("Fehler beim Excel-Export:\n{ex}").format(ex=str(ex)))

    def export_pdf(self):
        """Exportiert die aktuell gefilterten Einträge als PDF-Datei."""
        # pylint: disable=import-outside-toplevel
        from PyQt6.QtPrintSupport import QPrinter
        from PyQt6.QtGui import QTextDocument
        from PyQt6.QtCore import QSizeF

        file_name, _ = QFileDialog.getSaveFileName(
            self, tr("PDF Export"), "ueberstunden_export.pdf", tr("PDF Dateien (*.pdf)"))
        if not file_name:
            return
        try:
            export_entries = self._get_export_entries()
            total_min = sum(e.minutes for e in export_entries)

            rows_html = ""
            for i, e in enumerate(export_entries):
                d, zeitraum = self._export_row_data(e)
                bg = "#f9fafb" if i % 2 == 1 else "#ffffff"
                if e.minutes > 0:
                    color = "#059669"
                elif e.minutes < 0:
                    color = "#dc2626"
                else:
                    color = "inherit"
                rows_html += (
                    f"<tr style='background:{bg}'>"
                    f"<td>{d}</td><td>{zeitraum}</td>"
                    f"<td style='color:{color};text-align:right'>"
                    f"{format_time(e.minutes, show_plus=True)}</td>"
                    f"<td>{e.reason}</td></tr>"
                )

            html = f"""
            <html><head><meta charset="utf-8"><style>
                body {{ font-family: Arial, sans-serif; font-size: 11px; color: #111; }}
                h2 {{ font-size: 14px; margin-bottom: 4px; }}
                p.sub {{ color: #666; font-size: 10px; margin-top: 0; }}
                table {{ width: 100%; border-collapse: collapse; margin-top: 12px; }}
                th {{ background: #3b82f6; color: #fff; padding: 5px 8px; text-align: left; }}
                td {{ border-bottom: 1px solid #e5e7eb; padding: 4px 8px; }}
                tr.sum td {{ background: #dbeafe; font-weight: bold; }}
            </style></head><body>
            <h2>{tr('Überstunden-Nachweis')}</h2>
            <p class="sub">{self._get_export_title()} &nbsp;·&nbsp;
               {tr('Erstellt am {d}').format(
                   d=get_locale().toString(QDate.currentDate(), QLocale.FormatType.ShortFormat)
               )}</p>
            <table>
              <tr><th>{tr('Datum')}</th><th>{tr('Zeitraum')}</th>
                  <th>{tr('Überstunden')}</th><th>{tr('Anlass')}</th></tr>
              {rows_html}
              <tr class="sum">
                <td colspan="2">{tr('Gesamtsumme:')}</td>
                <td style='text-align:right'>
                  {format_time(total_min, show_plus=True)}</td><td></td>
              </tr>
            </table>
            </body></html>"""

            printer = QPrinter(QPrinter.PrinterMode.HighResolution)
            printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
            printer.setOutputFileName(file_name)

            doc = QTextDocument()
            doc.setHtml(html)
            doc.setPageSize(QSizeF(printer.pageRect(QPrinter.Unit.Point).size()))
            doc.print(printer)

            QMessageBox.information(self, tr("Erfolg"), tr("PDF erfolgreich exportiert!"))
        except Exception as ex:  # pylint: disable=broad-except
            QMessageBox.critical(self, tr("Fehler"),
                                 tr("Fehler beim PDF-Export:\n{ex}").format(ex=str(ex)))

    # pylint: disable=too-many-locals, too-many-statements, too-many-branches, too-many-nested-blocks
    def import_csv(self):
        """Importiert Einträge aus einer CSV-Datei und fügt sie zur Datenbank hinzu."""
        file_name, _ = QFileDialog.getOpenFileName(
            self, tr("CSV Import"), "", tr("CSV Dateien (*.csv)")
        )
        if not file_name:
            return
        try:
            with open(file_name, newline="", encoding="utf-8", errors="replace") as csvfile:
                content = csvfile.read()
                if not content:
                    return
                delimiter = ";" if ";" in content.split("\n")[0] else ","
                reader = csv.reader(StringIO(content), delimiter=delimiter)

                pending = []
                affected_dates = set()
                for row in reader:
                    if len(row) >= 1:
                        date_str = row[0].strip()
                        if date_str.lower() in ["datum", "date"] or not date_str:
                            continue
                        try:
                            minutes = (
                                int(row[1].strip()) if len(row) > 1 and row[1].strip() else 0
                            )
                        except ValueError:
                            minutes = 0
                        reason = row[2].strip() if len(row) > 2 else ""
                        start_str = row[3].strip() if len(row) > 3 else ""
                        end_str = row[4].strip() if len(row) > 4 else ""
                        try:
                            pause = int(row[5].strip()) if len(row) > 5 else 0
                        except ValueError:
                            pause = 0

                        parsed_date = ""
                        for fmt in ("%d.%m.%y", "%d.%m.%Y", "%Y-%m-%d"):
                            try:
                                parsed_date = datetime.strptime(
                                    date_str, fmt
                                ).strftime("%Y-%m-%d")
                                break
                            except ValueError:
                                pass
                        if not parsed_date:
                            logger.warning(
                                "Datum '%s' konnte in keinem Format geparst werden, "
                                "wird unverändert übernommen", date_str
                            )
                            parsed_date = date_str

                        pending.append(WorkEntry(
                            id=None, date=parsed_date, start=start_str,
                            end=end_str, pause=pause, minutes=minutes, reason=reason
                        ))
                        affected_dates.add(parsed_date)

            if not pending:
                QMessageBox.information(self, tr("Import"),
                                        tr("Keine importierbaren Einträge gefunden."))
                return

            preview_lines = [f"  {e.date}  {e.start}-{e.end}  {e.reason}" for e in pending[:5]]
            if len(pending) > 5:
                preview_lines.append(
                    f"  {tr('… und {n} weitere').format(n=len(pending) - 5)}"
                )

            reply = QMessageBox.question(
                self, tr("Import bestätigen"),
                tr(
                    "{n} Einträge gefunden:\n\n{preview}"
                    "\n\nJetzt importieren?\n"
                    "(Überstunden werden automatisch für jeden Tag berechnet/konsolidiert)"
                ).format(n=len(pending), preview=chr(10).join(preview_lines)),
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if reply != QMessageBox.StandardButton.Yes:
                return

            current_db = self.settings.get("db_path", DB_FILE)
            if os.path.exists(current_db):
                shutil.copy2(current_db, current_db + ".backup")

            for entry in pending:
                self.db.insert(entry)

            self.entries = self.db.load_all()
            for d in sorted(list(affected_dates)):
                self.recalculate_day(d)

            self.data_changed.emit()
            QMessageBox.information(
                self, tr("Erfolg"),
                tr(
                    "{n} Einträge importiert!\n"
                    "Tagessalden wurden automatisch berechnet.\n"
                    "Backup der Datenbank angelegt."
                ).format(n=len(pending))
            )

        except Exception as ex:  # pylint: disable=broad-except
            QMessageBox.critical(self, tr("Fehler"),
                                 tr("Fehler beim Import:\n{ex}").format(ex=str(ex)))
